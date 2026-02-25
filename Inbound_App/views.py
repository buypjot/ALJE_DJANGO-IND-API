from asyncio.log import logger
import base64
import io
import logging
import math
import mimetypes
import random
from time import time
import uuid
from ALJE_PROJECT.settings import MINIO_INBOUND_CONTAINER_QC_BUCKET_NAME, MINIO_INBOUND_PO_BUCKET_NAME
from rest_framework.decorators import api_view
from rest_framework.views import APIView
from rest_framework.response import Response
from django.http import JsonResponse, HttpResponseNotAllowed, Http404, FileResponse
from django.views.decorators.csrf import csrf_exempt
import json
from django.db import connection, connections, transaction
from datetime import date, datetime
from rest_framework import status
import traceback
from django.views.generic import View 
from django.utils.decorators import method_decorator
import re
from rest_framework import  status
from rest_framework.pagination import PageNumberPagination

from .models import BayanDocument, DocNO_Models, InboundPoDocument,PICK_ID_Models, Uniq_GatePass_tbl
from django.middleware.csrf import get_token
from urllib.parse import unquote
from django.views.decorators.http import require_GET
from django.views.decorators.http import require_http_methods
from django.http import HttpResponse
import os
from django.utils.encoding import smart_str
from rest_framework.parsers import MultiPartParser, FormParser
from django.views.decorators.http import require_POST
from django.db.models import Max
from django.http import JsonResponse, HttpResponse, StreamingHttpResponse
import base64
import uuid
import os
from datetime import datetime
from io import BytesIO
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework import status
import logging
import boto3
from django.conf import settings
import json




def test_cors(request):
    # Accessing request method and relevant headers using request.META
    method = request.method
    origin = request.META.get('HTTP_ORIGIN', 'No Origin Header')

    return JsonResponse({
        'message': 'CORS is working!',
        'method': method,
        'origin': origin,
    })

class StandardResultsSetPagination(PageNumberPagination):
    """
    Custom pagination class to define default page size and maximum page size.
    """
    page_size = 20

    page_size_query_param = 'page_size'
    max_page_size = 100



class GeneratePick_IdView(APIView):
    def get(self, request, *args, **kwargs):
        token = get_token(request)

        now = datetime.now()
        year_short = str(now.year)[-2:]  # '25'
        month = f"{now.month:02d}"      # '06'
        prefix = f"PICK_{year_short}{month}"

        latest_pick = PICK_ID_Models.objects.filter(PICK_ID__startswith=prefix).order_by('-id').first()

        if latest_pick:
            last_pick_id = latest_pick.PICK_ID
            match = re.match(rf"{prefix}(\d+)", last_pick_id)
            if match:
                last_number = int(match.group(1))
                next_number = last_number + 1
            else:
                next_number = 1
        else:
            next_number = 1

        next_pick_id = f"{prefix}{next_number}"  # Removed zero-padding

        PICK_ID_Models.objects.create(PICK_ID=next_pick_id, TOKEN=token)

        return Response({
            "PICK_ID": next_pick_id,
            "TOKEN": token
        }, status=status.HTTP_200_OK)
   
class Pick_IdView(APIView):
    def get(self, request, *args, **kwargs):
        # Get current year and month
        now = datetime.now()
        year_short = str(now.year)[-2:]  # '25'
        month = f"{now.month:02d}"      # '05'
        prefix = f"PICK_{year_short}{month}"
 
        # Default PICK_ID if no records
        default_pick_id = f"{prefix}000"
 
        # Get latest entry from table
        with connections['Inbound_db'].cursor() as cursor:
            cursor.execute("""
                SELECT TOP 1 [id], [PICK_ID], [TOKEN]
                FROM   [WHR_INBOUND_PICK_ID]
                ORDER BY [id] DESC;
            """)
            result = cursor.fetchone()
 
        # Determine DOC_NO to return
        if result and result[1]:
            pick_id = result[1]
        else:
            pick_id = default_pick_id
 
        return Response({'PICK_ID': pick_id})

def get_pending_po(request):
    po_number = request.GET.get('po_number')  # GET param: ?po_number=8403

    if not po_number:
        return JsonResponse({'error': 'po_number is required'}, status=400)

    with connections['default'].cursor() as cursor:
        cursor.execute("""
            SELECT *
            FROM   [XXALJEBUYP_PENDING_PO]
            WHERE PO_NUMBER = %s
        """, [po_number])

        columns = [col[0] for col in cursor.description]
        data = [dict(zip(columns, row)) for row in cursor.fetchall()]

    return JsonResponse(data, safe=False)

@api_view(['GET'])
def get_supplier_details(request, vendor_number):
    try:
        with connections['default'].cursor() as cursor:
            cursor.execute("""
                SELECT 
                    VENDOR_NAME,
                    VENDOR_SITE_CODE,
                    PO_NUMBER
                FROM  XXALJEBUYP_PENDING_PO
                WHERE VENDOR_NUMBER = %s
            """, [vendor_number])

            rows = cursor.fetchall()

            if not rows:
                return Response({'status': 'not_found'})

            # Extract vendor_name and vendor_site from first row
            vendor_name = rows[0][0]
            vendor_site = rows[0][1]

            # Extract all PO numbers (unique only)
            po_numbers = list({row[2] for row in rows if row[2] is not None})

            return Response({
                'vendor_name': vendor_name,
                'vendor_site': vendor_site,
                'po_numbers': po_numbers,
                'status': 'success'
            })

    except Exception as e:
        return Response({'status': 'error', 'message': str(e)})




@api_view(['POST'])
def save_warehouse_location(request):
    try:
        location_name = request.data.get('location_name')
        status_val = request.data.get('status')
        
        if not location_name or not status_val:
            return Response({'error': 'Missing required fields'}, status=400)
        
        with connections['Inbound_db'].cursor() as cursor:
            cursor.execute(
                "INSERT INTO   [WHR_LOCATION] ([LOCATION_NAME], [STATUS]) VALUES (%s, %s)",
                [location_name, status_val]
            )
        return Response({'success': True}, status=201)
    except Exception as e:
        return Response({'error': str(e)}, status=500)

@api_view(['GET'])
def get_warehouse_locations(request):
    try:
        with connections['Inbound_db'].cursor() as cursor:
            cursor.execute("SELECT [LOCATION_NAME], [STATUS] FROM   [WHR_LOCATION]")
            columns = [col[0] for col in cursor.description]
            locations = [dict(zip(columns, row)) for row in cursor.fetchall()]
        return Response(locations)
    except Exception as e:
        return Response({'error': str(e)}, status=500)


# views.py
@api_view(['GET'])
def get_location_options(request):
    try:
        with connections['Inbound_db'].cursor() as cursor:
            # Get distinct racks
            cursor.execute("SELECT DISTINCT [LOCATION_NAME] FROM   [WHR_LOCATION] WHERE [STATUS] = 'Rack'")
            racks = [row[0] for row in cursor.fetchall()]
            
            # Get distinct bins
            cursor.execute("SELECT DISTINCT [LOCATION_NAME] FROM   [WHR_LOCATION] WHERE [STATUS] = 'Bin'")
            bins = [row[0] for row in cursor.fetchall()]
            
            # Get distinct shelves
            cursor.execute("SELECT DISTINCT [LOCATION_NAME] FROM   [WHR_LOCATION] WHERE [STATUS] = 'Shelf'")
            shelves = [row[0] for row in cursor.fetchall()]

        return Response({
            'racks': racks,
            'bins': bins,
            'shelves': shelves,
        })
    except Exception as e:
        return Response({'error': str(e)}, status=500)

@api_view(['DELETE'])
def delete_warehouse_location(request):
    try:
        location_name = request.data.get('location_name')
        status_val = request.data.get('status')
        
        if not location_name or not status_val:
            return Response({'error': 'Missing required fields'}, status=400)
        
        with connections['Inbound_db'].cursor() as cursor:
            cursor.execute(
                "DELETE FROM   [WHR_LOCATION] WHERE [LOCATION_NAME] = %s AND [STATUS] = %s",
                [location_name, status_val]
            )
            if cursor.rowcount == 0:
                return Response({'error': 'Location not found'}, status=404)
        
        return Response({'success': True})
    except Exception as e:
        return Response({'error': str(e)}, status=500)
@csrf_exempt
def generate_document_no(request):
    """
    Generates a unique document number (DCYYMMSS) using MSSQL Sequence and AppLock.
    Handles concurrency and monthly reset safely.
    """
    try:
        doc_no = generate_next_doc_no()
        return JsonResponse({'status': 'success', 'DOC_NO': doc_no})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


def generate_next_doc_no():
    """
    Helper function to generate the next unique document number via Raw SQL.
    Returns the string DOC_NO or raises Exception.
    """
    now = datetime.now()
    year_yy = now.strftime("%y")
    month_mm = now.strftime("%m")
    prefix = f"DC{year_yy}{month_mm}" # e.g., 'DC2512'

    sql_batch = """
    SET NOCOUNT ON;
    DECLARE @Prefix NVARCHAR(20) = %s;
    DECLARE @NewDocNo NVARCHAR(50);
    DECLARE @Inserted BIT = 0;
    DECLARE @RetryCount INT = 0;
    
    -- Loop for Retry Logic (Safety net)
    WHILE @Inserted = 0 AND @RetryCount < 10
    BEGIN
        SET @RetryCount = @RetryCount + 1;
        BEGIN TRY
            BEGIN TRANSACTION;
            
            -- Acquire Exclusive Application Lock for 'DocNumberGen'
            -- This specifically prevents race conditions during the critical 'Resetcheck' phase
            -- Timeout 2 seconds -> if locked, waiting is better than failing immediately
            DECLARE @LockResult INT;
            EXEC @LockResult = sp_getapplock @Resource = 'DocNumberGen', @LockMode = 'Exclusive', @LockTimeout = 2000;
            
            IF @LockResult < 0
            BEGIN
                -- Failed to acquire lock, rollback and retry
                ROLLBACK TRANSACTION;
                CONTINUE;
            END
            
            -- CHECK FOR MONTHLY RESET
            -- If no record exists for this month (Prefix), restart sequence
            -- This check is safe because we hold the exclusive lock
            IF NOT EXISTS (SELECT 1 FROM [dbo].[DOCUMENTNUMBER_tbl] WHERE DOC_NO LIKE @Prefix + '%%')
            BEGIN
                ALTER SEQUENCE [dbo].[DOC_NO_SEQ] RESTART WITH 1;
            END
            
            -- Generate Next Number
            DECLARE @NextVal INT;
            SET @NextVal = NEXT VALUE FOR [dbo].[DOC_NO_SEQ];
            
            -- Format: DC + YY + MM + Sequence (01, 001, etc. - requirement says starts at 01)
            -- "01 -> Can grow to any length"
            -- Using '00' formatting for first 9 numbers (01..09), then normal
            
                SET @NewDocNo = @Prefix + 
                            CASE 
                                WHEN @NextVal < 10 THEN '0' + CAST(@NextVal AS NVARCHAR)
                                ELSE CAST(@NextVal AS NVARCHAR)
                            END;

            -- Insert validation (Double check, though Sequence should be unique unless reused)
            -- If we somehow generated a duplicate (unlikely with Sequence), constraints will catch it
            INSERT INTO [dbo].[DOCUMENTNUMBER_tbl] (DOC_NO)
            VALUES (@NewDocNo);
            
            SET @Inserted = 1;
            
            -- Release Lock
            EXEC sp_releaseapplock @Resource = 'DocNumberGen';
            
            COMMIT TRANSACTION;
        END TRY
        BEGIN CATCH
            -- Error Handling
            IF @@TRANCOUNT > 0 ROLLBACK TRANSACTION;
            -- If duplicate key or other transient error, loop will retry
            -- Wait briefly to reduce contention
            -- WAITFOR DELAY '00:00:00.005';
        END CATCH
    END
    
    SELECT @NewDocNo as DOC_NO;
    """

    with connections['Inbound_db'].cursor() as cursor:
        cursor.execute(sql_batch, [prefix])
        row = cursor.fetchone()
        
        if row and row[0]:
            return row[0]
        else:
            raise Exception("Failed to generate unique number after retries")






from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.utils import timezone
from django.db import connections, transaction
from datetime import datetime, date
import traceback

class InboundBayanFullView(APIView):
    
    def post(self, request, *args, **kwargs):
        # Debug: Print complete request data
        print("\n" + "="*80)
        print("🔵 INCOMING REQUEST TO InboundBayanFullView")
        print("="*80)
        print(f"📋 Request data type: {type(request.data)}")
        print(f"📋 Full request data: {request.data}")
        
        try:
            data = request.data
            header_data = data.get("header")
            details_data = data.get("details")
            serial_numbers_data = data.get("serial_numbers", [])

            print(f"\n📊 EXTRACTED DATA:")
            print(f"  Header data: {bool(header_data)}")
            if header_data:
                print(f"  Header type: {type(header_data)}")
            print(f"  Details count: {len(details_data) if details_data else 0}")
            print(f"  Serial numbers count: {len(serial_numbers_data)}")

            if not header_data or not details_data:
                return Response(
                    {"status": "error", "message": "Header and Details are required"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # --- GENERATE UNIQUE DOC_NO ---
            generated_doc_no = generate_next_doc_no()
            print(f"\n🆔 Generated Unique DOC_NO: {generated_doc_no}")

            # Normalize header (convert to list if it's a dict)
            if isinstance(header_data, dict):
                header_data = [header_data]

            # ---------------- UTIL FUNCTIONS ----------------
            def parse_date(value, field_name="unknown"):
                """
                Parse date from various formats and return a date object or None
                Includes detailed debugging
                """
                # Debug what's coming in
                print(f"\n  📅 [parse_date for {field_name}] Input: '{value}' (type: {type(value).__name__})")
                
                # Handle None
                if value is None:
                    print(f"    → Value is None, returning None")
                    return None
                
                # Handle empty string or "null" string
                if isinstance(value, str):
                    value = value.strip()
                    if value == "" or value.lower() in ("null", "none"):
                        print(f"    → Empty/null string, returning None")
                        return None
                
                # If it's already a date/datetime object
                if isinstance(value, (datetime, date)):
                    if isinstance(value, datetime):
                        result = value.date()
                    else:
                        result = value
                    print(f"    → Already date object: {result}")
                    return result
                
                # Try to parse string dates
                if isinstance(value, str):
                    # Try different formats (prioritize your app's format)
                    formats = [
                        "%Y-%m-%d",           # 2024-02-24 (most likely your format)
                        "%d-%b-%Y",           # 24-Feb-2024
                        "%d/%m/%Y",           # 24/02/2024
                        "%m/%d/%Y",           # 02/24/2024
                        "%d-%m-%Y",           # 24-02-2024
                        "%Y/%m/%d",           # 2024/02/24
                        "%d %b %Y",           # 24 Feb 2024
                        "%Y%m%d",              # 20240224
                    ]
                    
                    for fmt in formats:
                        try:
                            parsed = datetime.strptime(value, fmt).date()
                            print(f"    ✅ Successfully parsed with format '{fmt}' → {parsed}")
                            return parsed
                        except ValueError:
                            continue
                    
                    print(f"    ❌ Could not parse date: '{value}' with any format")
                    return None
                
                # If it's another type (number, etc)
                print(f"    ⚠️ Unexpected type: {type(value).__name__}, returning None")
                return None

            def parse_int(value, field_name="unknown"):
                if value in (None, "", "null"):
                    print(f"  🔢 [parse_int for {field_name}] None/empty → None")
                    return None
                try:
                    result = int(float(value)) if isinstance(value, (int, float, str)) else int(value)
                    print(f"  🔢 [parse_int for {field_name}] '{value}' → {result}")
                    return result
                except (ValueError, TypeError):
                    print(f"  🔢 [parse_int for {field_name}] Failed to parse '{value}' → None")
                    return None

            def parse_float(value, field_name="unknown"):
                if value in (None, "", "null"):
                    print(f"  💯 [parse_float for {field_name}] None/empty → 0.0")
                    return 0.0
                try:
                    result = float(value)
                    print(f"  💯 [parse_float for {field_name}] '{value}' → {result}")
                    return result
                except (ValueError, TypeError):
                    print(f"  💯 [parse_float for {field_name}] Failed to parse '{value}' → 0.0")
                    return 0.0

            # ---------------- COMMON AUDIT DATA ----------------
            current_time = timezone.now()
            today = date.today()
            client_ip = request.META.get("REMOTE_ADDR", "")
            created_by = request.data.get("user_name", "") or "SYSTEM"

            print(f"\n👤 Audit Info:")
            print(f"  Current time: {current_time}")
            print(f"  Today: {today}")
            print(f"  Client IP: {client_ip}")
            print(f"  Created by: {created_by}")

            # ---------------- PROCESS HEADER DATA ----------------
            print("\n" + "-"*60)
            print("📋 PROCESSING HEADER DATA")
            print("-"*60)
            
            header_values = []
            for row_index, row in enumerate(header_data):
                print(f"\n  Row {row_index + 1}:")
                
                # Extract ALL fields with debug
                bayan_no = row.get("BAYAN_NO", "")
                master_pl = row.get("MASTER_PL", "")
                house_pl = row.get("HOUSE_PL", "")
                bill_no = row.get("BILL_NO", "")
                bill_due_date = parse_date(row.get("BILL_DUE_DATE"), "BILL_DUE_DATE")
                bayan_date = parse_date(row.get("BAYAN_DATE"), "BAYAN_DATE")
                bl_date = parse_date(row.get("BL_DATE"), "BL_DATE")
                etd_date = parse_date(row.get("ETD_DATE"), "ETD_DATE")
                eta_date = parse_date(row.get("ETA_DATE"), "ETA_DATE")
                container_count = parse_int(row.get("CONTAINER_COUNT"), "CONTAINER_COUNT")
                clearance_date = parse_date(row.get("CLEARANCE_DATE"), "CLEARANCE_DATE")
                supplier_no = row.get("SUPPLIER_NO", "")
                supplier_name = row.get("SUPPLIER_NAME", "")
                supplier_site = row.get("SUPPLIER_SITE", "")
                pol = row.get("POL", "")
                pod = row.get("POD", "")
                line_name = row.get("LINE_NAME", "")
                invoice_no1 = row.get("INVOICE_NO1", "")
                lc_no1 = row.get("LC_NO1", "")
                value1 = parse_float(row.get("VALUE1"), "VALUE1")
                payment_term1 = row.get("PAYMENT_TERM1", "")
                incoterms1 = row.get("INCOTERMS1", "")
                invoice_no2 = row.get("INVOICE_NO2", "")
                lc_no2 = row.get("LC_NO2", "")
                value2 = parse_float(row.get("VALUE2"), "VALUE2")
                payment_terms2 = row.get("PAYMENT_TERMS2", "")
                incoterms2 = row.get("INCOTERMS2", "")
                invoice_no3 = row.get("INVOICE_NO3", "")
                lc_no3 = row.get("LC_NO3", "")
                value3 = parse_float(row.get("VALUE3"), "VALUE3")
                payment_term3 = row.get("PAYMENT_TERM3", "")
                incoterms3 = row.get("INCOTERMS3", "")
                initiator_no = row.get("INITIATOR_NO", "")
                initiator_name = row.get("INITIATOR_NAME", "")
                
                # Print all extracted values for debugging
                print(f"\n    📊 Extracted values for row {row_index + 1}:")
                print(f"      BAYAN_NO: '{bayan_no}'")
                print(f"      MASTER_PL: '{master_pl}'")
                print(f"      HOUSE_PL: '{house_pl}'")
                print(f"      BILL_NO: '{bill_no}'")
                print(f"      BILL_DUE_DATE: {bill_due_date}")
                print(f"      BAYAN_DATE: {bayan_date}")
                print(f"      BL_DATE: {bl_date}")
                print(f"      ETD_DATE: {etd_date}")
                print(f"      ETA_DATE: {eta_date}")
                print(f"      CONTAINER_COUNT: {container_count}")
                print(f"      CLEARANCE_DATE: {clearance_date}")
                print(f"      SUPPLIER_NO: '{supplier_no}'")
                print(f"      SUPPLIER_NAME: '{supplier_name}'")
                print(f"      SUPPLIER_SITE: '{supplier_site}'")
                
                header_values.append([
                    generated_doc_no,  # DOC_NO
                    bayan_no,           # BAYAN_NO
                    master_pl,          # MASTER_PL
                    house_pl,           # HOUSE_PL
                    bill_no,            # BILL_NO
                    bill_due_date,      # BILL_DUE_DATE
                    bayan_date,         # BAYAN_DATE
                    bl_date,            # BL_DATE
                    etd_date,           # ETD_DATE
                    'In Transit',       # STATUS
                    eta_date,           # ETA_DATE
                    container_count,    # CONTAINER_COUNT
                    clearance_date,     # CLEARANCE_DATE
                    supplier_no,        # SUPPLIER_NO
                    supplier_name,      # SUPPLIER_NAME
                    supplier_site,      # SUPPLIER_SITE
                    pol,                # POL
                    pod,                # POD
                    line_name,          # LINE_NAME
                    invoice_no1,        # INVOICE_NO1
                    lc_no1,             # LC_NO1
                    value1,             # VALUE1
                    payment_term1,      # PAYMENT_TERM1
                    incoterms1,         # INCOTERMS1
                    invoice_no2,        # INVOICE_NO2
                    lc_no2,             # LC_NO2
                    value2,             # VALUE2
                    payment_terms2,     # PAYMENT_TERMS2
                    incoterms2,         # INCOTERMS2
                    invoice_no3,        # INVOICE_NO3
                    lc_no3,             # LC_NO3
                    value3,             # VALUE3
                    payment_term3,      # PAYMENT_TERM3
                    incoterms3,         # INCOTERMS3
                    initiator_no,       # INITIATOR_NO
                    initiator_name,     # INITIATOR_NAME
                    None,               # GRN
                    None,               # GRN_DATE
                    today,              # DATE
                    current_time,       # CREATION_DATE
                    created_by,         # CREATION_BY
                    client_ip,          # CREATION_IP
                    current_time,       # UPDATION_DATE
                    created_by,         # UPDATION_BY
                    client_ip,          # UPDATION_IP
                    "",                 # ATTRIBUTE1
                    "",                 # ATTRIBUTE2
                    "",                 # ATTRIBUTE3
                    "",                 # ATTRIBUTE4
                    "",                 # ATTRIBUTE5
                    "",                 # FLAG1
                    ""                  # FLAG2
                ])

            # ---------------- PROCESS DETAILS DATA ----------------
            print("\n" + "-"*60)
            print("📋 PROCESSING DETAILS DATA")
            print("-"*60)
            
            detail_values = []
            for det_index, row in enumerate(details_data):
                print(f"\n  Detail Row {det_index + 1}:")
                
                detail_values.append([
                    generated_doc_no,  # DOC_NO
                    row.get("BAYAN_NO", ""),
                    row.get("PO_NUMBER", ""),
                    row.get("FRANCHISE", "-"),
                    row.get("CLASS", "-"),
                    row.get("SUBCLASS", "-"),
                    row.get("ITEM_CODE", ""),
                    row.get("ITEM_DESC", ""),
                    parse_float(row.get("PO_QTY") or 0, f"PO_QTY_{det_index}"),
                    parse_float(row.get("REC_QTY") or 0, f"REC_QTY_{det_index}"),
                    parse_float(row.get("BALANCE_QTY") or 0, f"BALANCE_QTY_{det_index}"),
                    parse_float(row.get("SHIPPED_QTY") or 0, f"SHIPPED_QTY_{det_index}"),
                    row.get("CONTAINER_NO", ""),
                    parse_float(row.get("ASSIGNED_QTY") or 0, f"ASSIGNED_QTY_{det_index}"),
                    parse_float(row.get("SCANNED_QTY") or 0, f"SCANNED_QTY_{det_index}"),
                    row.get("TRUCK_NUMBER", ""),
                    parse_date(row.get("ARRIVAL_DATE"), f"ARRIVAL_DATE_{det_index}"),
                    parse_date(row.get("EXIT_DATE"), f"EXIT_DATE_{det_index}"),
                    row.get("DRIVER_INFO", ""),
                    row.get("INITIATOR_NO", ""),
                    row.get("INITIATOR_NAME", ""),
                    "",  # WHR_SUPERUSER_NO
                    "",  # WHR_SUPERUSER_NAME
                    row.get("STATUS", "H"),
                    today,  # DATE
                    row.get("SEAL_NO", ""),
                    row.get("HS_CODE", ""),
                    parse_float(row.get("QTY_1") or 0, f"QTY_1_{det_index}"),
                    row.get("MEASURE_1", ""),
                    parse_float(row.get("QTY_2") or 0, f"QTY_2_{det_index}"),
                    row.get("MEASURE_2", ""),
                    parse_float(row.get("MW") or row.get("NW") or 0.0, f"MW_{det_index}"),
                    parse_float(row.get("GW") or 0, f"GW_{det_index}"),
                    parse_float(row.get("KGS") or 0, f"KGS_{det_index}"),
                    current_time,  # CREATION_DATE
                    created_by,    # CREATION_BY
                    client_ip,     # CREATION_IP
                    current_time,  # UPDATION_DATE
                    created_by,    # UPDATION_BY
                    client_ip,     # UPDATION_IP
                    row.get("ATTRIBUTE1", ""),
                    row.get("ATTRIBUTE2", ""),
                    row.get("ATTRIBUTE3", ""),
                    row.get("ATTRIBUTE4", ""),
                    row.get("ATTRIBUTE5", ""),
                    row.get("FLAG1", ""),
                    row.get("FLAG2", "")
                ])

            # ---------------- PROCESS SERIAL NUMBERS ----------------
            print("\n" + "-"*60)
            print("📋 PROCESSING SERIAL NUMBERS")
            print("-"*60)
            
            serial_values = []
            if serial_numbers_data:
                for ser_index, serial_row in enumerate(serial_numbers_data):
                    print(f"\n  Serial Row {ser_index + 1}:")
                    
                    serial_values.append([
                        generated_doc_no,  # DOC_NO
                        serial_row.get("BAYAN_NO", ""),
                        serial_row.get("PO_NUMBER", ""),
                        serial_row.get("MODEL_NO", ""),
                        serial_row.get("SERIAL_NO", ""),
                        serial_row.get("INITIATOR_NO", ""),
                        serial_row.get("INITIATOR_NAME", ""),
                        parse_date(serial_row.get("DATE") or today, f"SERIAL_DATE_{ser_index}"),
                        current_time,  # CREATION_DATE
                        created_by,    # CREATED_BY
                        client_ip,     # CREATED_IP
                        current_time,  # UPDATED_DATE
                        created_by,    # UPDATED_BY
                        serial_row.get("ATTRIBUTE1", ""),
                        serial_row.get("ATTRIBUTE2", ""),
                        serial_row.get("ATTRIBUTE3", ""),
                        serial_row.get("ATTRIBUTE4", ""),
                        serial_row.get("FLAG1", ""),
                        serial_row.get("FLAG2", "")
                    ])

            # ---------------- DB INSERT ----------------
            print("\n" + "-"*60)
            print("💾 EXECUTING DATABASE INSERTS")
            print("-"*60)
            
            with transaction.atomic():
                with connections['Inbound_db'].cursor() as cursor:
                    
                    print(f"\n  Inserting {len(header_values)} header rows...")
                    
                    # HEADER INSERT
                    cursor.executemany("""
                        INSERT INTO [WHR_INBOUND_BAYAN_HEADER_TBL] (
                            DOC_NO, BAYAN_NO, MASTER_PL, HOUSE_PL, BILL_NO,
                            BILL_DUE_DATE, BAYAN_DATE, BL_DATE, ETD_DATE, STATUS,
                            ETA_DATE, CONTAINER_COUNT, CLEARANCE_DATE, SUPPLIER_NO,
                            SUPPLIER_NAME, SUPPLIER_SITE, POL, POD, LINE_NAME,
                            INVOICE_NO1, LC_NO1, VALUE1, PAYMENT_TERM1, INCOTERMS1,
                            INVOICE_NO2, LC_NO2, VALUE2, PAYMENT_TERMS2, INCOTERMS2,
                            INVOICE_NO3, LC_NO3, VALUE3, PAYMENT_TERM3, INCOTERMS3,
                            INITIATOR_NO, INITIATOR_NAME, GRN, GRN_DATE, DATE,
                            CREATION_DATE, CREATION_BY, CREATION_IP,
                            UPDATION_DATE, UPDATION_BY, UPDATION_IP,
                            ATTRIBUTE1, ATTRIBUTE2, ATTRIBUTE3, ATTRIBUTE4, ATTRIBUTE5,
                            FLAG1, FLAG2
                        )
                        VALUES (
                            %s, %s, %s, %s, %s,
                            %s, %s, %s, %s, %s,
                            %s, %s, %s, %s,
                            %s, %s, %s, %s, %s,
                            %s, %s, %s, %s, %s,
                            %s, %s, %s, %s, %s,
                            %s, %s, %s, %s, %s,
                            %s, %s, %s, %s, %s,
                            %s, %s, %s,
                            %s, %s, %s,
                            %s, %s, %s, %s, %s,
                            %s, %s
                        )
                    """, header_values)
                    
                    print(f"  ✅ Header inserted: {cursor.rowcount} rows")

                    print(f"\n  Inserting {len(detail_values)} detail rows...")
                    
                    # DETAILS INSERT
                    cursor.executemany("""
                        INSERT INTO [WHR_INBOUND_BAYAN_DETAILS_TBL] (
                            DOC_NO, BAYAN_NO, PO_NUMBER, FRANCHISE, CLASS, SUBCLASS,
                            ITEM_CODE, ITEM_DESC, PO_QTY, REC_QTY, BALANCE_QTY, SHIPPED_QTY,
                            CONTAINER_NO, ASSIGNED_QTY, SCANNED_QTY, TRUCK_NUMBER,
                            ARRIVAL_DATE, EXIT_DATE, DRIVER_INFO,
                            INITIATOR_NO, INITIATOR_NAME, WHR_SUPERUSER_NO, WHR_SUPERUSER_NAME, STATUS, DATE,
                            SEAL_NO, HS_CODE, QTY_1, MEASURE_1, QTY_2, MEASURE_2,
                            MW, GW, KGS,
                            CREATION_DATE, CREATION_BY, CREATION_IP,
                            UPDATION_DATE, UPDATION_BY, UPDATION_IP,
                            ATTRIBUTE1, ATTRIBUTE2, ATTRIBUTE3, ATTRIBUTE4, ATTRIBUTE5,
                            FLAG1, FLAG2
                        )
                        VALUES (
                            %s, %s, %s, %s, %s, %s,
                            %s, %s, %s, %s, %s, %s,
                            %s, %s, %s, %s,
                            %s, %s, %s,
                            %s, %s, %s, %s, %s, %s,
                            %s, %s, %s, %s, %s, %s,
                            %s, %s, %s,
                            %s, %s, %s,
                            %s, %s, %s,
                            %s, %s, %s, %s, %s,
                            %s, %s
                        )
                    """, detail_values)
                    
                    print(f"  ✅ Details inserted: {cursor.rowcount} rows")

                    if serial_values:
                        print(f"\n  Inserting {len(serial_values)} serial rows...")
                        
                        cursor.executemany("""
                            INSERT INTO [WHR_INBOUND_BAYAN_SERIALNO_TBL] (
                                DOC_NO, BAYAN_NO, PO_NUMBER, MODEL_NO, SERIAL_NO,
                                INITIATOR_NO, INITIATOR_NAME, DATE,
                                CREATION_DATE, CREATED_BY, CREATED_IP,
                                UPDATED_DATE, UPDATED_BY, 
                                ATTRIBUTE1, ATTRIBUTE2, ATTRIBUTE3, ATTRIBUTE4,
                                FLAG1, FLAG2
                            )
                            VALUES (
                                %s, %s, %s, %s, %s,
                                %s, %s, %s,
                                %s, %s, %s,
                                %s, %s,
                                %s, %s, %s, %s,
                                %s, %s
                            )
                        """, serial_values)
                        
                        print(f"  ✅ Serials inserted: {cursor.rowcount} rows")

            print("\n" + "="*60)
            print("✅ ALL DATA SAVED SUCCESSFULLY!")
            print("="*60)
            print(f"📄 Document No: {generated_doc_no}")
            print(f"📊 Summary:")
            print(f"  - Header rows: {len(header_values)}")
            print(f"  - Detail rows: {len(detail_values)}")
            print(f"  - Serial rows: {len(serial_values)}")
            print("="*60)

            return Response(
                {
                    "status": "success",
                    "message": "Inbound bayan saved successfully",
                    "header_rows": len(header_values),
                    "detail_rows": len(detail_values),
                    "serial_rows": len(serial_values),
                    "doc_no": generated_doc_no,
                    "bayan_no": header_data[0].get("BAYAN_NO") if header_data else None,
                },
                status=status.HTTP_201_CREATED
            )

        except Exception as e:
            print("\n" + "="*60)
            print("❌ ERROR IN InboundBayanFullView")
            print("="*60)
            print(f"Error type: {type(e).__name__}")
            print(f"Error message: {str(e)}")
            print("\n📋 Traceback:")
            traceback.print_exc()
            
            # Debug: Print what we were trying to insert
            if 'header_values' in locals():
                print(f"\n🔍 Header values count: {len(header_values)}")
                if header_values:
                    print(f"🔍 Header values sample (first row):")
                    for i, val in enumerate(header_values[0][:20]):  # Print first 20 fields
                        print(f"  Field {i}: {val} (type: {type(val).__name__})")
            
            return Response(
                {
                    "status": "error", 
                    "message": str(e),
                    "details": traceback.format_exc()
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


            
def get_inbound_shipment_report(request):
    page = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('per_page', 15))
    offset = (page - 1) * per_page

    try:
        with connections['Inbound_db'].cursor() as cursor:
            # Count total rows (for pagination)
            cursor.execute("SELECT COUNT(*) FROM   [WHR_INBOUND_BAYAN_HEADER_TBL]")
            total = cursor.fetchone()[0]

            # Fetch all columns (*) with pagination
            cursor.execute(f"""
                SELECT *
                FROM   [WHR_INBOUND_BAYAN_HEADER_TBL]
                ORDER BY ID DESC
                OFFSET {offset} ROWS FETCH NEXT {per_page} ROWS ONLY;
            """)

            columns = [col[0] for col in cursor.description]
            results = [dict(zip(columns, row)) for row in cursor.fetchall()]

        return JsonResponse({
            'status': 'success',
            'data': results,
            'total': total,
            'page': page,
            'per_page': per_page,
            'total_pages': (total + per_page - 1) // per_page
        })

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

def get_existing_dOC_numbers(request):
    try:
        with connections['Inbound_db'].cursor() as cursor:
            cursor.execute("""
                SELECT DOC_NO 
                FROM   [WHR_INBOUND_BAYAN_HEADER_TBL]
                WHERE DOC_NO IS NOT NULL
            """)
            results = [row[0] for row in cursor.fetchall()]
            
        return JsonResponse({
            'data': results,
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def get_existing_bayan_numbers(request):
    try:
        with connections['Inbound_db'].cursor() as cursor:
            cursor.execute("""
                SELECT BAYAN_NO 
                FROM   [WHR_INBOUND_BAYAN_HEADER_TBL]
                WHERE DOC_NO IS NOT NULL
            """)
            results = [row[0] for row in cursor.fetchall()]
            
        return JsonResponse({
            'data': results,
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

class SaveContainerInfoView(APIView):
    @method_decorator(csrf_exempt, name='dispatch')
    def post(self, request, *args, **kwargs):
        try:
            data = request.data
            print("📦 Received Container Data:", data)

            required_fields = [ "DOC_NO", "CONTAINER_NO", "SIZE"]
            for field in required_fields:
                if field not in data or data[field] is None:
                    return Response({"status": "error", "message": f"Missing field: {field}"}, status=status.HTTP_400_BAD_REQUEST)

            values = [
                data.get("DOC_NO"),
                data.get("CONTAINER_NO"),
                data.get("SIZE")
            ]

            with connections['Inbound_db'].cursor() as cursor:
                cursor.execute("""
                    INSERT INTO   [WHR_INBOUND_CONTAINER_INFO] (
                        DOC_NO, CONTAINER_NO, SIZE
                    ) VALUES (%s, %s, %s)
                """, values)

            return Response({"status": "success", "message": "Container info saved successfully."}, status=status.HTTP_200_OK)

        except Exception as e:
            import traceback
            print("❌ Error saving container info:", str(e))
            print(traceback.format_exc())
            return Response({"status": "error", "message": "Internal server error."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class SaveProductInfoView(APIView):
    def post(self, request, *args, **kwargs):
        try:
            product_list = request.data
            print("📦 Received Product List:", product_list)

            if not isinstance(product_list, list) or not product_list:
                return Response({
                    "status": "error",
                    "message": "Request data must be a non-empty list of products."
                }, status=status.HTTP_400_BAD_REQUEST)

            required_fields = [
                "DOC_NO", "BAYAN_NO", "PO_NUMBER", "FRANCHISE", "CLASS",
                "SUBCLASS", "ITEM_CODE", "ITEM_DESC", "PO_QTY", "REC_QTY",
                "BALANCE_QTY", "SHIPPED_QTY", "CONTAINER_NO",
                "ASSIGNED_QTY", "SCANNED_QTY", "TRUCK_NUMBER",
                "DRIVER_INFO", "INITIATOR_NO", "INITIATOR_NAME", "STATUS"
            ]

            # Validate all products
            for idx, data in enumerate(product_list):
                for field in required_fields:
                    if field not in data:
                        return Response({
                            "status": "error",
                            "message": f"Missing field: '{field}' in product #{idx + 1}"
                        }, status=status.HTTP_400_BAD_REQUEST)

            with connections['Inbound_db'].cursor() as cursor:
                for data in product_list:
                    try:
                        values = [
                            data.get("DOC_NO"),
                            data.get("BAYAN_NO"),
                            data.get("PO_NUMBER"),
                            data.get("FRANCHISE"),
                            data.get("CLASS"),
                            data.get("SUBCLASS"),
                            data.get("ITEM_CODE"),
                            data.get("ITEM_DESC"),
                            float(data.get("PO_QTY")),
                            float(data.get("REC_QTY")),
                            float(data.get("BALANCE_QTY")),
                            float(data.get("SHIPPED_QTY")),
                            data.get("CONTAINER_NO") or "",
                            float(data.get("ASSIGNED_QTY")),
                            float(data.get("SCANNED_QTY")),
                            data.get("TRUCK_NUMBER") or "",
                            None if not data.get("ARRIVAL_DATE") else data.get("ARRIVAL_DATE"),
                            None if not data.get("EXIT_DATE") else data.get("EXIT_DATE"),
                            data.get("DRIVER_INFO") or "",
                            data.get("INITIATOR_NO"),
                            data.get("INITIATOR_NAME"),
                            data.get("STATUS"),
                        ]

                        cursor.execute("""
                            INSERT INTO   [WHR_INBOUND_BAYAN_DETAILS_TBL] (
                                DOC_NO, BAYAN_NO, PO_NUMBER, FRANCHISE, CLASS, SUBCLASS,
                                ITEM_CODE, ITEM_DESC, PO_QTY, REC_QTY, BALANCE_QTY, SHIPPED_QTY,
                                CONTAINER_NO, ASSIGNED_QTY, SCANNED_QTY, TRUCK_NUMBER,
                                ARRIVAL_DATE, EXIT_DATE, DRIVER_INFO,
                                INITIATOR_NO, INITIATOR_NAME, STATUS
                            )
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, values)

                    except Exception as insert_err:
                        return Response({
                            "status": "error",
                            "message": f"Failed to insert product #{product_list.index(data) + 1}: {str(insert_err)}"
                        }, status=status.HTTP_400_BAD_REQUEST)

            return Response({
                "status": "success",
                "message": "All product info saved successfully."
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({
                "status": "error",
                "message": f"Unexpected error: {str(e)}"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



@require_GET
def get_inbound_products(request):
    try:
        # Params
        doc_no = request.GET.get('doc_no')
        per_page = int(request.GET.get('per_page', 20))
        last_id = request.GET.get('last_id')  # keyset pagination anchor

        # Base query
        base_query = """
            FROM  WHR_INBOUND_BAYAN_DETAILS_TBL
            WHERE 1=1
        """
        params = []

        if doc_no:
            base_query += " AND DOC_NO = %s"
            params.append(doc_no)

        if last_id:
            base_query += " AND PRODUCT_ID > %s"
            params.append(last_id)

        # Main query (keyset pagination)
        data_query = f"""
            SELECT TOP {per_page}
                PRODUCT_ID, DOC_NO, PO_NUMBER, FRANCHISE,
                CLASS, SUBCLASS, ITEM_CODE, ITEM_DESC, 
                PO_QTY, REC_QTY, BALANCE_QTY, SHIPPED_QTY, 
                CONTAINER_NO, TRUCK_NUMBER, ARRIVAL_DATE, EXIT_DATE, DRIVER_INFO
            {base_query}
            ORDER BY PRODUCT_ID ASC
        """

        with connections['Inbound_db'].cursor() as cursor:
            cursor.execute(data_query, params)
            columns = [col[0] for col in cursor.description]
            data = [dict(zip(columns, row)) for row in cursor.fetchall()]

        # Count only if first page (expensive on lakhs of rows)
        total_count = None
        if not last_id:
            count_query = f"SELECT COUNT(*) {base_query}"
            with connections['Inbound_db'].cursor() as cursor:
                cursor.execute(count_query, params)
                total_count = cursor.fetchone()[0]

        return JsonResponse({
            'status': 'success',
            'data': data,
            'pagination': {
                'per_page': per_page,
                'last_id': data[-1]['PRODUCT_ID'] if data else None,
                'total': total_count
            }
        })

    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)
        

@require_GET
def get_all_bayan_details(request):
    try:
        # Pagination params
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 20))
        offset = (page - 1) * per_page

        # Query WHR_INBOUND_BAYAN_DETAILS_TBL with pagination
        query = """
            SELECT 
                PRODUCT_ID, DOC_NO, BAYAN_NO, PO_NUMBER, FRANCHISE, CLASS, SUBCLASS, 
                ITEM_CODE, ITEM_DESC, PO_QTY, REC_QTY, BALANCE_QTY, SHIPPED_QTY, 
                CONTAINER_NO, ASSIGNED_QTY, SCANNED_QTY, TRUCK_NUMBER, ARRIVAL_DATE, 
                EXIT_DATE, DRIVER_INFO, INITIATOR_NO, INITIATOR_NAME, STATUS
            FROM  WHR_INBOUND_BAYAN_DETAILS_TBL
            ORDER BY PRODUCT_ID
            OFFSET %s ROWS FETCH NEXT %s ROWS ONLY
        """
        params = [offset, per_page]

        with connections['Inbound_db'].cursor() as cursor:
            cursor.execute(query, params)
            columns = [col[0] for col in cursor.description]
            data = [dict(zip(columns, row)) for row in cursor.fetchall()]

        # DEBUG: Print sample data
        print(f"DEBUG: Returning {len(data)} products")
        if data:
            print(f"DEBUG: First product keys: {data[0].keys()}")
            print(f"DEBUG: First product DOC_NO: {data[0].get('DOC_NO')}")

        # Count total rows
        count_query = "SELECT COUNT(*) FROM  WHR_INBOUND_BAYAN_DETAILS_TBL"
        with connections['Inbound_db'].cursor() as cursor:
            cursor.execute(count_query)
            total_count = cursor.fetchone()[0]

        return JsonResponse({
            "status": "success",
            "data": data,
            "pagination": {
                "total": total_count,
                "page": page,
                "per_page": per_page,
                "total_pages": (total_count + per_page - 1) // per_page
            }
        })

    except Exception as e:
        print(f"ERROR in get_all_bayan_details: {str(e)}")
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)
    

# Add this to your Django views.py
@require_GET
def get_products_by_docno(request):
    try:
        doc_no = request.GET.get('doc_no', '')
        
        if not doc_no:
            return JsonResponse({
                "status": "error",
                "message": "DOC_NO parameter is required"
            }, status=400)
        
        # Query products by DOC_NO
        query = """
            SELECT TOP (1000)
                PRODUCT_ID, DOC_NO, BAYAN_NO, PO_NUMBER, FRANCHISE, CLASS, SUBCLASS, 
                ITEM_CODE, ITEM_DESC, PO_QTY, REC_QTY, BALANCE_QTY, SHIPPED_QTY, 
                CONTAINER_NO, ASSIGNED_QTY, SCANNED_QTY, TRUCK_NUMBER, ARRIVAL_DATE, 
                EXIT_DATE, DRIVER_INFO, INITIATOR_NO, INITIATOR_NAME, STATUS
            FROM  WHR_INBOUND_BAYAN_DETAILS_TBL
            WHERE DOC_NO = %s
            ORDER BY PRODUCT_ID
        """
        params = [doc_no]

        with connections['Inbound_db'].cursor() as cursor:
            cursor.execute(query, params)
            columns = [col[0] for col in cursor.description]
            data = [dict(zip(columns, row)) for row in cursor.fetchall()]

        print(f"DEBUG: Found {len(data)} products for DOC_NO: {doc_no}")
        
        return JsonResponse({
            "status": "success",
            "data": data,
            "count": len(data)
        })

    except Exception as e:
        print(f"ERROR in get_products_by_docno: {str(e)}")
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)
    
@csrf_exempt
def update_truck_details(request):
    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])

    try:
        data = json.loads(request.body)
        print(f"Incoming data: {data}")  # Log the incoming data

        product_id   = data.get('product_id')
        truck_no     = data.get('truck_no')
        arrival_date = data.get('arrival_date')
        exit_date    = data.get('exit_date')
        driver_info  = data.get('driver_info')
        bayan_no  = data.get('bayan_no')

        with connections['Inbound_db'].cursor() as cursor:
            cursor.execute("""
                UPDATE   [WHR_INBOUND_PRODUCT_INFO]
                SET TRUCK_NUMBER= %s,
                    ARRIVAL_DATE = %s,
                    EXIT_DATE = %s,
                    DRIVER_INFO = %s,
                    BAYAN_NO = %s
                WHERE PRODUCT_ID = %s
            """, [truck_no, arrival_date, exit_date, driver_info,bayan_no, product_id])

        return JsonResponse({'success': True})

    except Exception as e:
        print(f"Error updating truck details: {str(e)}")  # Log the error
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    

@csrf_exempt
def update_pickman_assignments(request):
    if request.method == 'POST':
        try:
            # Parse request data
            data = json.loads(request.body)
            logger.info(f"Received data: {data}")
            
            product_id = data.get('product_id')
            pickmen = data.get('pickmen', [])
            total_qty = data.get('total_qty', 0)
            whr_superuser = data.get('whr_superuser')
            whr_superuser_name = data.get('whr_superuser_name')
            
            logger.info(f"Product ID: '{product_id}' (type: {type(product_id)})")
            logger.info(f"WH Superuser: '{whr_superuser}', Name: '{whr_superuser_name}'")
            
            # Validate required fields
            if not product_id:
                logger.error("Missing product_id")
                return JsonResponse({'error': 'product_id is required'}, status=400)
            
            with connections['Inbound_db'].cursor() as cursor:
                # 1. First, check if the product exists
                cursor.execute(
                    "SELECT COUNT(*) FROM   [WHR_INBOUND_BAYAN_DETAILS_TBL] WHERE PRODUCT_ID = %s",
                    [product_id]
                )
                product_count = cursor.fetchone()[0]
                logger.info(f"Products found with ID '{product_id}': {product_count}")
                
                # 2. Delete existing assignments for this product
                cursor.execute(
                    "DELETE FROM   [WHR_INBOUND_ASSIGNED_PICKMAN] WHERE PRODUCT_ID = %s",
                    [product_id]
                )
                logger.info(f"Deleted existing assignments for product: {product_id}")
                
                # 3. Aggregate quantities by pickman
                pickman_aggregated = {}
                total_assigned_qty = 0
                for pickman in pickmen:
                    name = str(pickman.get('name', '')).strip()
                    qty = int(pickman.get('qty', 0))
                    
                    if name:
                        if name in pickman_aggregated:
                            pickman_aggregated[name] += qty
                        else:
                            pickman_aggregated[name] = qty
                        total_assigned_qty += qty
                
                logger.info(f"Aggregated pickmen: {pickman_aggregated}, total: {total_assigned_qty}")
                
                # 4. Insert new assignments
                for name, qty in pickman_aggregated.items():
                    cursor.execute(
                        """INSERT INTO   [WHR_INBOUND_ASSIGNED_PICKMAN] 
                        (PRODUCT_ID, ASSIGNED_PICKMAN, QTY) 
                        VALUES (%s, %s, %s)""",
                        [product_id, name, qty]
                    )
                
                logger.info(f"Inserted {len(pickman_aggregated)} pickman assignments")
                
                # 5. Update WHR_INBOUND_BAYAN_DETAILS_TBL table with superuser info
                # Use parameterized query to avoid SQL injection
                update_query = """
                    UPDATE   [WHR_INBOUND_BAYAN_DETAILS_TBL] 
                    SET ASSIGNED_QTY = %s, 
                        WHR_SUPERUSER_NO = %s,
                        WHR_SUPERUSER_NAME = %s
                    WHERE PRODUCT_ID = %s
                """
                
                logger.info(f"Executing UPDATE: {update_query}")
                logger.info(f"With values: {[total_assigned_qty, whr_superuser, whr_superuser_name, product_id]}")
                
                cursor.execute(update_query, [total_assigned_qty, whr_superuser, whr_superuser_name, product_id])
                
                rows_affected = cursor.rowcount
                logger.info(f"Rows affected by UPDATE: {rows_affected}")
                
                if rows_affected == 0:
                    logger.warning(f"No rows updated for PRODUCT_ID: {product_id}")
                    # Check if we need to insert instead of update
                    cursor.execute(
                        "SELECT COUNT(*) FROM   [WHR_INBOUND_BAYAN_DETAILS_TBL] WHERE PRODUCT_ID = %s",
                        [product_id]
                    )
                    exists = cursor.fetchone()[0] > 0
                    if not exists:
                        logger.error(f"Product ID {product_id} does not exist in the table")
                
                # 6. Calculate scan status
                scan_status = (
                    'Send Scan' if total_assigned_qty == 0 else
                    'Finish' if total_assigned_qty >= total_qty else
                    'Pending'
                )
                
                response_data = {
                    'status': 'success',
                    'scan_status': scan_status,
                    'total_scanned': total_assigned_qty,
                    'assigned_qty_updated': total_assigned_qty,
                    'rows_affected': rows_affected
                }
                
                logger.info(f"Returning response: {response_data}")
                return JsonResponse(response_data)
                
        except Exception as e:
            logger.error(f"Error in update_pickman_assignments: {str(e)}", exc_info=True)
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.db import connection

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.db import connection

@csrf_exempt
def get_pickmenlist(request):
    try:
        with connections['Inbound_db'].cursor() as cursor:
            cursor.execute("""
                SELECT EMP_NAME, EMP_USERNAME
                FROM   [WHR_USER_MANAGEMENT]
                WHERE EMP_ROLE = %s
                ORDER BY EMP_NAME
            """, ['Inbound Pickman'])  # ✅ Safe parameterized query

            rows = cursor.fetchall()

        # Convert to list of dicts
        result = [{'name': row[0], 'username': row[1]} for row in rows]

        return JsonResponse({
            'status': 'success',
            'count': len(result),   # ✅ helpful for bulk data
            'data': result
        }, safe=False)

    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)

@csrf_exempt
def save_pickman_data(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request method'}, status=405)

    try:
        data = json.loads(request.body)
        print("Received data:", data)  # Debug logging

        # Convert date
        date_str = data.get('DATE')
        try:
            formatted_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except:
            return JsonResponse({'error': 'Invalid date format'}, status=400)

        print("Assigned QTY from request:", data.get('ASSIGNED_QTY'))  # Debug logging

        insert_query = """
            INSERT INTO   [WHR_INBOUND_PICKMAN_TBL] (
                PICK_ID, DOC_NO, BAYAN_NO, CONTAINER_NO, DATE, PO_NUMBER, FRANCHISE,
                FAMILY, CLASS, SUBCLASS, ITEM_CODE, PO_QTY, SHIPPED_QTY, SCANNED_QTY,
                PRODUCT_CODE, SERIAL_NO, DAMAGE, LOCATION, PICK_MAN,
                WHR_MANAGER_STATUS, WHR_SUPERUSER_STATUS,
                WHR_SPARES_RECEIVER_STATUS, WHR_PRODUCT_RECEIVER_STATUS, ASSIGNED_QTY,PICKMAN_NO
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s)
        """

        values = (
            data.get('PICK_ID', ''),
            data.get('DOC_NO', ''),
            data.get('BAYAN_NO', ''),
            data.get('CONTAINER_NO', ''),
            formatted_date,
            data.get('PO_NUMBER', ''),
            data.get('FRANCHISE', ''),
            data.get('FAMILY', ''),
            data.get('CLASS', ''),
            data.get('SUBCLASS', ''),
            data.get('ITEM_CODE', ''),
            data.get('PO_QTY', 0),
            data.get('SHIPPED_QTY', 0),
            data.get('SCANNED_QTY', 0),
            data.get('PRODUCT_CODE', ''),
            data.get('SERIAL_NO', ''),
            data.get('DAMAGE', False),
            data.get('LOCATION', ''),
            data.get('PICK_MAN', ''),
            data.get('WHR_MANAGER_STATUS', ''),
            data.get('WHR_SUPERUSER_STATUS', ''),
            data.get('WHR_SPARES_RECEIVER_STATUS', ''),
            data.get('WHR_PRODUCT_RECEIVER_STATUS', ''),
            data.get('ASSIGNED_QTY', 0),    
            data.get('PICKMAN_NO', ''),

        )

        print("Values being inserted:", values)  # Debug logging

        with connections['Inbound_db'].cursor() as cursor:
            cursor.execute(insert_query, values)

        return JsonResponse({'message': 'Data saved successfully'}, status=200)

    except Exception as e:
        print("Error saving data:", str(e))  # Debug logging
        return JsonResponse({'error': str(e)}, status=400)

@csrf_exempt
def pickman_details(request, pickman_name):
    try:
        # Decode pickman_name if it comes URL-encoded
        pickman_name = unquote(pickman_name)

        with connections['Inbound_db'].cursor() as cursor:
            query = """
                SELECT 
                    ap.[PRODUCT_ID],
                    ap.[ASSIGNED_PICKMAN],
                    ap.[QTY] AS ASSIGNED_QTY,
                    pi.[DOC_NO],
                    pi.[PO_NUMBER],
                    pi.[ITEM_CODE],
                    pi.[ITEM_DESC],
                    pi.[CONTAINER_NO],
                    pi.[FRANCHISE],
                    pi.[PO_QTY],        -- Purchase Order Quantity
                    pi.[SHIPPED_QTY],   -- Shipped Quantity
                    pi.[SCANNED_QTY],   -- Scanned Quantity
                    pi.[PRODUCT_ID] AS DETAIL_PRODUCT_ID,
                    pi.[BAYAN_NO],
                    pi.[CLASS],
                    pi.[SUBCLASS]
                FROM   [WHR_INBOUND_ASSIGNED_PICKMAN] ap
                JOIN   [WHR_INBOUND_BAYAN_DETAILS_TBL] pi
                    ON ap.PRODUCT_ID = pi.PRODUCT_ID
                WHERE ap.ASSIGNED_PICKMAN = %s
                ORDER BY ap.PRODUCT_ID
            """
            cursor.execute(query, [pickman_name])
            columns = [col[0] for col in cursor.description]
            results = [dict(zip(columns, row)) for row in cursor.fetchall()]

        return JsonResponse({
            'status': 'success',
            'count': len(results),   # ✅ Helpful when handling thousands of records
            'data': results
        })

    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)

from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.db import connection

@csrf_exempt
def get_pickman_BAYAN_data(request):
    if request.method != 'GET':
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=400)

    try:
        with connections['Inbound_db'].cursor() as cursor:
            # Optimized query to calculate status for each DOC_NO + BAYAN_NO
            cursor.execute("""
                SELECT 
                    t1.*,
                    CASE 
                        WHEN t1.ASSIGNED_QTY = t2.total_scanned THEN 'Complete'
                        ELSE 'Incomplete'
                    END AS STATUS
                FROM 
                      [WHR_INBOUND_PICKMAN_TBL] t1
                INNER JOIN (
                    SELECT 
                        DOC_NO, 
                        BAYAN_NO, 
                        SUM(SCANNED_QTY) AS total_scanned
                    FROM 
                          [WHR_INBOUND_PICKMAN_TBL]
                    GROUP BY 
                        DOC_NO, BAYAN_NO
                ) t2 
                ON t1.DOC_NO = t2.DOC_NO AND t1.BAYAN_NO = t2.BAYAN_NO
                ORDER BY 
                    t1.DOC_NO, t1.BAYAN_NO
            """)

            columns = [col[0] for col in cursor.description]
            results = [dict(zip(columns, row)) for row in cursor.fetchall()]

        return JsonResponse({'status': 'success', 'data': results})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    
from django.http import JsonResponse
from django.db import connections
from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def get_inbound_shipment_docNo_details(request, doc_no):
    try:
        with connections['Inbound_db'].cursor() as cursor:
            query = """
                SELECT 
                    DOC_NO, HOUSE_PL, CONTAINER_COUNT, CLEARANCE_DATE,
                    SUPPLIER_NAME, SUPPLIER_NO, BL_DATE, POL, LC_NO1,
                    INVOICE_NO1, VALUE1, PAYMENT_TERM1, INCOTERMS1,
                    ETD_DATE, POD, BILL_NO, INVOICE_NO2, VALUE2,
                    PAYMENT_TERMS2, INCOTERMS2, ETA_DATE, LINE_NAME, BILL_DUE_DATE,
                    INVOICE_NO3, VALUE3, PAYMENT_TERM3, INCOTERMS3
                FROM   [WHR_INBOUND_BAYAN_HEADER_TBL]
                WHERE DOC_NO = %s
            """
            cursor.execute(query, [doc_no])
            row = cursor.fetchone()

            if not row:
                return JsonResponse({'status': 'error', 'message': 'Document not found'}, status=404)

            # Extract columns + row
            columns = [col[0] for col in cursor.description]
            document_data = dict(zip(columns, row))

        # ✅ Safe float conversion for numeric fields
        def safe_float(value):
            try:
                return float(value)
            except (TypeError, ValueError):
                return None  # keep None instead of 0.0 (to preserve meaning)

        numeric_fields = ['CONTAINER_COUNT']
        for field in numeric_fields:
            if field in document_data:
                document_data[field] = safe_float(document_data[field])

        return JsonResponse({'status': 'success', 'data': document_data})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
@csrf_exempt
def update_manager_status(request):
    if request.method == 'POST':
        try:
            # Parse JSON data from request body
            data = json.loads(request.body)
            pick_id = data.get('pick_id')
            status = data.get('status')
            
            # Validate required fields
            if not pick_id or not status:
                return JsonResponse(
                    {'success': False, 'error': 'Missing required fields (pick_id or status)'}, 
                    status=400
                )
            
            # Execute the database update
            with connections['Inbound_db'].cursor() as cursor:
                cursor.execute("""
                    UPDATE   [WHR_INBOUND_PICKMAN_TBL]
                    SET WHR_MANAGER_STATUS = %s
                    WHERE PICK_ID = %s
                """, [status, pick_id])
                
                # Check if any rows were affected
                if cursor.rowcount == 0:
                    return JsonResponse(
                        {'success': False, 'message': 'No records updated - PICK_ID not found'},
                        status=404
                    )
                    
            return JsonResponse({'success': True})
            
        except json.JSONDecodeError:
            return JsonResponse(
                {'success': False, 'error': 'Invalid JSON data'},
                status=400
            )
        except Exception as e:
            return JsonResponse(
                {'success': False, 'error': str(e)},
                status=500
            )
    
    # Handle non-POST requests
    return JsonResponse(
        {'success': False, 'error': 'Only POST method is allowed'},
        status=405
    )


@csrf_exempt
def update_superuser_status(request):
    if request.method == 'POST':
        try:
            # Parse JSON data from request body
            data = json.loads(request.body)
            pick_id = data.get('pick_id')
            status = data.get('status')
            
            # Validate required fields
            if not pick_id or not status:
                return JsonResponse(
                    {'success': False, 'error': 'Missing required fields (pick_id or status)'}, 
                    status=400
                )
            
            # Execute the database update
            with connections['Inbound_db'].cursor() as cursor:
                cursor.execute("""
                    UPDATE   [WHR_INBOUND_PICKMAN_TBL]
                    SET WHR_SUPERUSER_STATUS = %s
                    WHERE PICK_ID = %s
                """, [status, pick_id])
                
                # Check if any rows were affected
                if cursor.rowcount == 0:
                    return JsonResponse(
                        {'success': False, 'message': 'No records updated - PICK_ID not found'},
                        status=404
                    )
                    
            return JsonResponse({'success': True})
            
        except json.JSONDecodeError:
            return JsonResponse(
                {'success': False, 'error': 'Invalid JSON data'},
                status=400
            )
        except Exception as e:
            return JsonResponse(
                {'success': False, 'error': str(e)},
                status=500
            )
    


@csrf_exempt
def update_sparesreceiver_status(request):
    if request.method == 'POST':
        try:
            # Parse JSON data from request body
            data = json.loads(request.body)
            pick_id = data.get('pick_id')
            status = data.get('status')
            
            # Validate required fields
            if not pick_id or not status:
                return JsonResponse(
                    {'success': False, 'error': 'Missing required fields (pick_id or status)'}, 
                    status=400
                )
            
            # Execute the database update
            with connections['Inbound_db'].cursor() as cursor:
                cursor.execute("""
                    UPDATE   [WHR_INBOUND_PICKMAN_TBL]
                    SET WHR_SPARES_RECEIVER_STATUS = %s
                    WHERE PICK_ID = %s
                """, [status, pick_id])
                
                # Check if any rows were affected
                if cursor.rowcount == 0:
                    return JsonResponse(
                        {'success': False, 'message': 'No records updated - PICK_ID not found'},
                        status=404
                    )
                    
            return JsonResponse({'success': True})
            
        except json.JSONDecodeError:
            return JsonResponse(
                {'success': False, 'error': 'Invalid JSON data'},
                status=400
            )
        except Exception as e:
            return JsonResponse(
                {'success': False, 'error': str(e)},
                status=500
            )


    # Handle non-POST requests
    return JsonResponse(
        {'success': False, 'error': 'Only POST method is allowed'},
        status=405
    )

@csrf_exempt
def update_productreceiver_status(request):
    if request.method == 'POST':
        try:
            # Parse JSON data from request body
            data = json.loads(request.body)
            pick_id = data.get('pick_id')
            status = data.get('status')
            
            # Validate required fields
            if not pick_id or not status:
                return JsonResponse(
                    {'success': False, 'error': 'Missing required fields (pick_id or status)'}, 
                    status=400
                )
            
            # Execute the database update
            with connections['Inbound_db'].cursor() as cursor:
                cursor.execute("""
                    UPDATE   [WHR_INBOUND_PICKMAN_TBL]
                    SET WHR_PRODUCT_RECEIVER_STATUS = %s
                    WHERE PICK_ID = %s
                """, [status, pick_id])
                
                # Check if any rows were affected
                if cursor.rowcount == 0:
                    return JsonResponse(
                        {'success': False, 'message': 'No records updated - PICK_ID not found'},
                        status=404
                    )
                    
            return JsonResponse({'success': True})
            
        except json.JSONDecodeError:
            return JsonResponse(
                {'success': False, 'error': 'Invalid JSON data'},
                status=400
            )
        except Exception as e:
            return JsonResponse(
                {'success': False, 'error': str(e)},
                status=500
            )
def get_item_description(request, item_code):
    try:
        with connections['default'].cursor() as cursor:
            query = """
                SELECT DESCRIPTION, FRANCHISE, CLASS, SUBCLASS
                FROM [BUYP].[BUYP].[ALJE_ITEM_CATEGORIES_CPD_V]
                WHERE ITEM_CODE = %s
            """
            cursor.execute(query, [item_code])
            row = cursor.fetchone()

        if row:
            return JsonResponse({
                'item_code': item_code,
                'item_desc': row[0],  # DESCRIPTION
                'franchise': row[1],  # FRANCHISE
                'class': row[2],      # CLASS
                'subclass': row[3]    # SUBCLASS
            })
        else:
            return JsonResponse({
                'item_code': item_code,
                'item_desc': None,
                'franchise': None,
                'class': None,
                'subclass': None
            }, status=404)
    
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    
from django.http import JsonResponse
from django.db import connection
from rest_framework.views import APIView

class UniqueSupplierListView(APIView):
    def get(self, request, *args, **kwargs):
        try:
            with connections['default'].cursor() as cursor:
                cursor.execute("""
                    SELECT DISTINCT VENDOR_NAME
                    FROM   [XXALJEBUYP_PENDING_PO]
                    WHERE VENDOR_NAME IS NOT NULL
                    ORDER BY VENDOR_NAME
                """)
                rows = cursor.fetchall()

            # Convert into simple list
            vendor_list = [row[0] for row in rows]

            return JsonResponse({"status": "success", "vendors": vendor_list}, safe=False)

        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=500)

@csrf_exempt
def get_all_fully_scanned_pickman_details(request):
    try:
        with connections['Inbound_db'].cursor() as cursor:
            query = """
                SELECT k.*, p.ASSIGNED_QTY
                FROM   [WHR_INBOUND_BAYAN_DETAILS_TBL] p
                JOIN (
                    SELECT BAYAN_NO, ITEM_CODE, SUM(SCANNED_QTY) AS total_scanned
                    FROM   [WHR_INBOUND_PICKMAN_TBL]
                    GROUP BY BAYAN_NO, ITEM_CODE
                ) ksum 
                    ON p.BAYAN_NO = ksum.BAYAN_NO 
                   AND p.ITEM_CODE = ksum.ITEM_CODE
                JOIN   [WHR_INBOUND_PICKMAN_TBL] k 
                    ON k.BAYAN_NO = ksum.BAYAN_NO 
                   AND k.ITEM_CODE = ksum.ITEM_CODE
                WHERE p.ASSIGNED_QTY = ksum.total_scanned
                ORDER BY k.BAYAN_NO, k.ITEM_CODE
            """
            cursor.execute(query)
            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()
            data = [dict(zip(columns, row)) for row in rows]

        return JsonResponse({
            "status": "success",
            "count": len(data),   # ✅ helps in big datasets
            "matched_pickman_rows": data
        })

    except Exception as e:
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=500)


@method_decorator(csrf_exempt, name='dispatch')
class ProductPartsByPOView(View):
    def get(self, request):
        try:
            with connections['default'].cursor() as cursor:
                # Execute raw SQL query to get both PO_NUMBER and PROD_PART
                query = """
                SELECT PO_NUMBER, PROD_PART
                FROM   [XXALJEBUYP_PENDING_PO] 
                """
                cursor.execute(query)
                
                # Extract both PO_NUMBER and PROD_PART
                results = [
                    {'po_number': row[0], 'prod_part': row[1]}
                    for row in cursor.fetchall()
                ]
                
            return JsonResponse({
                'status': 'success',
                'data': results
            })
            
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=500)
        

@csrf_exempt
def check_po_exists(request, po_number):
    try:
        if not po_number:
            return JsonResponse({'exists': False, 'message': 'PO number not provided'}, status=400)

        query = """
            SELECT COUNT(*) FROM  WHR_INBOUND_PRODUCT_INFO
            WHERE PO_NUMBER = %s
        """
        with connections['Inbound_db'].cursor() as cursor:
            cursor.execute(query, [po_number])
            count = cursor.fetchone()[0]

        return JsonResponse({'exists': count > 0})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
def get_pickmen_for_products(request):
    if request.method == 'GET':
        try:
            # Extract comma-separated product_ids
            product_ids_str = request.GET.get('product_ids', '')
            product_ids = [pid.strip() for pid in product_ids_str.split(',') if pid.strip()]
            
            if not product_ids:
                return JsonResponse({'data': []})

            # Build placeholders safely for IN clause
            placeholders = ','.join(['%s'] * len(product_ids))
            query = f"""
                SELECT PRODUCT_ID, ASSIGNED_PICKMAN, QTY, ID
                FROM   [WHR_INBOUND_ASSIGNED_PICKMAN]
                WHERE PRODUCT_ID IN ({placeholders})
                ORDER BY ID
            """

            with connections['Inbound_db'].cursor() as cursor:
                cursor.execute(query, product_ids)
                columns = [col[0] for col in cursor.description]
                results = [dict(zip(columns, row)) for row in cursor.fetchall()]

            return JsonResponse({'data': results})

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'GET method required'}, status=400)

@csrf_exempt
def update_scanned_qty(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            product_id = data.get('product_id')
            new_scanned_qty = data.get('scanned_qty')  # The new quantity to add

            if not product_id or new_scanned_qty is None:
                return JsonResponse(
                    {'status': 'error', 'message': 'Missing product_id or scanned_qty'},
                    status=400
                )

            with connections['Inbound_db'].cursor() as cursor:
                # 1. First get the current SCANNED_QTY
                cursor.execute(
                    """
                    SELECT [SCANNED_QTY] 
                    FROM   [WHR_INBOUND_BAYAN_DETAILS_TBL]
                    WHERE [PRODUCT_ID] = %s
                    """,
                    [product_id]
                )
                row = cursor.fetchone()
                
                if not row:
                    return JsonResponse(
                        {'status': 'error', 'message': 'Product not found'},
                        status=404
                    )

                current_qty = row[0] or 0  # Handle NULL values
                updated_qty = current_qty + new_scanned_qty

                # 2. Update with the new total
                cursor.execute(
                    """
                    UPDATE   [WHR_INBOUND_BAYAN_DETAILS_TBL]
                    SET [SCANNED_QTY] = %s
                    WHERE [PRODUCT_ID] = %s
                    """,
                    [updated_qty, product_id]
                )

            return JsonResponse({
                'status': 'success',
                'message': f'SCANNED_QTY updated from {current_qty} to {updated_qty}'
            })

        except Exception as e:
            return JsonResponse(
                {'status': 'error', 'message': str(e)},
                status=500
            )

    return JsonResponse(
        {'status': 'error', 'message': 'Only POST allowed'},
        status=405
    )



def get_expense_cat(request):
    try:
        with connections['Inbound_db'].cursor() as cursor:
            cursor.execute("""
                SELECT DISTINCT [EXPENSE_CAT]
                FROM   [WHR_INBOUND_EXPENSE_CAT]
                ORDER BY [EXPENSE_CAT]
            """)
            rows = [row[0] for row in cursor.fetchall()]
        return JsonResponse({"expense_categories": rows})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


def get_names_by_expense_cat(request, cat):
    try:
        with connections['Inbound_db'].cursor() as cursor:
            cursor.execute("""
                SELECT [NAME]
                FROM   [WHR_INBOUND_EXPENSE_CAT_NAME]
                WHERE [EXPENSE_CAT] = %s
                ORDER BY [NAME]
            """, [cat])   # parameterized -> safe
            rows = [row[0] for row in cursor.fetchall()]
        return JsonResponse({"category": cat, "names": rows})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)



@method_decorator(csrf_exempt, name='dispatch')
class SaveExpenseView(APIView):
    def post(self, request, *args, **kwargs):
        try:
            data = request.data
            print("📥 Received Expense Data:", data)

            expenses = data.get("expense_data")
            if not expenses or not isinstance(expenses, list):
                return Response({
                    "status": "error",
                    "message": "Invalid or missing 'expense_data'. Must be a list."
                }, status=status.HTTP_400_BAD_REQUEST)

            # Get current datetime
            from django.utils import timezone
            current_datetime = timezone.now()

            with connections['Inbound_db'].cursor() as cursor:
                for expense in expenses:
                    # Required fields
                    required_fields = [
                        "BAYAN_NO", "EXPENSE_CAT", "NAME", 
                        "CURRENCY", "CONVERSION_RATE", "AMOUNT", "FINAL_AMOUNT",
                    ]
                    for field in required_fields:
                        if field not in expense or expense[field] in [None, ""]:
                            return Response({
                                "status": "error",
                                "message": f"Missing or empty field: {field}"
                            }, status=status.HTTP_400_BAD_REQUEST)

                    # Validate numeric fields
                    try:
                        conversion_rate = float(expense["CONVERSION_RATE"])
                        amount = float(expense["AMOUNT"])
                        final_amount = float(expense["FINAL_AMOUNT"])
                    except (TypeError, ValueError):
                        return Response({
                            "status": "error",
                            "message": "CONVERSION_RATE, AMOUNT, and FINAL_AMOUNT must be valid numbers"
                        }, status=status.HTTP_400_BAD_REQUEST)

                    # Get optional remarks field
                    remarks = expense.get("REMARKS", "")

                    # Insert into DB with current datetime
                    cursor.execute("""
                        INSERT INTO   [WHR_INBOUND_EXPENSE_DETAILS] (
                            BAYAN_NO, EXPENSE_CAT, NAME, CURRENCY, 
                            CONVERSION_RATE, AMOUNT, FINAL_AMOUNT, REMARKS, DATE
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, [
                        expense["BAYAN_NO"],
                        expense["EXPENSE_CAT"],
                        expense["NAME"],
                        expense["CURRENCY"],
                        conversion_rate,
                        amount,
                        final_amount,
                        remarks,
                        current_datetime
                    ])

            return Response({
                "status": "success",
                "message": "All expenses saved successfully."
            }, status=status.HTTP_200_OK)

        except Exception as e:
            import traceback
            print("❌ Error saving expenses:", str(e))
            print(traceback.format_exc())
            return Response({
                "status": "error",
                "message": "Internal server error."
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def get_inbound_expenses(request):
    try:
        # Query params
        per_page = int(request.GET.get('per_page', 20))
        last_id = request.GET.get('last_id')  # For keyset pagination
        search = request.GET.get('search', '')
        category = request.GET.get('category', '')
        currency = request.GET.get('currency', '')
        from_date = request.GET.get('from_date', '')
        to_date = request.GET.get('to_date', '')
        sort_by = request.GET.get('sort_by', 'id')
        sort_order = request.GET.get('sort_order', 'desc')

        # Allowed sort columns
        valid_sort_columns = [
            'id', 'BAYAN_NO', 'EXPENSE_CAT', 'NAME',
            'CURRENCY', 'CONVERSION_RATE', 'AMOUNT',
            'FINAL_AMOUNT', 'DATE', 'REMARKS'
        ]
        if sort_by not in valid_sort_columns:
            sort_by = 'id'

        order_direction = 'ASC' if sort_order.lower() == 'asc' else 'DESC'

        # Base query
        base_query = """
            FROM   [WHR_INBOUND_EXPENSE_DETAILS] 
            WHERE (flag IS NULL OR flag != 'd')
        """

        conditions = []
        params = []

        # Filters
        if from_date:
            conditions.append("CAST(DATE AS DATE) >= %s")
            params.append(from_date)
        if to_date:
            conditions.append("CAST(DATE AS DATE) <= %s")
            params.append(to_date)
        if category:
            conditions.append("EXPENSE_CAT = %s")
            params.append(category)
        if currency:
            conditions.append("CURRENCY = %s")
            params.append(currency)

        if search:
            search_conditions = []
            for field in ['BAYAN_NO', 'EXPENSE_CAT', 'NAME', 'CURRENCY', 'REMARKS']:
                search_conditions.append(f"{field} LIKE %s")
                params.append(f"%{search}%")
            conditions.append("(" + " OR ".join(search_conditions) + ")")

        # Keyset pagination
        if last_id:
            conditions.append("id < %s" if order_direction == "DESC" else "id > %s")
            params.append(last_id)

        # Append conditions
        if conditions:
            base_query += " AND " + " AND ".join(conditions)

        # Main query (keyset pagination)
        data_query = f"""
            SELECT TOP {per_page}
                id,
                BAYAN_NO,
                EXPENSE_CAT,
                NAME,
                CURRENCY,
                CONVERSION_RATE,
                AMOUNT,
                FINAL_AMOUNT,
                REMARKS,
                CONVERT(VARCHAR, DATE, 120) as DATE
            {base_query}
            ORDER BY {sort_by} {order_direction}
        """

        with connections['Inbound_db'].cursor() as cursor:
            cursor.execute(data_query, params)
            columns = [col[0] for col in cursor.description]
            expenses = [dict(zip(columns, row)) for row in cursor.fetchall()]

        # Get total count only when first page (avoid slow COUNT(*) for every page)
        total_count = None
        if not last_id:
            with connections['Inbound_db'].cursor() as cursor:
                cursor.execute(f"SELECT COUNT(*) {base_query}", params)
                total_count = cursor.fetchone()[0]

        return Response({
            "status": "success",
            "data": expenses,
            "pagination": {
                "per_page": per_page,
                "last_id": expenses[-1]['id'] if expenses else None,
                "total": total_count
            },
            "filters": {
                "search": search,
                "category": category,
                "currency": currency,
                "from_date": from_date,
                "to_date": to_date,
                "sort_by": sort_by,
                "sort_order": sort_order
            }
        }, status=status.HTTP_200_OK)

    except Exception as e:
        import traceback
        print("❌ Error fetching expenses:", str(e))
        print(traceback.format_exc())
        return Response({
            "status": "error",
            "message": "Internal server error while fetching expenses."
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def get_expense_filters(request):
    try:
        with connections['Inbound_db'].cursor() as cursor:
            # ✅ Use indexed DISTINCT (make sure EXPENSE_CAT and CURRENCY are indexed)
            cursor.execute("""
                SELECT EXPENSE_CAT 
                FROM   [WHR_INBOUND_EXPENSE_DETAILS] 
                WHERE EXPENSE_CAT IS NOT NULL 
                GROUP BY EXPENSE_CAT
                ORDER BY EXPENSE_CAT
            """)
            categories = [row[0] for row in cursor.fetchall()]

            cursor.execute("""
                SELECT CURRENCY 
                FROM   [WHR_INBOUND_EXPENSE_DETAILS] 
                WHERE CURRENCY IS NOT NULL 
                GROUP BY CURRENCY
                ORDER BY CURRENCY
            """)
            currencies = [row[0] for row in cursor.fetchall()]

            # ✅ Get date range (fast because it only scans MIN/MAX)
            cursor.execute("""
                SELECT MIN(DATE), MAX(DATE) 
                FROM   [WHR_INBOUND_EXPENSE_DETAILS]
            """)
            min_date, max_date = cursor.fetchone()

        return Response({
            "status": "success",
            "categories": categories,
            "currencies": currencies,
            "date_range": {
                "min_date": min_date.strftime('%Y-%m-%d') if min_date else None,
                "max_date": max_date.strftime('%Y-%m-%d') if max_date else None
            }
        }, status=status.HTTP_200_OK)

    except Exception as e:
        import traceback
        print("❌ Error fetching filter options:", str(e))
        print(traceback.format_exc())
        return Response({
            "status": "error",
            "message": "Internal server error while fetching filter options."
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@csrf_exempt
def delete_expense(request):
    if request.method == "POST":
        try:
            data = request.POST or json.loads(request.body)
            expense_id = data.get("id")

            if not expense_id:
                return JsonResponse({"error": "ID is required"}, status=400)

            with connections['Inbound_db'].cursor() as cursor:
                # Update flag to 'd' instead of deleting
                cursor.execute("""
                    UPDATE   [WHR_INBOUND_EXPENSE_DETAILS] 
                    SET flag = 'd'
                    WHERE id = %s
                """, [expense_id])

            return JsonResponse({"message": f"Expense with ID {expense_id} marked as deleted successfully"})
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    return JsonResponse({"error": "Invalid request"}, status=405)

# 🔹 Update API by ID
@csrf_exempt
def update_expense(request):
    if request.method == "POST":
        try:
            data = request.POST or json.loads(request.body)
            expense_id = data.get("id")

            if not expense_id:
                return JsonResponse({"error": "ID is required"}, status=400)

            update_fields = {
                "BAYAN_NO": data.get("bayan_no"),
                "EXPENSE_CAT": data.get("expense_cat"),
                "NAME": data.get("name"),
                "CURRENCY": data.get("currency"),
                "AMOUNT": data.get("amount"),
                "CONVERSION_RATE": data.get("conversion_rate"),
                "FINAL_AMOUNT": data.get("final_amount"),
                "DATE": data.get("date"),
                "REMARKS": data.get("remarks"),
            }

            set_clause = []
            values = []
            for col, val in update_fields.items():
                if val is not None:  # only update provided fields
                    set_clause.append(f"{col} = %s")
                    values.append(val)

            if not set_clause:
                return JsonResponse({"error": "No fields to update"}, status=400)

            values.append(expense_id)  # WHERE condition

            query = f"""
                UPDATE WHR_INBOUND_EXPENSE_DETAILS
                SET {', '.join(set_clause)}
                WHERE id = %s
            """

            with connections['Inbound_db'].cursor() as cursor:
                cursor.execute(query, values)

            return JsonResponse({"message": f"Expense with ID {expense_id} updated successfully"})
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    return JsonResponse({"error": "Invalid request"}, status=405)



def getPendingInvoice_for_salesman(request, salesman_name):
    with connections['default'].cursor() as cursor:
        cursor.execute("""
            SELECT
                INVOICE_NUMBER,
                SUM(ISNULL(QUANTITY, 0)) AS Total_Quantity,
                SUM(ISNULL(DISPATCH_QTY, 0)) AS Total_Dispatch_Qty,
                SUM(ISNULL(RETURN_QTY, 0)) AS Total_Return_Qty,
                -- Calculate Dispatch + Return as Final_Request_Qty
                SUM(ISNULL(DISPATCH_QTY, 0)) + SUM(ISNULL(RETURN_QTY, 0)) AS Final_Request_Qty,
                -- Calculate Quantity - (Dispatch + Return) as Final_Qty
                SUM(ISNULL(QUANTITY, 0)) - (SUM(ISNULL(DISPATCH_QTY, 0)) + SUM(ISNULL(RETURN_QTY, 0))) AS Final_Qty
            FROM
                [BUYP].[BUYP].[XXALJE_UNDELIVERED_DATA_BUYP1]
            WHERE
                SALESMAN_NAME = %s
            GROUP BY
                INVOICE_NUMBER
            HAVING
                SUM(ISNULL(QUANTITY, 0)) - (SUM(ISNULL(DISPATCH_QTY, 0)) + SUM(ISNULL(RETURN_QTY, 0))) > 0
            ORDER BY
                INVOICE_NUMBER;
        """, [salesman_name])
       
        # Fetch all results
        results = cursor.fetchall()
       
        # Prepare the response data
        pending_invoices = []
        for row in results:
            pending_invoices.append({
                "invoice_number": row[0],
                "total_quantity": row[1],
                "total_dispatch_qty": row[2],
                "total_return_qty": row[3],
                "final_request_qty": row[4],
                "final_qty": row[5],
            })
       
        # Count of pending invoices
        pending_invoice_count = len(pending_invoices)
 
    return JsonResponse({
        "salesman_name": salesman_name,
        "pendingInvoice_count": pending_invoice_count,
        "pending_invoices": pending_invoices
    })



def getcompleted_dispatches_for_salesman(request, salesman_name):
    with connections['Inbound_db'].cursor() as cursor:
        cursor.execute("""
            SELECT COUNT(DISTINCT INVOICE_NUMBER)
            FROM  WHR_CREATE_DISPATCH
            WHERE TRUCK_SCAN_QTY = 0 AND SALESMAN_NAME = %s
        """, [salesman_name])
        count = cursor.fetchone()[0]

    return JsonResponse({
        "salesman_name": salesman_name,
        "unscanned_invoice_count": count
    })




def get_On_Progress_dispatches_for_salesman(request, salesman_name):
    with connections['Inbound_db'].cursor() as cursor:
        cursor.execute("""
            SELECT COUNT(DISTINCT INVOICE_NUMBER)
            FROM  WHR_CREATE_DISPATCH
            WHERE TRUCK_SCAN_QTY != 0 AND SALESMAN_NAME = %s
        """, [salesman_name])
        count = cursor.fetchone()[0]

    return JsonResponse({
        "salesman_name": salesman_name,
        "scanned_invoice_count": count
    })

def undelivered_customer_count(request, salesman_name):
    with connections['Inbound_db'].cursor() as cursor:
        cursor.execute("""
            SELECT COUNT(DISTINCT CUSTOMER_NAME)
            FROM BUYP.BUYP.XXALJE_UNDELIVERED_DATA_BUYP1
            WHERE SALESMAN_NAME = %s
        """, [salesman_name])
        customer_count = cursor.fetchone()[0]
 
    return JsonResponse({
        "salesman_name": salesman_name,
        "customer_count": customer_count
    })



@csrf_exempt
@require_http_methods(["POST"])
def save_currency(request):
    try:
        data = json.loads(request.body.decode("utf-8"))
        currency = data.get("Currency")

        if not currency:
            return JsonResponse({"success": False, "error": "Currency is required"}, status=400)

        # Insert into WHR_CURRENCY_DETAILS
        with connections['Inbound_db'].cursor() as cursor:
            cursor.execute("""
                INSERT INTO WHR_CURRENCY_DETAILS (CURRENCY_CODE)
                VALUES (%s)
            """, [currency])

        return JsonResponse({"success": True, "message": "Currency saved successfully!"}, status=201)

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)


@csrf_exempt
@require_http_methods(["GET"])
def get_currencies(request):
    try:
        with connections['Inbound_db'].cursor() as cursor:
            cursor.execute("SELECT id, CURRENCY_CODE FROM WHR_CURRENCY_DETAILS ORDER BY id DESC")
            rows = cursor.fetchall()
            data = [
                {"id": row[0], "code": row[1]}
                for row in rows
            ]
        return JsonResponse({"success": True, "currencies": data}, status=200)
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)


@csrf_exempt
@require_http_methods(["DELETE"])
def delete_currency(request, currency_id):
    try:
        with connections['Inbound_db'].cursor() as cursor:
            cursor.execute("DELETE FROM WHR_CURRENCY_DETAILS WHERE id = %s", [currency_id])

        return JsonResponse({"success": True, "message": "Currency deleted successfully!"}, status=200)

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)





# ✅ 1. API to list all documents

from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework import status
from django.conf import settings
import boto3
from botocore.exceptions import ClientError
from concurrent.futures import ThreadPoolExecutor




class InboundMinioUploadView(APIView):
    parser_classes = [MultiPartParser]
    def post(self, request):
        bayan_no = request.data.get('bayan_no', 'default')
        document_name = request.data.get('document_name', 'Untitled')
        files = request.FILES.getlist('file')
        if not files:
            return Response({"error": "No files provided"}, status=status.HTTP_400_BAD_REQUEST)
        endpoint = f"http{'s' if settings.MINIO_SECURE else ''}://{settings.MINIO_ENDPOINT}"
        # Initialize S3 Client (boto3 client is thread-safe)
        s3 = boto3.client(
            's3',
            endpoint_url=endpoint,
            aws_access_key_id=settings.MINIO_ACCESS_KEY,
            aws_secret_access_key=settings.MINIO_SECRET_KEY,
        )
        uploaded_files = []
        
        # Prepare arguments for parallel execution
        # Normalized folder name
        folder_name = f"bayan:{bayan_no}-{document_name}".replace(" ", "_")
        root_folder = "all_bayan"
        
        def upload_single_file(file_obj):
            """
            Helper function to upload a single file to MinIO.
            Returns a dictionary with result info or error.
            """
            key = f"{root_folder}/{folder_name}/{file_obj.name}"
            try:
                # Upload to MinIO
                s3.upload_fileobj(
                    Fileobj=file_obj,
                    Bucket=settings.MINIO_INBOUND_BUCKET_NAME,
                    Key=key,
                    ExtraArgs={'ACL': 'public-read'}
                )
                file_url = f"/{settings.MINIO_INBOUND_BUCKET_NAME}/{key}"
                return {
                    "status": "success",
                    "file_obj": file_obj, # Return obj to access name later if needed
                    "file_url": file_url,
                    "document_name": document_name # Pass context
                }
            except Exception as e:
                return {
                    "status": "error",
                    "file_name": file_obj.name,
                    "error": str(e)
                }
        # 1. PARALLEL UPLOAD TO MINIO
        # Adjust max_workers as needed (e.g., 5-10)
        with ThreadPoolExecutor(max_workers=5) as executor:
            results = list(executor.map(upload_single_file, files))
        # 2. SEQUENTIAL DB INSERT (Main Thread)
        for result in results:
            if result["status"] == "success":
                try:
                    # Save to MSSQL
                    record = BayanDocument.objects.create(
                        document_name=result["document_name"],
                        file_path=result["file_url"]
                    )
                    
                    uploaded_files.append({
                        "id": record.id,
                        "document_name": record.document_name,
                        "file_path": record.file_path,
                        "upload_date": record.upload_date
                    })
                except Exception as e:
                    # Handle DB error
                     uploaded_files.append({"name": result["file_obj"].name, "error": f"DB Error: {str(e)}"})
            else:
                # Handle Upload error
                uploaded_files.append({"name": result["file_name"], "error": result["error"]})
        return Response({"uploaded_files": uploaded_files}, status=201)


class BayanDocumentsListView(APIView):
    def get(self, request, bayan_no):
        # Get all documents for this bayan number
        documents = BayanDocument.objects.filter(file_path__contains=f"bayan:{bayan_no}-")
        
        # Group by document_name
        grouped_docs = {}
        for doc in documents:
            # Extract document_name from file_path or use the stored document_name
            if doc.document_name not in grouped_docs:
                grouped_docs[doc.document_name] = []
            grouped_docs[doc.document_name].append({
                'id': doc.id,
                'file_path': doc.file_path,
                'upload_date': doc.upload_date,
                'document_name': doc.document_name
            })
        
        return Response({"bayan_no": bayan_no, "documents": grouped_docs})

class BayanDocumentDownloadView(APIView):
    def get(self, request, document_id):
        try:
            document = BayanDocument.objects.get(id=document_id)
            
            # Generate presigned URL for download
            endpoint = f"http{'s' if settings.MINIO_SECURE else ''}://{settings.MINIO_ENDPOINT}"
            
            s3 = boto3.client(
                's3',
                endpoint_url=endpoint,
                aws_access_key_id=settings.MINIO_ACCESS_KEY,
                aws_secret_access_key=settings.MINIO_SECRET_KEY,
            )
            
            # Extract the key from file_path (remove bucket name)
            key = document.file_path.replace(f"/{settings.MINIO_INBOUND_BUCKET_NAME}/", "")
            
            # Generate presigned URL (valid for 1 hour)
            presigned_url = s3.generate_presigned_url(
                'get_object',
                Params={
                    'Bucket': settings.MINIO_INBOUND_BUCKET_NAME,
                    'Key': key
                },
                ExpiresIn=3600
            )
            
            return Response({"download_url": presigned_url})
            
        except BayanDocument.DoesNotExist:
            return Response({"error": "Document not found"}, status=404)
        except Exception as e:
            return Response({"error": str(e)}, status=500)


class BayanDocumentPreviewView(APIView):
    def get(self, request, document_id):
        try:
            document = BayanDocument.objects.get(id=document_id)

            endpoint = f"http{'s' if settings.MINIO_SECURE else ''}://{settings.MINIO_ENDPOINT}"

            s3 = boto3.client(
                "s3",
                endpoint_url=endpoint,
                aws_access_key_id=settings.MINIO_ACCESS_KEY,
                aws_secret_access_key=settings.MINIO_SECRET_KEY,
            )

            key = document.file_path.replace(f"/{settings.MINIO_INBOUND_BUCKET_NAME}/", "")

            obj = s3.get_object(Bucket=settings.MINIO_INBOUND_BUCKET_NAME, Key=key)
            file_stream = obj["Body"]

            filename = os.path.basename(key)
            ext = filename.lower().split(".")[-1]

            # Special handling for text-based files
            if ext in ["txt", "sql", "log", "csv"]:
                content = file_stream.read()
                try:
                    content = content.decode("utf-8")
                except UnicodeDecodeError:
                    content = content.decode("latin1", errors="ignore")
                return HttpResponse(content, content_type="text/plain")

            # Guess content type for others
            content_type, _ = mimetypes.guess_type(filename)
            if not content_type:
                content_type = "application/octet-stream"

            response = FileResponse(file_stream, content_type=content_type)
            response["X-Frame-Options"] = "ALLOWALL"
            response["Access-Control-Allow-Origin"] = "*"
            response["Content-Disposition"] = f'inline; filename="{filename}"'
            return response

        except BayanDocument.DoesNotExist:
            raise Http404("Document not found")
        except ClientError as e:
            return Response({"error": str(e)}, status=500)


class BayanDocumentDeleteView(APIView):
    def delete(self, request, document_id):
        try:
            document = BayanDocument.objects.get(id=document_id)

            # MinIO client
            endpoint = f"http{'s' if settings.MINIO_SECURE else ''}://{settings.MINIO_ENDPOINT}"
            s3 = boto3.client(
                "s3",
                endpoint_url=endpoint,
                aws_access_key_id=settings.MINIO_ACCESS_KEY,
                aws_secret_access_key=settings.MINIO_SECRET_KEY,
            )

            # Key from file_path
            key = document.file_path.replace(f"/{settings.MINIO_INBOUND_BUCKET_NAME}/", "")

            # Delete from MinIO
            s3.delete_object(Bucket=settings.MINIO_INBOUND_BUCKET_NAME, Key=key)

            # Delete from DB
            document.delete()

            return Response({"success": f"Document {document_id} deleted successfully"})
        except BayanDocument.DoesNotExist:
            return Response({"error": "Document not found"}, status=404)
        except Exception as e:
            return Response({"error": str(e)}, status=500)

# Expense Categories APIs
@csrf_exempt
@require_http_methods(["GET", "POST"])
def expense_categories(request):
    try:
        if request.method == 'GET':
            with connections['Inbound_db'].cursor() as cursor:
                cursor.execute("SELECT ID, EXPENSE_CAT FROM WHR_INBOUND_EXPENSE_CAT ORDER BY ID")
                rows = cursor.fetchall()
                
                categories = []
                for row in rows:
                    categories.append({
                        'ID': row[0],
                        'EXPENSE_CAT': row[1]
                    })
                
                return JsonResponse(categories, safe=False)
                
        elif request.method == 'POST':
            data = json.loads(request.body)
            expense_cat = data.get('EXPENSE_CAT')
            
            if not expense_cat:
                return JsonResponse({'error': 'EXPENSE_CAT is required'}, status=400)
            
            with connections['Inbound_db'].cursor() as cursor:
                cursor.execute(
                    "INSERT INTO WHR_INBOUND_EXPENSE_CAT (EXPENSE_CAT) OUTPUT INSERTED.ID VALUES (%s)", 
                    [expense_cat]
                )
                new_id = cursor.fetchone()[0]
                
            return JsonResponse({'ID': new_id, 'EXPENSE_CAT': expense_cat}, status=201)
            
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["GET", "PUT", "DELETE"])
def expense_category_detail(request, id):
    try:
        if request.method == 'GET':
            with connections['Inbound_db'].cursor() as cursor:
                cursor.execute(
                    "SELECT ID, EXPENSE_CAT FROM WHR_INBOUND_EXPENSE_CAT WHERE ID = %s", 
                    [id]
                )
                row = cursor.fetchone()
                
                if row:
                    category = {
                        'ID': row[0],
                        'EXPENSE_CAT': row[1]
                    }
                    return JsonResponse(category)
                else:
                    return JsonResponse({'error': 'Category not found'}, status=404)
                    
        elif request.method == 'PUT':
            data = json.loads(request.body)
            expense_cat = data.get('EXPENSE_CAT')
            
            if not expense_cat:
                return JsonResponse({'error': 'EXPENSE_CAT is required'}, status=400)
            
            with connections['Inbound_db'].cursor() as cursor:
                cursor.execute(
                    "UPDATE WHR_INBOUND_EXPENSE_CAT SET EXPENSE_CAT = %s WHERE ID = %s", 
                    [expense_cat, id]
                )
                
                if cursor.rowcount == 0:
                    return JsonResponse({'error': 'Category not found'}, status=404)
                    
            return JsonResponse({'ID': id, 'EXPENSE_CAT': expense_cat})
            
        elif request.method == 'DELETE':
            with connections['Inbound_db'].cursor() as cursor:
                # Check if category is used in expense names
                cursor.execute(
                    "SELECT COUNT(*) FROM WHR_INBOUND_EXPENSE_CAT_NAME WHERE EXPENSE_CAT = %s", 
                    [id]
                )
                count = cursor.fetchone()[0]
                
                if count > 0:
                    return JsonResponse(
                        {'error': 'Cannot delete category. It is being used by expense names.'}, 
                        status=400
                    )
                
                cursor.execute(
                    "DELETE FROM WHR_INBOUND_EXPENSE_CAT WHERE ID = %s", 
                    [id]
                )
                
                if cursor.rowcount == 0:
                    return JsonResponse({'error': 'Category not found'}, status=404)
                    
            return JsonResponse({'message': 'Category deleted successfully'})
            
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
@require_http_methods(["GET", "POST"])
def expense_names(request):
    try:
        print(f"=== EXPENSE NAMES REQUEST ===")
        print(f"Method: {request.method}")
        
        if request.method == 'GET':
            print("Processing ALL expense names")
            
            with connections['Inbound_db'].cursor() as cursor:
                # Query to get expense names with their category names
                cursor.execute("""
                    SELECT 
                        ID, 
                        EXPENSE_CAT,  -- This will now be the category name
                        NAME 
                    FROM WHR_INBOUND_EXPENSE_CAT_NAME 
                    ORDER BY ID
                """)
                
                rows = cursor.fetchall()
                print(f"Found {len(rows)} expense names")
                
                expense_names = []
                for row in rows:
                    expense_names.append({
                        'ID': row[0],
                        'EXPENSE_CAT': row[1],  # This is the category name
                        'NAME': row[2],
                        'CATEGORY_NAME': row[1]  # Same as EXPENSE_CAT since we're storing names
                    })
                
                return JsonResponse(expense_names, safe=False)
                
        elif request.method == 'POST':
            data = json.loads(request.body)
            expense_cat = data.get('EXPENSE_CAT')  # This will be the category name
            name = data.get('NAME')
            
            if not expense_cat or not name:
                return JsonResponse({'error': 'EXPENSE_CAT and NAME are required'}, status=400)
            
            with connections['Inbound_db'].cursor() as cursor:
                # Insert new expense name with the category NAME (not ID)
                cursor.execute(
                    "INSERT INTO WHR_INBOUND_EXPENSE_CAT_NAME (EXPENSE_CAT, NAME) OUTPUT INSERTED.ID VALUES (%s, %s)", 
                    [expense_cat, name]  # Save the category name directly
                )
                new_id = cursor.fetchone()[0]
                
            return JsonResponse({
                'ID': new_id, 
                'EXPENSE_CAT': expense_cat,  # Return the category name
                'NAME': name,
                'CATEGORY_NAME': expense_cat  # Same as EXPENSE_CAT
            }, status=201)
            
    except Exception as e:
        print(f"Error in expense_names: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)
    
        
@csrf_exempt
@require_http_methods(["GET", "PUT", "DELETE"])
def expense_name_detail(request, id):
    try:
        if request.method == 'GET':
            with connections['Inbound_db'].cursor() as cursor:
                cursor.execute("""
                    SELECT en.ID, en.EXPENSE_CAT, en.NAME, ec.EXPENSE_CAT as CATEGORY_NAME 
                    FROM WHR_INBOUND_EXPENSE_CAT_NAME en
                    INNER JOIN WHR_INBOUND_EXPENSE_CAT ec ON en.EXPENSE_CAT = ec.ID
                    WHERE en.ID = %s
                """, [id])
                row = cursor.fetchone()
                
                if row:
                    expense_name = {
                        'ID': row[0],
                        'EXPENSE_CAT': row[1],
                        'NAME': row[2],
                        'CATEGORY_NAME': row[3]
                    }
                    return JsonResponse(expense_name)
                else:
                    return JsonResponse({'error': 'Expense name not found'}, status=404)
                    
        elif request.method == 'PUT':
            data = json.loads(request.body)
            expense_cat = data.get('EXPENSE_CAT')
            name = data.get('NAME')
            
            if not expense_cat or not name:
                return JsonResponse({'error': 'EXPENSE_CAT and NAME are required'}, status=400)
            
            with connections['Inbound_db'].cursor() as cursor:
                # Verify category exists
                cursor.execute(
                    "SELECT ID FROM WHR_INBOUND_EXPENSE_CAT WHERE ID = %s", 
                    [expense_cat]
                )
                if not cursor.fetchone():
                    return JsonResponse({'error': 'Category not found'}, status=400)
                
                cursor.execute(
                    "UPDATE WHR_INBOUND_EXPENSE_CAT_NAME SET EXPENSE_CAT = %s, NAME = %s WHERE ID = %s", 
                    [expense_cat, name, id]
                )
                
                if cursor.rowcount == 0:
                    return JsonResponse({'error': 'Expense name not found'}, status=404)
                
                # Get category name for response
                cursor.execute(
                    "SELECT EXPENSE_CAT FROM WHR_INBOUND_EXPENSE_CAT WHERE ID = %s", 
                    [expense_cat]
                )
                category_name = cursor.fetchone()[0]
                
            return JsonResponse({
                'ID': id, 
                'EXPENSE_CAT': expense_cat, 
                'NAME': name,
                'CATEGORY_NAME': category_name
            })
            
        elif request.method == 'DELETE':
            with connections['Inbound_db'].cursor() as cursor:
                cursor.execute(
                    "DELETE FROM WHR_INBOUND_EXPENSE_CAT_NAME WHERE ID = %s", 
                    [id]
                )
                
                if cursor.rowcount == 0:
                    return JsonResponse({'error': 'Expense name not found'}, status=404)
                    
            return JsonResponse({'message': 'Expense name deleted successfully'})
            
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

import io
from urllib.parse import quote
import boto3
from botocore.client import Config
import openpyxl
from django.conf import settings
from django.db import connection, transaction
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
from rest_framework import status
from .models import InboundContainerBayanDocument

class ExcelContainerUploadView(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request):
        try:
            # 📌 Required Parameters
            bayan_no = request.data.get("bayan_no") or ""
            doc_no = request.data.get("doc_no")
            initiator_no = request.data.get("initiator_no")
            initiator_name = request.data.get("initiator_name")

            missing = [k for k, v in {
                "doc_no": doc_no,
                "initiator_no": initiator_no,
                "initiator_name": initiator_name
            }.items() if not v]

            if missing:
                return Response(
                    {"error": f"Missing fields: {', '.join(missing)}"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # 📌 Files
            excel_file = request.FILES.get("excel_file")
            doc_files = request.FILES.getlist("documents")

            if not excel_file:
                return Response({"error": "Excel file is required"}, status=status.HTTP_400_BAD_REQUEST)

            # ✅ MinIO client
            endpoint = f"http{'s' if settings.MINIO_SECURE else ''}://{settings.MINIO_ENDPOINT}"
            s3 = boto3.client(
                "s3",
                endpoint_url=endpoint,
                aws_access_key_id=settings.MINIO_ACCESS_KEY,
                aws_secret_access_key=settings.MINIO_SECRET_KEY,
                config=Config(signature_version="s3v4"),
            )

            # 📌 Creator Info
            created_by = request.user.username if request.user.is_authenticated else initiator_no
            created_ip = request.META.get("REMOTE_ADDR", "0.0.0.0")

            # 📌 Upload Excel to MinIO
            excel_safe_name = quote(excel_file.name)
            excel_key = f"AllContainerInfo/Document:{doc_no}/{excel_safe_name}"
            excel_bytes = excel_file.read()
            s3.upload_fileobj(io.BytesIO(excel_bytes), settings.MINIO_INBOUND_CONTAINER_BUCKET_NAME, excel_key)
            excel_url = f"/inboundalje/{excel_key}"

            # 📌 Save Excel metadata in DB
            InboundContainerBayanDocument.objects.using('Inbound_db').create(
                DOC_NO=doc_no,
                BAYAN_NO=bayan_no,
                DOCUMENT_NAME=excel_safe_name,
                FILE_PATH=excel_url,
                INITIATOR_NO=initiator_no,
                INITIATOR_NAME=initiator_name,
                CREATED_BY=created_by,
                CREATED_IP=created_ip,
                UPLOAD_DATE=timezone.now()
            )

            # 📌 Upload Supporting Documents
            for file_obj in doc_files:
                safe_name = quote(file_obj.name)
                key = f"all_bayan/bayan:{bayan_no}-Document/{safe_name}"
                s3.upload_fileobj(file_obj, settings.MINIO_INBOUND_CONTAINER_BUCKET_NAME, key)
                file_url = f"/inboundalje/{key}"

                InboundContainerBayanDocument.objects.using('Inbound_db').create(
                    DOC_NO=doc_no,
                    BAYAN_NO=bayan_no,
                    DOCUMENT_NAME=safe_name,
                    FILE_PATH=file_url,
                    INITIATOR_NO=initiator_no,
                    INITIATOR_NAME=initiator_name,
                    CREATED_BY=created_by,
                    CREATED_IP=created_ip,
                    UPLOAD_DATE=timezone.now()
                )

            # 📌 Parse Excel for container & serial rows
            wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
            
            # Get sheet names
            sheet_names = wb.sheetnames
            print(f"Sheet names in Excel: {sheet_names}")
            
            if len(sheet_names) >= 2:
                container_sheet = wb[sheet_names[0]]  # First sheet
                serial_sheet = wb[sheet_names[1]]     # Second sheet
            else:
                return Response({"error": "Excel must contain at least 2 sheets"}, 
                               status=status.HTTP_400_BAD_REQUEST)

            # --- Container rows (Sheet1) ---
            container_values = []
            for row_idx, row in enumerate(container_sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue
                
                try:
                    # Clean and format values
                    po_number = str(row[0]).strip() if row[0] else ''
                    container_no = str(row[1]).strip() if row[1] else ''
                    seal_no = str(row[2]).strip() if row[2] else ''
                    model_no = str(row[3]).strip() if row[3] else ''
                    hs_code = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                    qty_1 = float(row[5]) if len(row) > 5 and row[5] is not None else 0.0
                    measure_1 = str(row[6]).strip() if len(row) > 6 and row[6] else ''
                    qty_2 = float(row[7]) if len(row) > 7 and row[7] is not None else 0.0
                    measure_2 = str(row[8]).strip() if len(row) > 8 and row[8] else ''
                    mw = float(row[9]) if len(row) > 9 and row[9] is not None else 0.0
                    gw = float(row[10]) if len(row) > 10 and row[10] is not None else 0.0
                    kgs = float(row[11]) if len(row) > 11 and row[11] is not None else 0.0
                    
                    # Build SQL values string
                    values = f"""('{doc_no}', '{bayan_no}', GETDATE(), '{po_number}', 
                                 '{container_no}', '{seal_no}', '{model_no}', '{hs_code}', 
                                 {qty_1}, '{measure_1}', {qty_2}, '{measure_2}', 
                                 {mw}, {gw}, {kgs}, 
                                 '{initiator_no}', '{initiator_name}', 
                                 GETDATE(), '{created_by}', '{created_ip}')"""
                    container_values.append(values)
                    
                except (ValueError, TypeError, IndexError) as e:
                    print(f"Error parsing container row {row_idx}: {e}")
                    continue

            # --- Serial rows (Sheet2) ---
            serial_values = []
            for row_idx, row in enumerate(serial_sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue
                
                try:
                    model_no = str(row[0]).strip() if row[0] else ''
                    serial_no = str(row[1]).strip() if row[1] else ''
                    
                    # Build SQL values string
                    values = f"""('{doc_no}', '{bayan_no}', GETDATE(), '', 
                                 '{model_no}', '{serial_no}', 
                                 '{initiator_no}', '{initiator_name}', 
                                 GETDATE(), '{created_by}', '{created_ip}')"""
                    serial_values.append(values)
                    
                except (ValueError, TypeError, IndexError) as e:
                    print(f"Error parsing serial row {row_idx}: {e}")
                    continue

            print(f"Total container rows to insert: {len(container_values)}")
            print(f"Total serial rows to insert: {len(serial_values)}")

            # 📌 Bulk Insert into container/serial tables using raw SQL
            with transaction.atomic(using='Inbound_db'):
                with connections['Inbound_db'].cursor() as cursor:

                    if serial_values:
                        # Build single INSERT statement for all serial rows
                        serial_sql = f"""
                            INSERT INTO WHR_INBOUND_BAYAN_SERIALNO_TBL
                            (DOC_NO, BAYAN_NO, DATE, PO_NUMBER,
                             MODEL_NO, SERIAL_NO,
                             INITIATOR_NO, INITIATOR_NAME,
                             CREATION_DATE, CREATED_BY, CREATED_IP)
                            VALUES {','.join(serial_values)}
                        """
                        cursor.execute(serial_sql)
                        print(f"Inserted {len(serial_values)} serial rows")

            return Response({
                "success": True,
                "excel_file_url": excel_url,
                "inserted_containers": len(container_values),
                "inserted_serials": len(serial_values),
                "inserted_documents": 1 + len(doc_files),
                "doc_no": doc_no,
                "bayan_no": bayan_no
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"Error in ExcelContainerUploadView: {e}")
            print(f"Traceback: {error_details}")
            return Response({"error": str(e), "details": error_details}, 
                           status=status.HTTP_400_BAD_REQUEST)
        
        
class ExcelContainerReportView(APIView):

    # =========================================================
    # 🔁 COMMON PAGINATION HELPER
    # =========================================================
    def _fetch_paginated(
        self,
        table,
        key,
        offset,
        limit,
        bayan_no=None,
        doc_no=None,
        select_fields="*"
    ):
        where_clause = "WHERE 1=1"
        params = []

        if bayan_no:
            where_clause += " AND BAYAN_NO = %s"
            params.append(bayan_no)

        if doc_no:
            where_clause += " AND DOC_NO = %s"
            params.append(doc_no)

        with connections['Inbound_db'].cursor() as cursor:
            # 🔢 TOTAL COUNT
            count_query = f"""
                SELECT COUNT(*)
                FROM   [{table}]
                {where_clause}
            """
            cursor.execute(count_query, params)
            total = cursor.fetchone()[0]

            # 📄 DATA WITH PAGINATION
            data_query = f"""
                SELECT {select_fields}
                FROM   [{table}]
                {where_clause}
                ORDER BY {key} DESC
                OFFSET %s ROWS FETCH NEXT %s ROWS ONLY
            """
            cursor.execute(data_query, params + [offset, limit])

            columns = [col[0] for col in cursor.description]
            rows = [dict(zip(columns, row)) for row in cursor.fetchall()]

        return total, rows

    # =========================================================
    # 📥 GET API
    # =========================================================
    def get(self, request):
        try:
            # ---------------- REQUEST PARAMS ----------------
            bayan_no = request.GET.get("bayan_no", "").strip()
            doc_no = request.GET.get("doc_no", "").strip()

            limit = int(request.GET.get("limit", 10))
            container_offset = int(request.GET.get("container_offset", 0))
            serial_offset = int(request.GET.get("serial_offset", 0))
            po_info_offset = int(request.GET.get("po_info_offset", 0))

            if not bayan_no and not doc_no:
                return Response(
                    {"success": False, "error": "Either bayan_no or doc_no is required"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # =================================================
            # 🧾 HEADER INFO
            # =================================================
            po_header_info = {}
            with connections['Inbound_db'].cursor() as cursor:
                query = """
                    SELECT TOP 1
                        DOC_NO, BAYAN_NO, MASTER_PL, HOUSE_PL, BILL_NO,
                        BILL_DUE_DATE, BAYAN_DATE, BL_DATE, ETD_DATE, ETA_DATE,
                        CONTAINER_COUNT, CLEARANCE_DATE,
                        SUPPLIER_NO, SUPPLIER_NAME, SUPPLIER_SITE,
                        POL, POD, LINE_NAME,
                        INVOICE_NO1, LC_NO1, VALUE1, PAYMENT_TERM1, INCOTERMS1,
                        INVOICE_NO2, LC_NO2, VALUE2, PAYMENT_TERMS2, INCOTERMS2,
                        INVOICE_NO3, LC_NO3, VALUE3, PAYMENT_TERM3, INCOTERMS3,
                        INITIATOR_NO, INITIATOR_NAME, DATE,
                        GRN, GRN_DATE,RECEIPT_DATE ,
                        STATUS  -- <--- YOU MUST ADD THIS COLUMN !!!-- Added GRN fields here if they exist in header too, just in case
                    FROM   [WHR_INBOUND_BAYAN_HEADER_TBL]
                    WHERE 1=1
                """
                params = []

                if bayan_no:
                    query += " AND BAYAN_NO = %s"
                    params.append(bayan_no)

                if doc_no:
                    query += " AND DOC_NO = %s"
                    params.append(doc_no)

                cursor.execute(query, params)
                row = cursor.fetchone()

                if row:
                    columns = [col[0] for col in cursor.description]
                    po_header_info = dict(zip(columns, row))

            # =================================================
            # 📎 DOCUMENTS
            # =================================================
            documents = InboundContainerBayanDocument.objects.using('Inbound_db').all()
            if bayan_no:
                documents = documents.filter(BAYAN_NO=bayan_no)
            if doc_no:
                documents = documents.filter(DOC_NO=doc_no)

            document_list = list(documents.values(
                "id", "DOC_NO", "BAYAN_NO", "DOCUMENT_NAME", "FILE_PATH",
                "INITIATOR_NO", "INITIATOR_NAME",
                "CREATED_BY", "CREATED_IP", "UPLOAD_DATE"
            ))

            # =================================================
            # 📦 CONTAINERS
            # =================================================
            container_total, container_data = self._fetch_paginated(
                table="WHR_INBOUND_BAYAN_DETAILS_TBL",
                key="PRODUCT_ID",
                offset=container_offset,
                limit=limit,
                bayan_no=bayan_no,
                doc_no=doc_no,
                select_fields="""
                    PRODUCT_ID, DOC_NO, BAYAN_NO, CONTAINER_NO, SEAL_NO,
                    ITEM_CODE, HS_CODE,
                    QTY_1, MEASURE_1, QTY_2, MEASURE_2,
                    MW, GW, KGS,
                    INITIATOR_NO, INITIATOR_NAME, CREATION_DATE
                """
            )

            # =================================================
            # 🔢 SERIAL NUMBERS
            # =================================================
            serial_total, serial_data = self._fetch_paginated(
                table="WHR_INBOUND_BAYAN_SERIALNO_TBL",
                key="ID",
                offset=serial_offset,
                limit=limit,
                bayan_no=bayan_no,
                doc_no=doc_no,
                select_fields="""
                    ID, DOC_NO, BAYAN_NO, MODEL_NO, SERIAL_NO,
                    INITIATOR_NO, INITIATOR_NAME, CREATION_DATE
                """
            )

            # =================================================
            # 🧾 PO INFO (UPDATED)
            # =================================================
            po_info_total, po_info_data = self._fetch_paginated(
                table="WHR_INBOUND_BAYAN_DETAILS_TBL",
                key="PRODUCT_ID",
                offset=po_info_offset,
                limit=limit,
                bayan_no=bayan_no,
                doc_no=doc_no,
                select_fields="""
                    PRODUCT_ID, DOC_NO, BAYAN_NO, PO_NUMBER,
                    FRANCHISE, CLASS, SUBCLASS,
                    ITEM_CODE, ITEM_DESC,
                    PO_QTY, REC_QTY, BALANCE_QTY,
                    SHIPPED_QTY, CONTAINER_NO,
                    STATUS -- <--- ADDED STATUS FIELD HERE
                """
            )

            # =================================================
            # ✅ FINAL RESPONSE
            # =================================================
            return Response({
                "success": True,
                "po_header_info": po_header_info,
                "documents": document_list,
                "containers": container_data,
                "serials": serial_data,
                "po_info": po_info_data,
                "total_counts": {
                    "containers": container_total,
                    "serials": serial_total,
                    "poInfo": po_info_total
                },
                "limit": limit,
                "container_offset": container_offset,
                "serial_offset": serial_offset,
                "po_info_offset": po_info_offset
            }, status=status.HTTP_200_OK)

        except Exception as e:
            import traceback
            print(traceback.format_exc())
            return Response(
                {"success": False, "error": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
class UpdatePOHeaderView(APIView):
    def post(self, request):
        try:
            data = request.data
            
            # Get the identifier (bayan_no or doc_no)
            bayan_no = data.get('bayan_no', '').strip()
            doc_no = data.get('doc_no', '').strip()
            
            if not bayan_no and not doc_no:
                return Response(
                    {"success": False, "error": "Either bayan_no or doc_no is required"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # List of fields that cannot be updated
            non_editable_fields = ['DOC_NO', 'SUPPLIER_NAME', 'SUPPLIER_SITE']
            
            # List of valid fields that can be updated
            valid_fields = [
                'BAYAN_NO', 'MASTER_PL', 'HOUSE_PL', 'BILL_NO', 'BILL_DUE_DATE', 'BAYAN_DATE',
                'BL_DATE', 'ETD_DATE', 'ETA_DATE', 'CONTAINER_COUNT', 'CLEARANCE_DATE',
                'SUPPLIER_NO', 'POL', 'POD', 'LINE_NAME', 'INVOICE_NO1', 'LC_NO1',
                'VALUE1', 'PAYMENT_TERM1', 'INCOTERMS1', 'INVOICE_NO2', 'LC_NO2',
                'VALUE2', 'PAYMENT_TERMS2', 'INCOTERMS2', 'INVOICE_NO3', 'LC_NO3',
                'VALUE3', 'PAYMENT_TERM3', 'INCOTERMS3', 'INITIATOR_NO', 'INITIATOR_NAME',
                'DATE',
                'GRN', 'GRN_DATE', 'RECEIPT_DATE'
            ]
            
            # -- REFINED LOGIC: Check actual presence of values --
            
            # Helper to check if value exists
            def has_value(key):
                val = data.get(key)
                return val is not None and str(val).strip() != ''

            is_bayan_update = 'BAYAN_NO' in data
            is_grn_update = has_value('GRN') or has_value('GRN_DATE')
            is_receipt_update = 'RECEIPT_DATE' in data
            has_receipt_date = has_value('RECEIPT_DATE')
            is_clearance_update = 'CLEARANCE_DATE' in data
            has_clearance_date = has_value('CLEARANCE_DATE')
            
            new_bayan_no = data.get('BAYAN_NO')
            clearance_date = data.get('CLEARANCE_DATE')
            receipt_date = data.get('RECEIPT_DATE')
            
            with transaction.atomic(using='Inbound_db'):
                with connections['Inbound_db'].cursor() as cursor:
                    # Fetch current record state to check for actual changes and get DOC_NO
                    curr_db_bayan = None
                    curr_db_clearance = None
                    curr_db_status = None
                    actual_doc_no = doc_no
                    
                    if actual_doc_no:
                        cursor.execute(
                            "SELECT BAYAN_NO, CLEARANCE_DATE, STATUS FROM [WHR_INBOUND_BAYAN_HEADER_TBL] WHERE DOC_NO = %s", 
                            [actual_doc_no]
                        )
                    elif bayan_no:
                        cursor.execute(
                            "SELECT DOC_NO, BAYAN_NO, CLEARANCE_DATE, STATUS FROM [WHR_INBOUND_BAYAN_HEADER_TBL] WHERE BAYAN_NO = %s", 
                            [bayan_no]
                        )
                    
                    row = cursor.fetchone()
                    if row:
                        if actual_doc_no:
                            curr_db_bayan, curr_db_clearance, curr_db_status = row
                        else:
                            actual_doc_no, curr_db_bayan, curr_db_clearance, curr_db_status = row

                    # Helper to check if a value is actually changing compared to DB
                    def is_changing(new_val, db_val):
                        if new_val is None: return False
                        n = str(new_val).strip()
                        o = str(db_val or '').strip()
                        return n != '' and n != o

                    is_new_bayan = is_changing(new_bayan_no, curr_db_bayan)
                    is_new_clearance = is_changing(clearance_date, curr_db_clearance)
                    
                    # === REFINED ROBUST UPDATE LOGIC ===
                    # 1. Collect all valid fields to update
                    update_fields_map = {}
                    for field, value in data.items():
                        if field in valid_fields and field not in non_editable_fields:
                            if has_value(field):
                                update_fields_map[field] = value
                    
                    if not update_fields_map:
                        return Response(
                            {"success": False, "error": "No valid fields to update"},
                            status=status.HTTP_400_BAD_REQUEST
                        )

                    # 2. Determine implied status transition
                    # Workflow: In Transit -> Under Clearance -> WH Transportation -> Pending Offloading -> Pending GRN -> Completed
                    status_order = {
                        'None': 0, '': 0, 'In Transit': 1, 'Under Clearance': 2, 
                        'WH Transportation': 3, 'Pending Offloading': 4, 
                        'Pending GRN': 5, 'Completed': 6
                    }
                    
                    current_status = curr_db_status or ''
                    current_rank = status_order.get(current_status, 0)
                    
                    target_status = None
                    target_rank = 0
                    
                    # Logic: What is the most advanced milestone we have data for?
                    if has_value('GRN') or has_value('GRN_DATE'):
                        target_status = 'Completed'
                    elif has_value('RECEIPT_DATE'):
                        target_status = 'Pending GRN'
                    elif has_value('CLEARANCE_DATE'):
                        target_status = 'WH Transportation'
                    elif is_new_bayan:
                        target_status = 'Under Clearance'
                        
                    if target_status:
                        target_rank = status_order.get(target_status, 0)
                    
                    # 3. Only change status if it's a forward progression
                    if target_rank > current_rank:
                        update_fields_map['STATUS'] = target_status
                    else:
                        target_status = current_status # Keep existing

                    # 4. Construct and execute the Header update
                    set_clauses = [f"{field} = %s" for field in update_fields_map.keys()]
                    header_params = list(update_fields_map.values())
                    
                    where_clause = "WHERE DOC_NO = %s"
                    header_params.append(actual_doc_no)
                    
                    query = f"UPDATE [WHR_INBOUND_BAYAN_HEADER_TBL] SET {', '.join(set_clauses)} {where_clause}"
                    cursor.execute(query, header_params)
                    
                    if cursor.rowcount == 0:
                        return Response(
                            {"success": False, "error": "No record found to update"},
                            status=status.HTTP_404_NOT_FOUND
                        )

                    # 5. SIDE EFFECTS: Sync BAYAN_NO and STATUS
                    if is_new_bayan:
                        # Propagate BAYAN_NO to details and serials
                        cursor.execute(
                            "UPDATE [WHR_INBOUND_BAYAN_DETAILS_TBL] SET BAYAN_NO = %s WHERE DOC_NO = %s",
                            [new_bayan_no, actual_doc_no]
                        )
                        cursor.execute(
                            "UPDATE [WHR_INBOUND_BAYAN_SERIALNO_TBL] SET BAYAN_NO = %s WHERE DOC_NO = %s",
                            [new_bayan_no, actual_doc_no]
                        )

                    # Sync status to Details table if it changed
                    if target_status != current_status:
                        cursor.execute(
                            "UPDATE [WHR_INBOUND_BAYAN_DETAILS_TBL] SET STATUS = %s WHERE DOC_NO = %s",
                            [target_status, actual_doc_no]
                        )

                    return Response({
                        "success": True, 
                        "message": f"Updated successfully. Status: {target_status if target_status else current_status}",
                        "updated_fields": list(update_fields_map.keys())
                    })
           
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            return Response(
                {"success": False, "error": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )




class EditContainerDocumentView(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request):
        try:
            # 📌 Required Parameters
            bayan_no = request.data.get("bayan_no")
            doc_no = request.data.get("doc_no")
            initiator_no = request.data.get("initiator_no")
            initiator_name = request.data.get("initiator_name")

            missing = [k for k, v in {
                "bayan_no": bayan_no,
                "doc_no": doc_no,
                "initiator_no": initiator_no,
                "initiator_name": initiator_name
            }.items() if not v]

            if missing:
                return Response(
                    {"error": f"Missing fields: {', '.join(missing)}"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # 📌 Files
            excel_file = request.FILES.get("excel_file")
            doc_files = request.FILES.getlist("documents")

            if not excel_file:
                return Response({"error": "Excel file is required"}, status=status.HTTP_400_BAD_REQUEST)

            # ✅ MinIO client
            endpoint = f"http{'s' if settings.MINIO_SECURE else ''}://{settings.MINIO_ENDPOINT}"
            s3 = boto3.client(
                "s3",
                endpoint_url=endpoint,
                aws_access_key_id=settings.MINIO_ACCESS_KEY,
                aws_secret_access_key=settings.MINIO_SECRET_KEY,
                config=Config(signature_version="s3v4"),
            )

            # 📌 Creator Info
            created_by = request.user.username if request.user.is_authenticated else initiator_no
            created_ip = request.META.get("REMOTE_ADDR", "0.0.0.0")

            # 📌 Delete existing data
            with transaction.atomic(using='Inbound_db'):
                # Delete existing container records
                with connections['Inbound_db'].cursor() as cursor:
                    cursor.execute(
                        "DELETE FROM WHR_INBOUND_BAYAN_CONTAINER_TBL WHERE DOC_NO = %s AND BAYAN_NO = %s",
                        [doc_no, bayan_no]
                    )
                
                # Delete existing serial records
                with connections['Inbound_db'].cursor() as cursor:
                    cursor.execute(
                        "DELETE FROM WHR_INBOUND_BAYAN_SERIALNO_TBL WHERE DOC_NO = %s AND BAYAN_NO = %s",
                        [doc_no, bayan_no]
                    )
                
                # Get existing documents to delete from MinIO
                existing_docs = InboundContainerBayanDocument.objects.using('Inbound_db').filter(
                    DOC_NO=doc_no, 
                    BAYAN_NO=bayan_no
                )
                
                # Delete files from MinIO
                for doc in existing_docs:
                    try:
                        file_key = doc.FILE_PATH.replace('/inboundalje/', '')
                        file_key = unquote(file_key)
                        s3.delete_object(
                            Bucket=settings.MINIO_INBOUND_CONTAINER_BUCKET_NAME,
                            Key=file_key
                        )
                    except Exception as e:
                        print(f"Error deleting file from MinIO: {e}")
                
                # Delete document records from database
                existing_docs.delete()

            # 📌 Upload new Excel to MinIO
            excel_safe_name = quote(excel_file.name)
            excel_key = f"AllContainerInfo/bayan:{bayan_no}-Document/{excel_safe_name}"
            excel_bytes = excel_file.read()
            s3.upload_fileobj(io.BytesIO(excel_bytes), settings.MINIO_INBOUND_CONTAINER_BUCKET_NAME, excel_key)
            excel_url = f"/inboundalje/{excel_key}"

            # 📌 Save Excel metadata in DB
            InboundContainerBayanDocument.objects.using('Inbound_db').create(
                DOC_NO=doc_no,
                BAYAN_NO=bayan_no,
                DOCUMENT_NAME=excel_safe_name,
                FILE_PATH=excel_url,
                INITIATOR_NO=initiator_no,
                INITIATOR_NAME=initiator_name,
                CREATED_BY=created_by,
                CREATED_IP=created_ip,
                UPLOAD_DATE=timezone.now()
            )

            # 📌 Upload Supporting Documents
            for file_obj in doc_files:
                safe_name = quote(file_obj.name)
                key = f"all_bayan/bayan:{bayan_no}-Document/{safe_name}"
                s3.upload_fileobj(file_obj, settings.MINIO_INBOUND_CONTAINER_BUCKET_NAME, key)
                file_url = f"/inboundalje/{key}"

                InboundContainerBayanDocument.objects.using('Inbound_db').create(
                    DOC_NO=doc_no,
                    BAYAN_NO=bayan_no,
                    DOCUMENT_NAME=safe_name,
                    FILE_PATH=file_url,
                    INITIATOR_NO=initiator_no,
                    INITIATOR_NAME=initiator_name,
                    CREATED_BY=created_by,
                    CREATED_IP=created_ip,
                    UPLOAD_DATE=timezone.now()
                )

            # 📌 Parse Excel for container & serial rows
            wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
            sheet3 = wb.worksheets[2]  # Container sheet
            sheet4 = wb.worksheets[3]  # Serial sheet

            container_rows, serial_rows = [], []

            # --- Container rows ---
            for row in sheet3.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                container_rows.append((
                    doc_no, bayan_no,
                    str(row[0]), str(row[1]),   # CONTAINER_NO, SEAL_NO
                    str(row[2]), str(row[3]),   # MODEL_NO, HS_CODE
                    row[4], str(row[5]),        # QTY_1, MEASURE_1
                    row[6], str(row[7]),        # QTY_2, MEASURE_2
                    row[8], row[9], row[10],    # MW, GW, KGS
                    initiator_no, initiator_name,
                    created_by, created_ip
                ))

            # --- Serial rows ---
            for row in sheet4.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                serial_rows.append((
                    doc_no, bayan_no,
                    str(row[0]), str(row[1]),   # MODEL_NO, SERIAL_NO
                    initiator_no, initiator_name,
                    created_by, created_ip
                ))

            # 📌 Bulk Insert into container/serial tables
            with transaction.atomic(using='Inbound_db'):
                with connections['Inbound_db'].cursor() as cursor:
                    if container_rows:
                        cursor.executemany("""
                            INSERT INTO WHR_INBOUND_BAYAN_CONTAINER_TBL
                            (DOC_NO, BAYAN_NO, DATE,
                             CONTAINER_NO, SEAL_NO, MODEL_NO, HS_CODE,
                             QTY_1, MEASURE_1, QTY_2, MEASURE_2,
                             MW, GW, KGS,
                             INITIATOR_NO, INITIATOR_NAME,
                             CREATION_DATE, CREATED_BY, CREATED_IP)
                            VALUES (%s, %s, GETDATE(),
                                    %s, %s, %s, %s,
                                    %s, %s, %s, %s,
                                    %s, %s, %s,
                                    %s, %s,
                                    GETDATE(), %s, %s)
                        """, container_rows)

                    if serial_rows:
                        cursor.executemany("""
                            INSERT INTO WHR_INBOUND_BAYAN_SERIALNO_TBL
                            (DOC_NO, BAYAN_NO, DATE,
                             MODEL_NO, SERIAL_NO,
                             INITIATOR_NO, INITIATOR_NAME,
                             CREATION_DATE, CREATED_BY, CREATED_IP)
                            VALUES (%s, %s, GETDATE(),
                                    %s, %s,
                                    %s, %s,
                                    GETDATE(), %s, %s)
                        """, serial_rows)

            return Response({
                "success": True,
                "excel_file_url": excel_url,
                "inserted_containers": len(container_rows),
                "inserted_serials": len(serial_rows),
                "inserted_documents": 1 + len(doc_files),  # Excel + supporting docs
                "doc_no": doc_no,
                "bayan_no": bayan_no
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        



# class DownloadDocumentView(APIView):
#     def get(self, request, document_id):
#         try:
#             # Get document from database
#             document = InboundContainerBayanDocument.objects.using('Inbound_db').get(id=document_id)
            
#             # Initialize MinIO client
#             endpoint = f"http{'s' if settings.MINIO_SECURE else ''}://{settings.MINIO_ENDPOINT}"
#             s3 = boto3.client(
#                 "s3",
#                 endpoint_url=endpoint,
#                 aws_access_key_id=settings.MINIO_ACCESS_KEY,
#                 aws_secret_access_key=settings.MINIO_SECRET_KEY,
#                 config=Config(signature_version="s3v4"),
#             )
            
#             # Extract key from file path (remove '/inboundalje/' prefix)
#             file_key = document.FILE_PATH.replace('/inboundalje/', '')
#             file_key = unquote(file_key)  # URL decode the key
            
#             # Get file from MinIO
#             response = s3.get_object(
#                 Bucket=settings.MINIO_INBOUND_CONTAINER_BUCKET_NAME,
#                 Key=file_key
#             )
            
#             # Get file content
#             file_content = response['Body'].read()
            
#             # Create response with file
#             response = HttpResponse(file_content, content_type='application/octet-stream')
#             response['Content-Disposition'] = f'attachment; filename="{document.DOCUMENT_NAME}"'
            
#             return response
            
#         except InboundContainerBayanDocument.DoesNotExist:
#             return Response({"error": "Document not found"}, status=status.HTTP_404_NOT_FOUND)
#         except Exception as e:
#             return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)



class DownloadDocumentView(APIView):

    def get(self, request, document_id):
        # 1️⃣ Fetch document from Inbound DB
        document = (
            InboundContainerBayanDocument.objects
            .using('Inbound_db')
            .filter(id=document_id)
            .first()
        )

        if not document:
            return Response(
                {"error": "Document not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        try:
            # 2️⃣ Initialize MinIO client
            protocol = "https" if settings.MINIO_SECURE else "http"
            endpoint = f"{protocol}://{settings.MINIO_ENDPOINT}"

            s3 = boto3.client(
                "s3",
                endpoint_url=endpoint,
                aws_access_key_id=settings.MINIO_ACCESS_KEY,
                aws_secret_access_key=settings.MINIO_SECRET_KEY,
                config=Config(signature_version="s3v4"),
            )

            # 3️⃣ Extract object key
            file_key = unquote(
                document.FILE_PATH.replace('/inboundalje/', '')
            )

            # 4️⃣ Get object from MinIO
            s3_response = s3.get_object(
                Bucket=settings.MINIO_INBOUND_CONTAINER_BUCKET_NAME,
                Key=file_key
            )

            # 5️⃣ Stream file to response
            response = HttpResponse(
                s3_response['Body'].read(),
                content_type=s3_response.get(
                    "ContentType", "application/octet-stream"
                )
            )

            response['Content-Disposition'] = (
                f'attachment; filename="{document.DOCUMENT_NAME}"'
            )

            return response

        except ClientError as e:
            return Response(
                {"error": "File not found in storage"},
                status=status.HTTP_404_NOT_FOUND
            )

        except Exception as e:
            return Response(
                {"error": str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )


@csrf_exempt
@require_POST
def update_shipped_qty(request):
    """
    Update SHIPPED_QTY in WHR_INBOUND_BAYAN_DETAILS_TBL
    Only edit shipped quantity and automatically update balance quantity
    """
    try:
        # Parse request data
        data = json.loads(request.body)
        product_id = data.get('product_id')
        shipped_qty = data.get('shipped_qty')
        initiator_no = data.get('initiator_no', '')
        initiator_name = data.get('initiator_name', '')
        
        # Validate required fields
        if not product_id or shipped_qty is None:
            return JsonResponse({
                'success': False,
                'error': 'Product ID and shipped quantity are required'
            }, status=400)
        
        # Validate shipped_qty is numeric and positive
        try:
            shipped_qty = float(shipped_qty)
            if shipped_qty < 0:
                return JsonResponse({
                    'success': False,
                    'error': 'Shipped quantity cannot be negative'
                }, status=400)
        except (ValueError, TypeError):
            return JsonResponse({
                'success': False,
                'error': 'Shipped quantity must be a valid number'
            }, status=400)
        
        # Update the database
        with connections['Inbound_db'].cursor() as cursor:
            # First, get the current values for validation
            cursor.execute("""
                SELECT PO_QTY, REC_QTY, BALANCE_QTY, SHIPPED_QTY 
                FROM WHR_INBOUND_BAYAN_DETAILS_TBL 
                WHERE PRODUCT_ID = %s
            """, [product_id])
            
            row = cursor.fetchone()
            if not row:
                return JsonResponse({
                    'success': False,
                    'error': 'Product not found'
                }, status=404)
            
            po_qty = row[0] or 0
            rec_qty = row[1] or 0
            current_balance_qty = row[2] or 0
            current_shipped_qty = row[3] or 0
            
            # Validate that shipped quantity doesn't exceed PO quantity
            
            # Calculate new balance quantity
            new_balance_qty = po_qty - rec_qty - shipped_qty
            
            # Validate that balance quantity doesn't go negative
            if new_balance_qty < 0:
                return JsonResponse({
                    'success': False,
                    'error': f'Insufficient quantity. Balance would be negative: {new_balance_qty}'
                }, status=400)
            
            # Update only the shipped quantity and recalculate balance
            cursor.execute("""
                UPDATE WHR_INBOUND_BAYAN_DETAILS_TBL 
                SET SHIPPED_QTY = %s,
                    INITIATOR_NO = %s,
                    INITIATOR_NAME = %s,
                    DATE = GETDATE()
                WHERE PRODUCT_ID = %s
            """, [shipped_qty, initiator_no, initiator_name, product_id])
        
        return JsonResponse({
            'success': True,
            'message': 'Shipped quantity updated successfully',
            'data': {
                'product_id': product_id,
                'old_shipped_qty': current_shipped_qty,
                'new_shipped_qty': shipped_qty,
                'old_balance_qty': current_balance_qty,
                'new_balance_qty': new_balance_qty,
                'po_qty': po_qty,
                'rec_qty': rec_qty
            }
        })
        
    except Exception as e:
        logger.error(f"Error updating shipped quantity: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'Internal server error: {str(e)}'

        }, status=500)
    

 
def get_supplier_by_DocNo(request, doc_no):
    if request.method != "GET":
        return JsonResponse({"status": "error", "message": "GET method required"}, status=400)
 
    try:
        sql = """
            SELECT SUPPLIER_NO, SUPPLIER_NAME, SUPPLIER_SITE
            FROM WHR_INBOUND_BAYAN_HEADER_TBL
            WHERE DOC_NO = %s
        """
 
        with connections['Inbound_db'].cursor() as cursor:
            cursor.execute(sql, [doc_no])
            row = cursor.fetchone()
 
        if not row:
            return JsonResponse({"status": "error", "message": "No record found"}, status=404)
 
        data = {
            "SUPPLIER_NO": row[0],
            "SUPPLIER_NAME": row[1],
            "SUPPLIER_SITE": row[2]
        }
 
        return JsonResponse({"status": "success", "data": data})
 
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=500)

from sqlite3 import Cursor, IntegrityError, OperationalError
def get_shipment_details_by_docno(request):
    """
    Single optimized API to fetch complete shipment details by DOC_NO only
    Fetches all data from details table including container information
    """
    try:
        doc_no = request.GET.get('doc_no')
        
        if not doc_no:
            return JsonResponse({
                'success': False,
                'message': 'DOC_NO parameter is required'
            }, status=400)
        
        with connections['Inbound_db'].cursor() as cursor:
            # ==============================================
            # 1. SINGLE QUERY TO GET ALL ITEMS WITH CONTAINER INFO
            # ==============================================
            cursor.execute("""
                SELECT 
                    d.PRODUCT_ID,
                    d.DOC_NO,
                    d.BAYAN_NO,
                    d.PO_NUMBER,
                    d.FRANCHISE,
                    d.CLASS,
                    d.SUBCLASS,
                    d.ITEM_CODE,
                    d.ITEM_DESC,
                    d.PO_QTY,
                    d.REC_QTY,
                    d.BALANCE_QTY,
                    d.SHIPPED_QTY,
                    d.CONTAINER_NO,
                    d.SEAL_NO,
                    d.HS_CODE,
                    d.QTY_1,
                    d.MEASURE_1,
                    d.QTY_2,
                    d.MEASURE_2,
                    d.MW,
                    d.GW,
                    d.KGS,
                    d.ASSIGNED_QTY,
                    d.SCANNED_QTY,
                    d.TRUCK_NUMBER,
                    d.ARRIVAL_DATE,
                    d.EXIT_DATE,
                    d.DRIVER_INFO,
                    CASE
                        WHEN d.STATUS IN ('Pending Offloading', 'Pending GRN', 'Completed') 
                        THEN d.STATUS
                        WHEN h.STATUS IN ('Pending Offloading', 'Pending GRN', 'Completed')
                        THEN 'WH Transportation'
                        WHEN d.STATUS IS NOT NULL AND d.STATUS <> 'H' AND d.STATUS <> ''
                        THEN d.STATUS
                        ELSE h.STATUS
                    END AS STATUS,
                    d.DATE,
                    -- Count serials for this item
                    (SELECT COUNT(*) 
                     FROM   [WHR_INBOUND_BAYAN_SERIALNO_TBL] s
                     WHERE s.DOC_NO = d.DOC_NO 
                     AND s.PO_NUMBER = d.PO_NUMBER) as SERIAL_COUNT
                FROM   [WHR_INBOUND_BAYAN_DETAILS_TBL] d
                INNER JOIN [WHR_INBOUND_BAYAN_HEADER_TBL] h ON d.DOC_NO = h.DOC_NO
                WHERE d.DOC_NO = %s
                ORDER BY d.BAYAN_NO, d.PO_NUMBER, d.ITEM_CODE
            """, [doc_no])
            
            columns = [col[0] for col in cursor.description]
            items = [dict(zip(columns, row)) for row in cursor.fetchall()]
            
            if not items:
                return JsonResponse({
                    'success': False,
                    'message': f'No data found for DOC_NO: {doc_no}'
                }, status=404)
            
            # ==============================================
            # 2. FETCH ALL SERIAL NUMBERS (Single Query)
            # ==============================================
            cursor.execute("""
                SELECT 
                    PO_NUMBER,
                    MODEL_NO,
                    SERIAL_NO,
                    CREATION_DATE
                FROM   [WHR_INBOUND_BAYAN_SERIALNO_TBL]
                WHERE DOC_NO = %s
                ORDER BY PO_NUMBER, SERIAL_NO
            """, [doc_no])
            
            columns = [col[0] for col in cursor.description]
            all_serials = [dict(zip(columns, row)) for row in cursor.fetchall()]
            
            # Group serials by PO_NUMBER for quick lookup
            serials_by_po = {}
            for serial in all_serials:
                po = serial['PO_NUMBER']
                if po not in serials_by_po:
                    serials_by_po[po] = []
                serials_by_po[po].append({
                    'serial_no': serial['SERIAL_NO'],
                    'model_no': serial['MODEL_NO'],
                    'created_date': serial['CREATION_DATE']
                })
            
            # ==============================================
            # 3. GET UNIQUE CONTAINERS FROM DETAILS TABLE
            # ==============================================
            cursor.execute("""
                SELECT DISTINCT
                    CONTAINER_NO,
                    BAYAN_NO,
                    SEAL_NO,
                    HS_CODE,
                    QTY_1,
                    MEASURE_1,
                    QTY_2,
                    MEASURE_2,
                    MW,
                    GW,
                    KGS
                FROM   [WHR_INBOUND_BAYAN_DETAILS_TBL]
                WHERE DOC_NO = %s 
                AND CONTAINER_NO IS NOT NULL 
                AND CONTAINER_NO != ''
                ORDER BY CONTAINER_NO
            """, [doc_no])
            
            columns = [col[0] for col in cursor.description]
            containers = [dict(zip(columns, row)) for row in cursor.fetchall()]
            
            # ==============================================
            # 4. FETCH SUMMARY STATISTICS (Single Query)
            # ==============================================
            cursor.execute("""
                SELECT 
                    COUNT(DISTINCT BAYAN_NO) as TOTAL_BAYANS,
                    COUNT(DISTINCT PO_NUMBER) as TOTAL_POS,
                    COUNT(DISTINCT CASE 
                        WHEN CONTAINER_NO IS NOT NULL AND CONTAINER_NO != '' 
                        THEN CONTAINER_NO 
                    END) as TOTAL_CONTAINERS,
                    COUNT(*) as TOTAL_ITEMS,
                    SUM(PO_QTY) as TOTAL_PO_QTY,
                    SUM(REC_QTY) as TOTAL_REC_QTY,
                    SUM(BALANCE_QTY) as TOTAL_BALANCE_QTY,
                    SUM(SHIPPED_QTY) as TOTAL_SHIPPED_QTY,
                    MIN(DATE) as EARLIEST_DATE,
                    MAX(DATE) as LATEST_DATE,
                    -- Get most recent status
                    (SELECT TOP 1 STATUS 
                     FROM   [WHR_INBOUND_BAYAN_DETAILS_TBL] 
                     WHERE DOC_NO = %s 
                     ORDER BY DATE DESC) as CURRENT_STATUS
                FROM   [WHR_INBOUND_BAYAN_DETAILS_TBL]
                WHERE DOC_NO = %s
            """, [doc_no, doc_no])
            
            summary_row = cursor.fetchone()
            summary_columns = [col[0] for col in cursor.description]
            summary = dict(zip(summary_columns, summary_row))
            
            # ==============================================
            # 5. GROUP DATA BY BAYAN_NO FOR ORGANIZATION
            # ==============================================
            bayan_groups = {}
            
            for item in items:
                bayan_no = item['BAYAN_NO']
                
                if bayan_no not in bayan_groups:
                    bayan_groups[bayan_no] = {
                        'bayan_no': bayan_no,
                        'items': [],
                        'containers': [],
                        'summary': {
                            'item_count': 0,
                            'shipped_qty': 0,
                            'po_count': set(),
                            'container_count': set()
                        }
                    }
                
                # Add container to bayan's container set if exists
                if item['CONTAINER_NO']:
                    bayan_groups[bayan_no]['summary']['container_count'].add(item['CONTAINER_NO'])
                
                # Add PO to bayan's PO set
                bayan_groups[bayan_no]['summary']['po_count'].add(item['PO_NUMBER'])
                
                # Get serials for this item
                item_serials = serials_by_po.get(item['PO_NUMBER'], [])
                
                # Prepare container details from the item itself
                container_details = {}
                if item['CONTAINER_NO']:
                    container_details = {
                        'container_no': item['CONTAINER_NO'],
                        'seal_no': item['SEAL_NO'],
                        'hs_code': item['HS_CODE'],
                        'qty_1': item['QTY_1'],
                        'measure_1': item['MEASURE_1'],
                        'qty_2': item['QTY_2'],
                        'measure_2': item['MEASURE_2'],
                        'mw': item['MW'],
                        'gw': item['GW'],
                        'kgs': item['KGS']
                    }
                
                # Add item with all details
                bayan_groups[bayan_no]['items'].append({
                    'product_id': item['PRODUCT_ID'],
                    'po_number': item['PO_NUMBER'],
                    'franchise': item['FRANCHISE'],
                    'class': item['CLASS'],
                    'subclass': item['SUBCLASS'],
                    'item_code': item['ITEM_CODE'],
                    'item_desc': item['ITEM_DESC'],
                    'po_qty': item['PO_QTY'],
                    'rec_qty': item['REC_QTY'],
                    'balance_qty': item['BALANCE_QTY'],
                    'shipped_qty': item['SHIPPED_QTY'],
                    'container_no': item['CONTAINER_NO'],
                    'seal_no': item['SEAL_NO'],
                    'hs_code': item['HS_CODE'],
                    'qty_1': item['QTY_1'],
                    'measure_1': item['MEASURE_1'],
                    'qty_2': item['QTY_2'],
                    'measure_2': item['MEASURE_2'],
                    'mw': item['MW'],
                    'gw': item['GW'],
                    'kgs': item['KGS'],
                    'assigned_qty': item['ASSIGNED_QTY'],
                    'scanned_qty': item['SCANNED_QTY'],
                    'truck_number': item['TRUCK_NUMBER'],
                    'arrival_date': item['ARRIVAL_DATE'],
                    'exit_date': item['EXIT_DATE'],
                    'driver_info': item['DRIVER_INFO'],
                    'status': item['STATUS'],
                    'date': item['DATE'],
                    'serial_numbers': item_serials,
                    'serial_count': item['SERIAL_COUNT']
                })
                
                # Update bayan summary
                bayan_groups[bayan_no]['summary']['item_count'] += 1
                bayan_groups[bayan_no]['summary']['shipped_qty'] += (item['SHIPPED_QTY'] or 0)
            
            # Convert sets to counts for JSON serialization
            for bayan in bayan_groups.values():
                bayan['summary']['po_count'] = len(bayan['summary']['po_count'])
                bayan['summary']['container_count'] = len(bayan['summary']['container_count'])
                
                # Get unique containers for this bayan from containers list
                bayan_containers = [c for c in containers if c['BAYAN_NO'] == bayan['bayan_no']]
                bayan['containers'] = [{
                    'container_no': c['CONTAINER_NO'],
                    'seal_no': c['SEAL_NO'],
                    'hs_code': c['HS_CODE']
                } for c in bayan_containers]
            
            # ==============================================
            # 6. PREPARE FINAL RESPONSE
            # ==============================================
            response_data = {
                'success': True,
                'doc_no': doc_no,
                'summary': {
                    'total_bayans': summary['TOTAL_BAYANS'] or 0,
                    'total_pos': summary['TOTAL_POS'] or 0,
                    'total_containers': summary['TOTAL_CONTAINERS'] or 0,
                    'total_items': summary['TOTAL_ITEMS'] or 0,
                    'total_po_qty': float(summary['TOTAL_PO_QTY'] or 0),
                    'total_rec_qty': float(summary['TOTAL_REC_QTY'] or 0),
                    'total_balance_qty': float(summary['TOTAL_BALANCE_QTY'] or 0),
                    'total_shipped_qty': float(summary['TOTAL_SHIPPED_QTY'] or 0),
                    'completion_percentage': round(
                        (float(summary['TOTAL_REC_QTY'] or 0) / float(summary['TOTAL_PO_QTY'] or 1) * 100), 2
                    ) if summary['TOTAL_PO_QTY'] else 0,
                    'earliest_date': summary['EARLIEST_DATE'],
                    'latest_date': summary['LATEST_DATE'],
                    'current_status': summary['CURRENT_STATUS']
                },
                'bayans': list(bayan_groups.values()),
                'all_containers': [{
                    'container_no': c['CONTAINER_NO'],
                    'bayan_no': c['BAYAN_NO'],
                    'seal_no': c['SEAL_NO'],
                    'hs_code': c['HS_CODE']
                } for c in containers],
                'total_serials': len(all_serials),
                'timestamp': datetime.now().isoformat()
            }
            
            # Add bayan numbers list
            response_data['bayan_numbers'] = list(bayan_groups.keys())
            
            # Add item count for first bayan (for backward compatibility)
            if bayan_groups:
                first_bayan = list(bayan_groups.values())[0]
                response_data['items'] = first_bayan['items']
            
            return JsonResponse(response_data, safe=False)
            
    except OperationalError as e:
        return JsonResponse({
            'success': False,
            'message': f'Database error: {str(e)}'
        }, status=500)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Unexpected error: {str(e)}'
        }, status=500)
    
######################################################  INBOUND  ##########################################################################      

from django.shortcuts import render

# Create your views here.
from django.http import JsonResponse
from django.db import IntegrityError, transaction, connection
from .models import Uniq_GatePass_tbl
from datetime import datetime, date, time
import json
from django.views.decorators.csrf import csrf_exempt
from django.db.models.functions import Length

def get_next_gatepass_no():
    """Helper function to generate the next unique Gate Pass Number."""
    while True:
        try:
            # 1. Get current Year (YY) and Month (MM)
            now = datetime.now()
            year_str = now.strftime("%y")  # e.g., '25'
            month_str = now.strftime("%m") # e.g., '12'
            prefix = f"GTP{year_str}{month_str}" # e.g., 'GTP2512'

            # 2. Find the latest sequence number for this Month/Year
            # We filter by starts_with prefix
            # Annotate with length to sort correctly (so GTP...100 > GTP...99)
            last_entry = Uniq_GatePass_tbl.objects.using('Inbound_db').filter(
                GatPass_No__startswith=prefix
            ).annotate(
                text_len=Length('GatPass_No')
            ).order_by('-text_len', '-GatPass_No').first()

            if last_entry:
                # Extract the last 2 digits
                last_no = last_entry.GatPass_No
                # Assuming format is always GTPYYMMSS (9 chars total), 
                # or just take the part after the prefix
                sequence_str = last_no[len(prefix):] 
                if sequence_str.isdigit():
                    next_seq = int(sequence_str) + 1
                else:
                    # Fallback if for some reason it's not a digit, though it should be
                    next_seq = 1
            else:
                next_seq = 1

            # 3. Format the new GatePass_No with zero-padding (2 digits)
            new_gate_pass_no = f"{prefix}{next_seq:02d}"

            # 4. Attempt to create the record
            # atomic=True ensures that if this fails, it rolls back
            # But here we rely on IntegrityError from unique constraint
            with transaction.atomic(using='Inbound_db'):
                gate_pass = Uniq_GatePass_tbl.objects.using('Inbound_db').create(GatPass_No=new_gate_pass_no)
            
            return gate_pass.GatPass_No

        except IntegrityError:
            # Duplicate found! The loop will restart, 
            # fetch the NEW latest entry, and try the next number.
            continue
        except Exception as e:
            # Do not swallow other errors, otherwise it causes infinite loops
            raise e

def generate_gate_pass(request):
    gate_pass_no = get_next_gatepass_no()
    return JsonResponse({
        'status': 'success',
        'GatePass_No': gate_pass_no
    })

from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.db import connections
import json
import logging

logger = logging.getLogger(__name__)


@csrf_exempt
def insert_inbound_gatepass(request):
    if request.method != 'POST':
        return JsonResponse(
            {'status': 'error', 'message': 'Invalid method'},
            status=405
        )

    try:
        data = json.loads(request.body)
        # Generate GatePass No
        gate_pass_no = get_next_gatepass_no()

        items_list = data.get('items_details') or data.get('Item_deatils')
        if not items_list or not isinstance(items_list, list):
            items_list = [{
                'PO_Qty': data.get('PO_Qty'),
                'Item_Code': data.get('Item_Code'),
                'Item_Decription': data.get('Item_Decription')
            }]

        insert_sql = """
            INSERT INTO [BUYP_INBOUND].[dbo].[WHR_Inbound_Genrate_GatePass_tbl](
                Doc_No, GatePass_No, Truck_No, Driver_Name, Driver_MobileNo, 
                Arivale_DriverTime, ExceptedExit_dateTime, Remarks, 
                Inound_Iniater_No, Inound_Reciverd_No, Inound_Reciverd_Name, Inbound_Iniater_Name, 
                Creaetion_Date, Creation_By, Updation_Date, 
                Container_No, Bayan_No, Supplier_No, Supplier_Name,
                PO_Number, PO_Qty, Item_Code, Item_Decription
            ) VALUES (
                %s, %s, %s, %s, %s, 
                %s, %s, %s, 
                %s, %s, %s, %s, 
                GETDATE(), %s, GETDATE(), 
                %s, %s, %s, %s,
                %s, %s, %s, %s
            )
        """

        update_bayan_sql = """
            UPDATE [BUYP_INBOUND].[dbo].[WHR_INBOUND_BAYAN_DETAILS_TBL]
            SET STATUS = 'Pending Offloading'
            WHERE CONTAINER_NO = %s
              AND PO_NUMBER   = %s
              AND ITEM_CODE   = %s
              AND DOC_NO      = %s
        """

        update_header_sql = """
            UPDATE [BUYP_INBOUND].[dbo].[WHR_INBOUND_BAYAN_HEADER_TBL]
            SET STATUS = 'Pending Offloading'
            WHERE DOC_NO = %s AND (STATUS IS NULL OR STATUS IN ('In Transit', 'Under Clearance', 'WH Transportation'))
        """

        with connections['Inbound_db'].cursor() as cursor:
            # Sync header status if it's at an earlier stage
            doc_no = data.get('Doc_No')
            if doc_no:
                cursor.execute(update_header_sql, [doc_no])

            for item in items_list:
                if not isinstance(item, dict):
                    continue

                insert_params = [
                    data.get('Doc_No', ''),
                    gate_pass_no,
                    data.get('Truck_No', ''),
                    data.get('Driver_Name', ''),
                    data.get('Driver_MobileNo', ''),
                    data.get('Arivale_DriverTime') or None,
                    data.get('ExceptedExit_dateTime') or None,
                    data.get('Remarks', ''),
                    data.get('Inound_Iniater_No', ''),
                    data.get('Inound_Reciverd_No', ''),
                    data.get('Inound_Reciverd_Name', ''),
                    data.get('Inbound_Iniater_Name', ''),
                    data.get('Creation_By', ''),
                    data.get('Container_No', ''),
                    data.get('Bayan_No') or None,
                    data.get('Supplier_No', ''),
                    data.get('Supplier_Name', ''),
                    data.get('PO_Number', ''),
                    item.get('PO_Qty') or None,
                    item.get('Item_Code', ''),
                    item.get('Item_Decription', '')
                ]

                # INSERT Gatepass
                cursor.execute(insert_sql, insert_params)

                update_params = [
                    data.get('Container_No', ''),
                    data.get('PO_Number', ''),
                    item.get('Item_Code', ''),
                    data.get('Doc_No', '')
                ]

                # UPDATE Bayan status
                cursor.execute(update_bayan_sql, update_params)

        return JsonResponse({
            'status': 'success',
            'GatePass_No': gate_pass_no,
            'items_count': len(items_list)
        })

    except Exception as e:
        logger.exception("Inbound gatepass insert failed")
        return JsonResponse(
            {
                'status': 'error',
                'message': 'Failed to insert inbound gatepass'
            },
            status=500
        )

@csrf_exempt
def save_po_documents_minio(request):
    """
    Upload PO documents to MinIO and save metadata in DB with PO numbers
    """
    print("\n" + "="*80)
    print("DEBUG: save_po_documents_minio called")
    print("="*80)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        # Get parameters
        doc_no = request.POST.get('doc_no', '').strip()
        po_number = request.POST.get('po_number', '').strip()
        initiator_no = request.POST.get('initiator_no', 'unknown')
        initiator_name = request.POST.get('initiator_name', 'Unknown')
        
        print(f"\n📊 Parameters extracted:")
        print(f"  doc_no: {doc_no}")
        print(f"  po_number: {po_number}")
        print(f"  initiator_no: {initiator_no}")
        print(f"  initiator_name: {initiator_name}")

        if not doc_no:
            return JsonResponse({'error': 'doc_no is required'}, status=400)
        
        if not po_number:
            return JsonResponse({'error': 'po_number is required'}, status=400)

        # Check for bucket configuration
        if not hasattr(settings, 'MINIO_INBOUND_PO_BUCKET_NAME') or not settings.MINIO_INBOUND_PO_BUCKET_NAME:
            return JsonResponse(
                {'error': 'MINIO_INBOUND_PO_BUCKET_NAME not configured'},
                status=500
            )
        
        bucket_name = settings.MINIO_INBOUND_PO_BUCKET_NAME
        print(f"✓ Using bucket: {bucket_name}")

        # MinIO configuration
        endpoint = f"http{'s' if getattr(settings, 'MINIO_SECURE', False) else ''}://{settings.MINIO_ENDPOINT}"
        
        try:
            s3_client = boto3.client(
                service_name='s3',
                endpoint_url=endpoint,
                aws_access_key_id=settings.MINIO_ACCESS_KEY,
                aws_secret_access_key=settings.MINIO_SECRET_KEY,
                region_name='us-east-1'
            )
        except Exception as e:
            return JsonResponse({'error': f'MinIO configuration error: {e}'}, status=500)

        # Check/create bucket
        try:
            s3_client.head_bucket(Bucket=bucket_name)
        except ClientError as e:
            if e.response['Error']['Code'] == '404':
                try:
                    s3_client.create_bucket(Bucket=bucket_name)
                except Exception as create_error:
                    return JsonResponse(
                        {'error': f'Failed to create bucket: {create_error}'},
                        status=500
                    )
            else:
                return JsonResponse({'error': f'Bucket check failed: {e}'}, status=500)

        # Get uploaded files
        uploaded_files = []
        for key in request.FILES:
            uploaded_files.extend(request.FILES.getlist(key))
        
        if not uploaded_files:
            return JsonResponse({'error': 'No documents uploaded'}, status=400)

        saved_files = []
        errors = []

        # Get existing document count for this PO to generate sequence numbers
        existing_count = InboundPoDocument.objects.filter(
            DOC_NO=doc_no, 
            PO_NUMBER=po_number
        ).count()
        
        print(f"📊 Existing documents for PO {po_number}: {existing_count}")

        for index, uploaded_file in enumerate(uploaded_files, 1):
            try:
                print(f"\n📄 Processing file {index}/{len(uploaded_files)}: {uploaded_file.name}")
                
                # Generate sequential filename: DocumentName_8372-1.ext
                sequence_number = existing_count + index
                original_name = uploaded_file.name
                file_name_without_ext, file_ext = os.path.splitext(original_name)
                
                # Create new filename: OriginalName_PO-Sequence.ext
                # Examples: 
                #   Invoice.pdf → Invoice_8372-1.pdf
                #   PackingList.jpg → PackingList_8372-2.jpg
                new_filename = f"{file_name_without_ext}_{po_number}-{sequence_number}{file_ext}"
                
                # Create object path
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                unique_id = uuid.uuid4().hex[:8]
                object_name = f"{doc_no}/{po_number}/{timestamp}_{unique_id}_{new_filename}"
                
                print(f"  Original name: {original_name}")
                print(f"  New filename: {new_filename}")
                print(f"  Object name: {object_name}")

                # Read and upload file
                file_content = uploaded_file.read()
                
                s3_client.put_object(
                    Bucket=bucket_name,
                    Key=object_name,
                    Body=file_content,
                    ContentType=uploaded_file.content_type or 'application/octet-stream',
                    Metadata={
                        'original_name': original_name,
                        'doc_no': doc_no,
                        'po_number': po_number,
                        'sequence_number': str(sequence_number),
                        'display_name': new_filename,  # New field for display
                        'uploaded_by': initiator_name
                    }
                )

                # Generate presigned URL
                presigned_url = s3_client.generate_presigned_url(
                    'get_object',
                    Params={'Bucket': bucket_name, 'Key': object_name},
                    ExpiresIn=7 * 24 * 60 * 60  # 7 days
                )

                # Save to database
                doc_obj = InboundPoDocument.objects.create(
                    DOC_NO=doc_no,
                    PO_NUMBER=po_number,
                    DOCUMENT_NAME=new_filename,  # Save the new display name
                    FILE_PATH=object_name,
                    INITIATOR_NO=initiator_no,
                    INITIATOR_NAME=initiator_name,
                    CREATED_BY=initiator_name,
                    CREATED_IP=request.META.get('REMOTE_ADDR', 'unknown')
                )

                saved_files.append({
                    'id': doc_obj.id,
                    'po_number': po_number,
                    'original_name': original_name,
                    'display_name': new_filename,  # Show display name
                    'object_name': object_name,
                    'file_url': presigned_url,
                    'bucket_name': bucket_name,
                    'sequence_number': sequence_number,
                    'size': uploaded_file.size
                })

            except Exception as e:
                error_msg = str(e)
                print(f"❌ ERROR processing file {uploaded_file.name}: {error_msg}")
                errors.append({
                    'file_name': uploaded_file.name,
                    'error': error_msg
                })

        # Prepare response
        if saved_files:
            status = 'partial_success' if errors else 'success'
        else:
            status = 'error'

        response = {
            'status': status,
            'doc_no': doc_no,
            'po_number': po_number,
            'bucket_name': bucket_name,
            'total_files': len(uploaded_files),
            'successful_uploads': len(saved_files),
            'saved_files': saved_files
        }

        if errors:
            response['errors'] = errors

        return JsonResponse(response, status=201)

    except Exception as e:
        print(f"\n❌ UNEXPECTED ERROR: {e}")
        import traceback
        traceback.print_exc()
        
        return JsonResponse(
            {
                'status': 'error',
                'message': str(e)
            },
            status=500
        )







def get_all_inbound_gatepass(request):
    """
    Optimized view to retrieve TOP 1000 records.
    """
    try:
        sql = """
            SELECT TOP (1000) [id]
                ,[Doc_No]
                ,[GatePass_No]
                ,[Truck_No]
                ,[Driver_Name]
                ,[Driver_MobileNo]
                ,[Arivale_DriverTime]
                ,[ExceptedExit_dateTime]
                ,[Remarks]
                ,[Inound_Iniater_No]
                ,[Inbound_Iniater_Name]
                ,[Creaetion_Date]
                ,[Creation_By]
                ,[updation_By]
                ,[Updation_Date]
                ,[Atrribute1]
                ,[Attribute2]
                ,[Attibute3]
                ,[Attribute4]
                ,[Attribute5]
                ,[Flag1]
                ,[Flag2]
                ,[Container_No]
                ,[Bayan_No]
                ,[Supplier_No]
                ,[Supplier_Name]
                ,[Inound_Reciverd_No]
                ,[Inound_Reciverd_Name]
                ,[Status]
                ,[PO_Number]
                ,[PO_Qty]
                ,[Item_Code]
                ,[Item_Decription]
            FROM   [WHR_Inbound_Genrate_GatePass_tbl]
            ORDER BY [id] DESC
        """
        with connections['Inbound_db'].cursor() as cursor:
            cursor.execute(sql)
            columns = [col[0] for col in cursor.description]
            rows = cursor.fetchall()
           
        results = []
        for row in rows:
            results.append(dict(zip(columns, row)))
           
        return JsonResponse(results, safe=False)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
 
def get_inbound_gatepass_by_id(request, gatepass_no):
    """
    Optimized view to retrieve a single record by GatePass_No.
    """
    try:
        sql = """
            SELECT [id]
                ,[Doc_No]
                ,[GatePass_No]
                ,[Truck_No]
                ,[Driver_Name]
                ,[Driver_MobileNo]
                ,[Arivale_DriverTime]
                ,[ExceptedExit_dateTime]
                ,[Remarks]
                ,[Inound_Iniater_No]
                ,[Inbound_Iniater_Name]
                ,[Creaetion_Date]
                ,[Creation_By]
                ,[updation_By]
                ,[Updation_Date]
                ,[Atrribute1]
                ,[Attribute2]
                ,[Attibute3]
                ,[Attribute4]
                ,[Attribute5]
                ,[Flag1]
                ,[Flag2]
                ,[Container_No]
                ,[Bayan_No]
                ,[Supplier_No]
                ,[Supplier_Name]
                ,[Inound_Reciverd_No]
                ,[Inound_Reciverd_Name]
                ,[Status]
                ,[PO_Number]
                ,[PO_Qty]
                ,[Item_Code]
                ,[Item_Decription]
            FROM   [WHR_Inbound_Genrate_GatePass_tbl]
            WHERE [GatePass_No] = %s
        """
        with connections['Inbound_db'].cursor() as cursor:
            cursor.execute(sql, [gatepass_no])
            columns = [col[0] for col in cursor.description]
            row = cursor.fetchone()
           
        if row:
            result = dict(zip(columns, row))
            return JsonResponse(result)
        else:
            return JsonResponse({'status': 'error', 'message': 'Record not found'}, status=404)
           
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)




@csrf_exempt
def get_container_data(request):
    """
    Optimized view to fetch container & item details
    Designed to handle high concurrency (~1000 users)
    """
    if request.method != 'POST':
        return JsonResponse(
            {'status': 'error', 'message': 'POST method required'},
            status=405
        )
 
    try:
        body = request.body
        if not body:
            return JsonResponse(
                {'status': 'error', 'message': 'Empty request body'},
                status=400
            )
 
        data = json.loads(body)
        doc_no = data.get('DOC_NO')
        container_no = data.get('CONTAINER_NO')
 
        if not doc_no or not container_no:
            return JsonResponse(
                {'status': 'error', 'message': 'Missing DOC_NO or CONTAINER_NO'},
                status=400
            )
 
        result_data = {
            'container_info': {},
            'item_details': [],
            'has_multiple_items': False
        }
 
        with connections['Inbound_db'].cursor() as cursor:
 
            # 1️⃣ Fetch container info
            cursor.execute("""
                SELECT TOP (1)
                    [PRODUCT_ID], [DOC_NO], [BAYAN_NO], [DATE], [PO_NUMBER],
                    [CONTAINER_NO], [SEAL_NO], [ITEM_CODE], [HS_CODE],
                    [QTY_1], [MEASURE_1], [QTY_2], [MEASURE_2],
                    [MW], [GW], [KGS],
                    [INITIATOR_NO], [INITIATOR_NAME],
                    [CREATION_DATE], [CREATION_BY], [CREATION_IP],
                    [UPDATION_DATE], [UPDATION_BY]
                FROM   [WHR_INBOUND_BAYAN_DETAILS_TBL]
                WHERE [DOC_NO] = %s AND [CONTAINER_NO] = %s
            """, [doc_no, container_no])
 
            row = cursor.fetchone()
            if row:
                columns = [col[0] for col in cursor.description]
                result_data['container_info'] = dict(zip(columns, row))
 
            # 2️⃣ Fetch item details
            cursor.execute("""
                SELECT
                    [PO_NUMBER], [ITEM_CODE], [ITEM_DESC],
                    [PO_QTY], [REC_QTY], [BALANCE_QTY],
                    [SHIPPED_QTY], [CONTAINER_NO],
                    [FRANCHISE], [CLASS], [SUBCLASS]
                FROM   [WHR_INBOUND_BAYAN_DETAILS_TBL]
                WHERE [DOC_NO] = %s AND [CONTAINER_NO] = %s
                ORDER BY [ITEM_CODE]
            """, [doc_no, container_no])
 
            rows = cursor.fetchall()
            if rows:
                columns = [col[0] for col in cursor.description]
                result_data['item_details'] = [
                    dict(zip(columns, r)) for r in rows
                ]
 
        # 3️⃣ Business rule checks
        item_count = len(result_data['item_details'])
        result_data['has_multiple_items'] = item_count > 1
 
        # 4️⃣ Fallback logic (unchanged)
        if item_count == 0 and result_data['container_info']:
            c = result_data['container_info']
            result_data['item_details'].append({
                'PO_NUMBER': c.get('PO_NUMBER'),
                'ITEM_CODE': c.get('MODEL_NO'),
                'ITEM_DESC': f"Container Item - {c.get('MODEL_NO', 'N/A')}",
                'PO_QTY': c.get('QTY_1'),
                'SHIPPED_QTY': c.get('QTY_1'),
                'CONTAINER_NO': c.get('CONTAINER_NO'),
            })
 
        return JsonResponse({
            'status': 'success',
            'data': result_data
        })
 
    except json.JSONDecodeError:
        return JsonResponse(
            {'status': 'error', 'message': 'Invalid JSON body'},
            status=400
        )
    except Exception as e:
        print(traceback.format_exc())
        return JsonResponse(
            {'status': 'error', 'message': 'Internal server error'},
            status=500
        )





@method_decorator(csrf_exempt, name='dispatch')
class InboundOnProgressDataView(View):
    def get(self, request):
        try:
            # Pagination parameters
            page = int(request.GET.get('page', 1))
            page_size = int(request.GET.get('page_size', 50))
           
            if page < 1:
                page = 1
            # For streaming, we can allow larger page sizes, but still keep a sane limit to prevent DB abuse.
            if page_size < 1:
                page_size = 50
            if page_size > 50000: # Increased limit significantly for "single page" views of large chunks
                page_size = 50000
 
            offset = (page - 1) * page_size
 
            # Generator function to stream the response
            def data_generator():
                yield '{'
                yield f'"page": {page}, "page_size": {page_size}, "results": ['
               
                # REMOVED 'DISTINCT' to ensure all lines are shown
                # UPDATED 'ORDER BY' to ensure stable pagination
                sql_query = """
                    SELECT
                    d.PO_NUMBER,
                    d.ITEM_CODE,
                    d.SUBCLASS,
                    d.FRANCHISE,
                    d.SHIPPED_QTY,
                    h.MASTER_PL,
                    h.ETA_DATE,

                    CASE
                        -- 1. If item has reached an advanced operational stage, use its specific status
                        WHEN d.STATUS IN ('Pending Offloading', 'Pending GRN', 'Completed') 
                        THEN d.STATUS
                        
                        -- 2. If header is at an advanced stage but this item hasn't marked arrival yet,
                        -- it means this specific container is still in the 'WH Transportation' phase.
                        WHEN h.STATUS IN ('Pending Offloading', 'Pending GRN', 'Completed')
                        THEN 'WH Transportation'
                        
                        -- 3. For earlier stages, if item has a specific status (not default 'H'), use it
                        WHEN d.STATUS IS NOT NULL AND d.STATUS <> 'H' AND d.STATUS <> ''
                        THEN d.STATUS
                        
                        -- 4. Otherwise, follow the shipment header status
                        ELSE h.STATUS
                    END AS STATUS,

                    h.GRN,
                    h.GRN_DATE
                FROM [WHR_INBOUND_BAYAN_DETAILS_TBL] d
                INNER JOIN [WHR_INBOUND_BAYAN_HEADER_TBL] h
                    ON d.DOC_NO = h.DOC_NO
                ORDER BY
                    d.PO_NUMBER,
                    d.ITEM_CODE,
                    d.SUBCLASS,
                    d.FRANCHISE,
                    d.SHIPPED_QTY

                    OFFSET %s ROWS
                    FETCH NEXT %s ROWS ONLY;
                """
 
                # Manual connection management for streaming (to keep cursor open during yield)
                # In Django View, connection is usually managed per request, but let's be safe with standard execution.
                # However, 'yield' pauses execution. We must ensure the cursor isn't closed prematurely.
                # Since we are inside the 'data_generator', we open the cursor here.
               
                with connections['Inbound_db'].cursor() as cursor:
                    cursor.execute(sql_query, [offset, page_size])
                   
                    if cursor.description:
                        columns = [col[0] for col in cursor.description]
                       
                        first = True
                        while True:
                            # Fetch in chunks to keep server memory low
                            rows = cursor.fetchmany(1000)
                            if not rows:
                                break
                           
                            for row in rows:
                                record = dict(zip(columns, row))
                                if not first:
                                    yield ','
                                else:
                                    first = False
                               
                                # Manually serialize to string to stream immediately
                                # using default=str to handle Dates/Decimals
                                yield json.dumps(record, default=str)
               
                yield ']}'
 
            # Use StreamingHttpResponse
            response = StreamingHttpResponse(data_generator(), content_type='application/json')
            # Disable buffering in proxies if possible (optional)
            response['X-Accel-Buffering'] = 'no'
            return response
 
        except ValueError:
             return JsonResponse({'error': 'Invalid page or page_size parameters'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)


# -------------------------------------------------------------------------
# NEW BULK INSERT VIEW (Function-Based)
# -------------------------------------------------------------------------
@api_view(['POST'])
def inbound_container_qc_insert(request):
    """
    Insert container QC data with image count and MinIO location
    Handles container numbers with slashes
    """
    try:
        payload = request.data
        
        # Extract Header Data
        doc_no = payload.get('DocNo')
        date_val = payload.get('Date')
        container_no = payload.get('Container_No')  # Original: GCXU5089740/40HQ
        gatepass_no = payload.get('GatePass_No')
        bayan_no = payload.get('Bayan_No')
        supplier_name = payload.get('Supplier_Name')
        supplier_no = payload.get('Supplier_No')        
        container_status = payload.get('status')  # Renamed from status to container_status
        inbound_initiator_name = payload.get('Inbound_Initiater_Name')
        inbound_initiator_no = payload.get('Inbound_Initiator_No')
        inbound_receiver_no = payload.get('Inbound_Reciver_No')
        inbound_receiver_name = payload.get('Inound_Reciver_Name')
        gatepass_generated_by_no = payload.get('GatePassGeneratedBy_No')
        gatepass_generated_by_name = payload.get('GatePassGeneartedBy_Name')
        image_count = payload.get('Image_Count', 0)
        minioimage_location = payload.get('MinioImage_Location', '')
        remarks = payload.get('Remarks')
        
        # Audit info
        creation_by = payload.get('Creation_by', '')
        creation_ip = request.META.get('REMOTE_ADDR', '')

        item_details = payload.get('ItemDetails', [])
        
        if not item_details:
            return Response(
                {"status": "error", "message": "No ItemDetails provided"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Clean container number for MinIO (replace slashes)
        safe_container_no = container_no.replace('/', '_') if container_no else ''
        
        # Prepare list of parameters for bulk execution
        params_list = []
        for item in item_details:
            row_params = (
                doc_no,
                date_val,
                container_no,  # Store original in database
                gatepass_no,
                bayan_no,
                item.get('po_number'),
                item.get('Item_Code'),
                item.get('Item_Desc'),
                item.get('PO_QTY'),
                item.get('Prod_Part'),
                supplier_name,
                supplier_no,
                container_status,  # Use renamed variable
                inbound_initiator_name,
                inbound_initiator_no,
                inbound_receiver_no,
                inbound_receiver_name,
                gatepass_generated_by_no,
                gatepass_generated_by_name,
                remarks,
                image_count,
                minioimage_location,  # This should be the cleaned version
                creation_by,
                creation_ip
            )
            params_list.append(row_params)

        # Raw SQL Insert
        sql = """
            INSERT INTO   [Inbound_Receiver_Container_Qc_Tbl] (
                [Doc_No],
                [Date],
                [Container_No],
                [GatePass_No],
                [Bayan_No],
                [PO_Number],
                [Item_Code],
                [Item_Desc],
                [PO_QTY],
                [Prod_Part],
                [Supplier_Name],
                [Supplier_No],
                [Status],
                [Inbound_Initiater_Name],
                [Inbound_Initiator_No],
                [Inbound_Reciver_No],
                [Inound_Reciver_Name],
                [GatePassGeneratedBy_No],
                [GatePassGeneartedBy_Name],
                [Remarks],
                [Image_Count],
                [MinioImage_Location],
                [Creation_by],
                [Creation_date],
                [Creation_Ip]
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, GETDATE(), %s)
        """

        with connections['Inbound_db'].cursor() as cursor:
            # 1. Bulk Insert
            cursor.executemany(sql, params_list)
            
            # 2. Update Status in WHR_Inbound_Genrate_GatePass_tbl
            update_sql = """
                UPDATE   [WHR_Inbound_Genrate_GatePass_tbl]
                SET [Status] = 'Container_QC'
                WHERE [Doc_No] = %s 
                  AND [Container_No] = %s  -- Match with original container number
                  AND [GatePass_No] = %s
            """
            cursor.execute(update_sql, [doc_no, container_no, gatepass_no])
            rows_updated = cursor.rowcount

        return Response({
            "status": "success", 
            "message": "Data inserted and status updated successfully",
            "rows_inserted": len(params_list),
            "status_updated_rows": rows_updated,
            "image_count": image_count,
            "minio_location": minioimage_location,
            "container_no_original": container_no,
            "container_no_safe": safe_container_no
        }, status=status.HTTP_201_CREATED)

    except Exception as e:
        import traceback
        logger.error(f"Bulk insert failed: {str(e)}")
        logger.error(traceback.format_exc())
        return Response({
            "status": "error", 
            "message": f"Insert failed: {str(e)}"
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
# -------------------------------------------------------------------------
# INBOUND PRODUCT DAMAGE INSERT VIEW (Raw SQL)
# -------------------------------------------------------------------------
@api_view(['POST'])
def inbound_product_damage_insert(request):
    """
    Inserts product damage details and updates related tables.
    Logic:
    1. Parse Header Data (Common for all items).
    2. Parse Item_Details (List of specific damage items).
    3. Bulk Insert into [Inbound_Receiver_Product_Damage_tbl].
    4. Update [WHR_Inbound_Genrate_GatePass_tbl] status to 'Damaged'.
    5. Update [Inbound_Receiver_Container_Qc_Tbl] Po_Damage_Qty.
    """
    try:
        data = request.data
        
        # --- 1. Header Level Data extraction ---
        doc_no = data.get('Doc_No')
        container_no = data.get('Container_No')
        gate_pass_no = data.get('GatePass_No')
        
        # Validation
        if not all([doc_no, container_no, gate_pass_no]):
             return Response({"status": "error", "message": "Doc_No, Container_No, and Gate_Pass_No are required"}, status=status.HTTP_400_BAD_REQUEST)

        date_val = data.get('Date')
        po_number = data.get('Po_Number')
        bayan_no = data.get('Bayan_No')
        item_code = data.get('Item_Code')
        item_desc = data.get('ItemDesc')
        po_qty = data.get('PO_QTY')
        prod_part = data.get('Prod_Part')
        supplier_no = data.get('Supplier_No')
        supplier_name = data.get('Supplier_Name')
        inbound_initiator_no = data.get('Inbound_Initiator_No')
        inbound_initiator_name = data.get('Inbound_Initiator_Name')
        receiver_no = data.get('Receiver_No')
        receiver_name = data.get('Recevicer_Name')
        gp_gen_by_no = data.get('Gate_Pass_Generated_ByNo')
        gp_gen_by_name = data.get('Gate_Pass_Generated_ByName')
        po_damage_qty = data.get('Po_Damage_QTY') # Total damage qty for the QC update? Or per item? 
                                                  # Usually typical to come from header if used in update query.
        po_scanned_qty = data.get('Po_Scanned_Qty')
        # ... other header fields ...
        status_val = data.get('Status', 'Damaged') # Default to Damaged if not sent
        created_by = data.get('Created_By')
        created_ip = request.META.get('REMOTE_ADDR')

        # --- 2. Item Details Extraction ---
        item_details = data.get('Item_Details', [])
        if not item_details or not isinstance(item_details, list):
             return Response({"status": "error", "message": "Item_Details must be a non-empty list"}, status=status.HTTP_400_BAD_REQUEST)

        # Prepare Bulk Insert Params
        current_date = datetime.now() # Captured once to ensure consistency across Insert and Updates
        
        insert_params = []
        for item in item_details:
            row_params = [
                doc_no, date_val, container_no, gate_pass_no, po_number, 
                bayan_no, item_code, item_desc, po_qty, prod_part, 
                supplier_no, supplier_name, inbound_initiator_no, inbound_initiator_name, 
                receiver_no, receiver_name, gp_gen_by_no, gp_gen_by_name, 
                po_damage_qty, po_scanned_qty,
                
                # Item Specific Fields
                item.get('Damage_Type'),
                item.get('Damage_ProductNo'),
                item.get('Damage_serailNo'), 
                item.get('Damage_Image_Count'),
                item.get('Damage_Image_Location'),
                
                status_val,
                created_by, current_date, created_ip 
            ]
            insert_params.append(row_params)

        with connections['Inbound_db'].cursor() as cursor:
            # --- 3. Bulk Insert ---
            insert_sql = """
                INSERT INTO   [Inbound_Receiver_Product_Damage_tbl] (
                    [Doc_No], [Date], [Container_No], [GatePass_No], [Po_Number],
                    [Bayan_No], [Item_Code], [ItemDesc], [PO_QTY], [Prod_Part],
                    [Supplier_No], [Supplier_Name], [Inbound_Initiator_No], [Inbound_Initiator_Name],
                    [Receiver_No], [Recevicer_Name], [Gate_Pass_Generated_ByNo], [Gate_Pass_Generated_ByName],
                    [Po_Damage_QTY], [Po_Scanned_Qty],
                    [Damage_Type], [Damage_ProductNo], [Damage_serailNo], [Damage_Image_Count], [Damage_Image_Location],
                    [Status], [Created_By], [Created_Date], [Created_IP]
                )
                VALUES (
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s,
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, %s
                )
            """
            cursor.executemany(insert_sql, insert_params)

            # --- 4. Update Gate Pass Status ---
            # Reuse created_by (as updation_By) and current_date (as Updation_Date)
            update_gp_sql = """
                UPDATE   [WHR_Inbound_Genrate_GatePass_tbl]
                SET [Damage_Status] = 'Damaged',
                    [updation_By] = %s,
                    [Updation_Date] = %s
                WHERE [Doc_No] = %s
                  AND [Container_No] = %s
                  AND [GatePass_No] = %s
            """
            cursor.execute(update_gp_sql, [created_by, current_date, doc_no, container_no, gate_pass_no])

            # --- 5. Update QC Table Damage Qty ---
            # Reuse created_by (as updation_By) and current_date (as Updation_Date)
            update_qc_sql = """
                UPDATE   [Inbound_Receiver_Container_Qc_Tbl]
                SET [Po_Damage_Qty] = %s,
                    [updation_By] = %s,
                    [Updation_Date] = %s
                WHERE [Doc_No] = %s
                  AND [Container_No] = %s
                  AND [GatePass_No] = %s
            """
            cursor.execute(update_qc_sql, [po_damage_qty, created_by, current_date, doc_no, container_no, gate_pass_no])
            
        return Response({
            "status": "success",
            "message": "Damage details inserted and status updated successfully",
            "rows_inserted": len(insert_params)
        }, status=status.HTTP_201_CREATED)

    except Exception as e:
        return Response(
            {"status": "error", "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.db import connections
@api_view(['POST'])
def get_damage_qty(request):
    """
    Returns the total damage quantity for a specific item in a container.
    """
    try:
        data = request.data
        doc_no = data.get('Doc_No')
        container_no = data.get('Container_No')
        gate_pass_no = data.get('GatePass_No')
        item_code = data.get('Item_Code')
        po_number = data.get('Po_Number')
        
        if not all([doc_no, container_no, gate_pass_no, item_code]):
             return Response({"status": "error", "message": "Missing required fields"}, status=400)
        with connections['Inbound_db'].cursor() as cursor:
            sql = """
                SELECT COUNT(*) as damage_qty
                FROM [Inbound_Receiver_Product_Damage_tbl]
                WHERE [Doc_No] = %s
                  AND [Container_No] = %s
                  AND [GatePass_No] = %s
                  AND [Item_Code] = %s
                  AND [Po_Number] = %s
            """
            cursor.execute(sql, [doc_no, container_no, gate_pass_no, item_code, po_number])
            result = cursor.fetchone()
            qty = result[0] if result else 0
            
        return Response({
            "status": "success",
            "damage_qty": qty,
            "item_code": item_code
        })
    except Exception as e:
        return Response({"status": "error", "message": str(e)}, status=500)



logger = logging.getLogger(__name__)

@api_view(['POST'])
@permission_classes([AllowAny])
def upload_container_images_to_minio(request):
    """
    Upload container QC images to MinIO bucket
    Preserves container numbers with slashes like 'GCXU5089740/40HQ'
    """
    try:
        data = request.data
        gatepass_no = data.get('gatepass_no', '').strip()
        container_no = data.get('container_no', '').strip()
        doc_no = data.get('doc_no', '').strip()
        images = data.get('images', [])

        if not gatepass_no or not container_no:
            return Response({
                'status': 'error',
                'message': 'Gatepass number and Container number are required'
            }, status=status.HTTP_400_BAD_REQUEST)

        if not images:
            # No images to upload
            return Response({
                'status': 'success',
                'message': 'No images to upload',
                'image_count': 0,
                'minio_location': '',
                'urls': []
            }, status=status.HTTP_200_OK)

        # Initialize S3 client for MinIO
        endpoint = f"http{'s' if settings.MINIO_SECURE else ''}://{settings.MINIO_ENDPOINT}"
        
        s3 = boto3.client(
            's3',
            endpoint_url=endpoint,
            aws_access_key_id=settings.MINIO_ACCESS_KEY,
            aws_secret_access_key=settings.MINIO_SECRET_KEY,
            config=boto3.session.Config(signature_version='s3v4')
        )

        uploaded_urls = []
        
        # Keep the original container number with slash
        # MinIO will treat slash as folder separator
        folder_name = f"{gatepass_no}/{container_no}"
        
        # MinIO location path for database (using original container number)
        minio_location = f"{gatepass_no}/{container_no}/"

        for image_data in images:
            try:
                image_index = image_data.get('image_index', 1)
                image_base64 = image_data.get('image_data', '')
                original_name = image_data.get('image_name', f'image_{image_index}')

                if not image_base64:
                    continue

                # Decode base64 image
                image_bytes = base64.b64decode(image_base64)
                
                # Generate unique filename
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                unique_id = str(uuid.uuid4())[:8]
                file_extension = 'jpg'  # Default extension
                
                # Try to get extension from original name
                if '.' in original_name:
                    file_extension = original_name.split('.')[-1].lower()
                
                # Create filename
                file_name = f"{timestamp}_{unique_id}.{file_extension}"
                key = f"{folder_name}/{file_name}"
                
                # Upload to MinIO
                s3.put_object(
                    Bucket=settings.MINIO_INBOUND_CONTAINER_QC_BUCKET_NAME,
                    Key=key,
                    Body=BytesIO(image_bytes),
                    ContentType=f'image/{file_extension}',
                    ACL='public-read'
                )

                # Generate public URL
                image_url = f"{endpoint}/{settings.MINIO_INBOUND_CONTAINER_QC_BUCKET_NAME}/{key}"
                
                uploaded_urls.append({
                    'url': image_url,
                    'minio_path': key,
                    'original_name': original_name,
                    'image_index': image_index,
                    'size': len(image_bytes),
                })

                logger.info(f"Uploaded to MinIO: {key}")

            except Exception as img_error:
                logger.error(f"Error uploading image {image_index}: {str(img_error)}")
                continue

        # Return only what's needed for main QC table
        return Response({
            'status': 'success',
            'message': f'Successfully uploaded {len(uploaded_urls)} images',
            'image_count': len(uploaded_urls),
            'minio_location': minio_location,  # Make sure this is returned
            'urls': uploaded_urls,
            'gatepass_no': gatepass_no,
            'container_no': container_no,  # Original with slash
            'doc_no': doc_no
        }, status=status.HTTP_200_OK)

    except Exception as e:
        logger.error(f"Upload error: {str(e)}")
        return Response({
            'status': 'error',
            'message': f'Upload failed: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# @csrf_exempt
# @require_http_methods(["POST"])  # Change from GET to POST
# def get_product_barcode(request):
#     try:
#         data = json.loads(request.body)
#         item_code = data.get("ITEM_CODE")
        
#         if not item_code:
#             return JsonResponse(
#                 {"status": "error", "message": "ITEM_CODE is required"},
#                 status=400
#             )

#         query = """
#             SELECT TOP 1 PRODUCT_BARCODE, SERIAL_STATUS
#             FROM BUYP.ALJE_ITEM_CATEGORIES_CPD_V
#             WHERE ITEM_CODE = %s
#         """

#         with connections['Inbound_db'].cursor() as cursor:
#             cursor.execute(query, [item_code])
#             row = cursor.fetchone()

#         if row and row[0]:
#             return JsonResponse({
#                 "status": "success",
#                 "ITEM_CODE": item_code,
#                 "PRODUCT_BARCODE": row[0]
#             })

#         return JsonResponse({
#             "status": "success",
#             "ITEM_CODE": item_code,
#             "PRODUCT_BARCODE": None,
#             "message": "No product code found"
#         })

#     except json.JSONDecodeError:
#         return JsonResponse(
#             {"status": "error", "message": "Invalid JSON"},
#             status=400
#         )
#     except Exception as e:
#         return JsonResponse(
#             {"status": "error", "message": str(e)},
#             status=500
#         )

@csrf_exempt
@require_http_methods(["POST"])
def get_product_barcode(request):
    try:
        data = json.loads(request.body)
        item_code = data.get("ITEM_CODE")

        if not item_code:
            return JsonResponse(
                {"status": "error", "message": "ITEM_CODE is required"},
                status=400
            )

        query = """
            SELECT TOP 1
                CASE
                    WHEN PRODUCT_BARCODE IS NULL
                         OR LTRIM(RTRIM(PRODUCT_BARCODE)) = ''
                         OR LTRIM(RTRIM(PRODUCT_BARCODE)) IN ('0', '00', '000', '0000', '00000')
                         OR PRODUCT_BARCODE NOT LIKE '%%[^0]%%'
                    THEN '00'
                    ELSE LTRIM(RTRIM(PRODUCT_BARCODE))
                END AS PRODUCT_BARCODE,
                ISNULL(SERIAL_STATUS, 'N') AS SERIAL_STATUS
            FROM BUYP.ALJE_ITEM_CATEGORIES_CPD_V
            WHERE ITEM_CODE = %s
        """

        with connections['default'].cursor() as cursor:
            cursor.execute(query, [item_code])
            row = cursor.fetchone()

        if row:
            product_barcode = row[0]
            serial_status = row[1]
            
            # If product_barcode is '00', it means accept any barcode
            return JsonResponse({
                "status": "success",
                "ITEM_CODE": item_code,
                "PRODUCT_BARCODE": product_barcode,
                "SERIAL_STATUS": serial_status,
                "message": "Any barcode accepted" if product_barcode == '00' else "Product barcode found"
            })

        return JsonResponse({
            "status": "success",
            "ITEM_CODE": item_code,
            "PRODUCT_BARCODE": "00",
            "SERIAL_STATUS": "N",
            "message": "No product found - any barcode accepted"
        })

    except json.JSONDecodeError:
        return JsonResponse(
            {"status": "error", "message": "Invalid JSON"},
            status=400
        )
    except Exception as e:
        return JsonResponse(
            {"status": "error", "message": str(e)},
            status=500
        )


@api_view(['POST'])
def get_received_qty(request):
    try:
        # Filter strictly by the key identifiers for this Gate Pass Line Item
        doc_no = request.data.get('Doc_No')
        container_no = request.data.get('Container_No')
        gate_pass_no = request.data.get('GatePass_No')
        po_number = request.data.get('Po_Number')
        item_code = request.data.get('Item_Code')
        
        # Validation
        if not all([doc_no, container_no, gate_pass_no, po_number, item_code]):
            # If any key is missing, return 0 or error. Returning 0 is safer for now.
             return Response({"status": "error", "message": "Missing keys for lookup"}, status=400)
        with connections['Inbound_db'].cursor() as cursor:
            # Count rows in your Product Scanning Table
            cursor.execute("""
                SELECT COUNT(*) 
                FROM [Inbound_Receiver_Product_Scaning_tbl]
                WHERE [Doc_No] = %s
                  AND [Container_No] = %s
                  AND [GatePass_No] = %s
                  AND [Po_Number] = %s
                  AND [Item_Code] = %s
            """, [doc_no, container_no, gate_pass_no, po_number, item_code])
            
            row = cursor.fetchone()
            total_received = row[0] if row else 0
            
        return Response({
            "status": "success",
            "received_qty": total_received
        })
    except Exception as e:
        return Response({"status": "error", "message": str(e)}, status=500)

# -------------------------------------------------------------------------
# INBOUND PRODUCT SCANNING INSERT VIEW (New Requirement)
# -------------------------------------------------------------------------
@api_view(['POST'])
def inbound_product_scanning_insert(request):
    """
    Inserts product scanning details and updates related tables.
    Logic:
    1. Parse Header Data.
    2. Parse Item_Details.
    3. Bulk Insert into [Inbound_Receiver_Product_Scaning_tbl].
    4. Update [Inbound_Receiver_Container_Qc_Tbl] Po_Scanned_Qty.
    5. Conditional Status Update based on total quantity check.
    """
    try:
        data = request.data
        
        # --- 1. Header Level Data extraction ---
        doc_no = data.get('Doc_No')
        container_no = data.get('Container_No')
        gate_pass_no = data.get('GatePass_No')
        item_code = data.get('Item_Code')  # This is important!
        
        # Validation
        if not all([doc_no, container_no, gate_pass_no, item_code]):
            return Response({
                "status": "error", 
                "message": "Doc_No, Container_No, Gate_Pass_No and Item_Code are required"
            }, status=status.HTTP_400_BAD_REQUEST)

        date_val = data.get('Date')
        po_number = data.get('Po_Number')
        bayan_no = data.get('Bayan_No')
        item_desc = data.get('ItemDesc')
        po_qty = data.get('PO_QTY')
        prod_part = data.get('Prod_Part')
        supplier_no = data.get('Supplier_No')
        supplier_name = data.get('Supplier_Name')
        inbound_initiator_no = data.get('Inbound_Initiator_No')
        inbound_initiator_name = data.get('Inbound_Initiator_Name')
        receiver_no = data.get('Receiver_No')
        receiver_name = data.get('Recevicer_Name')
        gp_gen_by_no = data.get('Gate_Pass_Generated_ByNo')
        gp_gen_by_name = data.get('Gate_Pass_Generated_ByName')
        
        po_scanned_qty = data.get('Po_Scanned_Qty')
        status_val = data.get('Status', 'Scanned') 
        created_by = data.get('Created_By')
        created_ip = request.META.get('REMOTE_ADDR')

        # --- 2. Item Details Extraction ---
        item_details = data.get('Item_Details', [])
        if not item_details or not isinstance(item_details, list):
            return Response({
                "status": "error", 
                "message": "Item_Details must be a non-empty list"
            }, status=status.HTTP_400_BAD_REQUEST)

        # Prepare Bulk Insert Params
        current_date = datetime.now()
        
        insert_params = []
        for item in item_details:
            row_params = [
                doc_no, date_val, container_no, gate_pass_no, po_number, 
                bayan_no, item_code, item_desc, po_qty, prod_part, 
                supplier_no, supplier_name, inbound_initiator_no, inbound_initiator_name, 
                receiver_no, receiver_name, gp_gen_by_no, gp_gen_by_name, 
                po_scanned_qty,
                
                # Item Specific Fields
                item.get('ProductNo'),
                item.get('SerailNo'), 
                
                status_val,
                created_by, current_date, created_ip 
            ]
            insert_params.append(row_params)

        with connections['Inbound_db'].cursor() as cursor:
            # --- 3. Bulk Insert ---
            insert_sql = """
                INSERT INTO [Inbound_Receiver_Product_Scaning_tbl] (
                    [Doc_No], [Date], [Container_No], [GatePass_No], [Po_Number],
                    [Bayan_No], [Item_Code], [ItemDesc], [PO_QTY], [Prod_Part],
                    [Supplier_No], [Supplier_Name], [Inbound_Initiator_No], [Inbound_Initiator_Name],
                    [Receiver_No], [Recevicer_Name], [Gate_Pass_Generated_ByNo], [Gate_Pass_Generated_ByName],
                    [Po_Scanned_Qty],
                    [ProductNo], [SerailNo],
                    [Status], [Created_By], [Created_Date], [Created_IP]
                )
                VALUES (
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s, %s,
                    %s,
                    %s, %s,
                    %s, %s, %s, %s
                )
            """
            cursor.executemany(insert_sql, insert_params)

            # --- 4. Update QC Table Scanned Qty (with Item_Code) ---
            update_qc_sql = """
                UPDATE qc
                SET qc.Po_Scanned_Qty = (
                    SELECT COUNT(*)
                    FROM Inbound_Receiver_Product_Scaning_tbl ps
                    WHERE ps.Doc_No = qc.Doc_No
                    AND ps.Container_No = qc.Container_No
                    AND ps.GatePass_No = qc.GatePass_No
                    AND ps.Item_Code = qc.Item_Code  -- Added Item_Code condition
                ),
                [updation_By] = %s,
                [Updation_Date] = %s
                FROM Inbound_Receiver_Container_Qc_Tbl qc
                WHERE qc.Doc_No = %s
                AND qc.Container_No = %s
                AND qc.GatePass_No = %s
                AND qc.Item_Code = %s;  -- Added Item_Code condition
            """
            cursor.execute(update_qc_sql, [
                created_by, current_date, 
                doc_no, container_no, gate_pass_no, item_code
            ])
            
            # --- 5. Post-Update Check & Conditional Status Update ---
            check_sql = """
                SELECT 
                    CASE 
                        WHEN PO_QTY <= SUM(ISNULL(Po_Damage_Qty, 0) + ISNULL(Po_Scanned_Qty, 0))
                            THEN 'Product_Scanned_Finished'
                        ELSE 'Product_Scanned_Pending'
                    END AS Scan_Status
                FROM Inbound_Receiver_Container_Qc_Tbl
                WHERE Doc_No = %s
                AND Container_No = %s
                AND GatePass_No = %s
                AND Item_Code = %s  -- Added Item_Code condition
                GROUP BY PO_QTY
            """
            cursor.execute(check_sql, [doc_no, container_no, gate_pass_no, item_code])
            row = cursor.fetchone()
            
            scan_status = row[0] if row else None
            
            # --- 6. Update GatePass table status (with Item_Code) ---
            if scan_status == 'Product_Scanned_Finished':
                update_status_sql = """
                    UPDATE [WHR_Inbound_Genrate_GatePass_tbl]
                    SET [Status] = 'Product_Scanned_Finished'
                    WHERE [Doc_No] = %s
                    AND [Container_No] = %s
                    AND [GatePass_No] = %s
                    AND [Item_Code] = %s  -- CRITICAL: Added Item_Code condition
                """
                cursor.execute(update_status_sql, [
                    doc_no, container_no, gate_pass_no, item_code
                ])
            
        return Response({
            "status": "success",
            "message": "Scanning details inserted and status checked successfully",
            "scan_status": scan_status,
            "rows_inserted": len(insert_params),
            "item_code": item_code  # Added for clarity
        }, status=status.HTTP_201_CREATED)

    except Exception as e:
        import traceback
        return Response(
            {"status": "error", "message": str(e), "trace": traceback.format_exc()},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
def upload_product_damage_images_to_minio(request):
    """
    Upload product damage images to MinIO bucket
    Each image is linked to a specific damage item with serial number
    """
    try:
        data = request.data
        gatepass_no = data.get('gatepass_no', '').strip()
        container_no = data.get('container_no', '').strip()
        doc_no = data.get('doc_no', '').strip()
        po_number = data.get('po_number', '').strip()
        item_code = data.get('item_code', '').strip()
        damage_serial_no = data.get('damage_serial_no', '').strip()
        product_no = data.get('product_no', '').strip()
        images = data.get('images', [])

        if not gatepass_no or not container_no:
            return Response({
                'status': 'error',
                'message': 'Gatepass number and Container number are required'
            }, status=status.HTTP_400_BAD_REQUEST)

        if not images:
            return Response({
                'status': 'success',
                'message': 'No images to upload',
                'image_count': 0,
                'minio_location': '',
                'urls': []
            }, status=status.HTTP_200_OK)

        # Initialize S3 client for MinIO
        endpoint = f"http{'s' if settings.MINIO_SECURE else ''}://{settings.MINIO_ENDPOINT}"
        
        s3 = boto3.client(
            's3',
            endpoint_url=endpoint,
            aws_access_key_id=settings.MINIO_ACCESS_KEY,
            aws_secret_access_key=settings.MINIO_SECRET_KEY,
            config=Config(signature_version='s3v4')
        )

        uploaded_urls = []
        
        # Create organized folder structure
        # Format: gatepass/container/po/item/serial/
        folder_name = f"{gatepass_no}/{container_no}/{po_number}/{item_code}/{damage_serial_no}/{product_no}"
        
        # MinIO location path for database
        minio_location = f"{folder_name}/"

        for image_data in images:
            try:
                image_index = image_data.get('image_index', 1)
                image_base64 = image_data.get('image_data', '')
                original_name = image_data.get('image_name', f'damage_image_{image_index}')
                damage_type = image_data.get('damage_type', 'Unknown')

                if not image_base64:
                    continue

                # Decode base64 image
                image_bytes = base64.b64decode(image_base64)
                
                # Generate unique filename
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                unique_id = str(uuid.uuid4())[:8]
                file_extension = 'jpg'
                
                # Try to get extension from original name
                if '.' in original_name:
                    file_extension = original_name.split('.')[-1].lower()
                    if file_extension not in ['jpg', 'jpeg', 'png', 'gif', 'bmp']:
                        file_extension = 'jpg'
                
                # Create filename with damage type
                file_name = f"{damage_type}_{timestamp}_{unique_id}.{file_extension}"
                key = f"{folder_name}/{file_name}"
                
                # Upload to MinIO
                s3.put_object(
                    Bucket=settings.MINIO_INBOUND_PRODUCT_DAMAGE_BUCKET_NAME,
                    Key=key,
                    Body=BytesIO(image_bytes),
                    ContentType=f'image/{file_extension}',
                    ACL='public-read'
                )

                # Generate public URL
                image_url = f"{endpoint}/{settings.MINIO_INBOUND_PRODUCT_DAMAGE_BUCKET_NAME}/{key}"
                
                uploaded_urls.append({
                    'url': image_url,
                    'minio_path': key,
                    'original_name': original_name,
                    'image_index': image_index,
                    'damage_type': damage_type,
                    'size': len(image_bytes),
                })

                logger.info(f"Uploaded product damage image to MinIO: {key}")

            except Exception as img_error:
                logger.error(f"Error uploading product damage image {image_index}: {str(img_error)}")
                continue

        return Response({
            'status': 'success',
            'message': f'Successfully uploaded {len(uploaded_urls)} damage images',
            'image_count': len(uploaded_urls),
            'minio_location': minio_location,
            'urls': uploaded_urls,
            'gatepass_no': gatepass_no,
            'container_no': container_no,
            'doc_no': doc_no,
            'po_number': po_number,
            'item_code': item_code,
            'damage_serial_no': damage_serial_no,
            'product_no': product_no
        }, status=status.HTTP_200_OK)

    except Exception as e:
        logger.error(f"Product damage upload error: {str(e)}")
        return Response({
            'status': 'error',
            'message': f'Upload failed: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



@csrf_exempt
def check_wh_transportation_status(request):
    """
    Checks if a record exists in WHR_INBOUND_BAYAN_HEADER_TBL with specific DOC_NO
    and STATUS in ['WH Transportation', 'Pending Offloading', 'Pending GRN', 'Completed'].
    Expects GET request with query parameter: ?doc_no=...
    Returns JSON: {'display': True, 'status': 'Found status'} or {'display': False}
    """
    if request.method != 'GET':
        return JsonResponse({'status': 'error', 'message': 'GET method required'}, status=405)
 
    try:
        # Get doc_no from query parameters (GET request)
        doc_no = request.GET.get('doc_no')
       
        if not doc_no:
            return JsonResponse({'status': 'error', 'message': 'Missing doc_no parameter'}, status=400)
 
        # Statuses to check
        valid_statuses = ['WH Transportation', 'Pending Offloading', 'Pending GRN', 'Completed']
        
        # Create placeholders for SQL IN clause
        placeholders = ','.join(['%s'] * len(valid_statuses))
        
        # Query to get both existence and the actual status
        sql = f"""
            SELECT TOP 1 STATUS
            FROM [WHR_INBOUND_BAYAN_HEADER_TBL]
            WHERE DOC_NO = %s
              AND STATUS IN ({placeholders});
        """
        
        with connections['Inbound_db'].cursor() as cursor:
            # Prepare parameters: doc_no + all statuses
            params = [doc_no] + valid_statuses
            cursor.execute(sql, params)
            row = cursor.fetchone()
           
        if row:
            return JsonResponse({
                'display': True,
                'status': row[0]  # Return the actual found status
            })
        else:
            return JsonResponse({'display': False})
 
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    
@csrf_exempt
def get_consolidated_gatepass_data(request):
    try:
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 50))
        offset_val = (page - 1) * page_size

        with connections['Inbound_db'].cursor() as cursor:
            # ---------------------------------------------------------
            # STEP 1: Get DISTINCT GatePass_No AND Date for Pagination
            # ---------------------------------------------------------
            # Added master.Creaetion_Date to the SELECT list
            pagination_sql = """
                WITH AllGatePasses AS (
                    SELECT DISTINCT TRIM(GatePass_No) as GatePass_No FROM [Inbound_Receiver_Container_Qc_Tbl] WHERE GatePass_No IS NOT NULL
                    UNION
                    SELECT DISTINCT TRIM(GatePass_No) as GatePass_No FROM [Inbound_Receiver_Product_Scaning_tbl] WHERE GatePass_No IS NOT NULL
                    UNION
                    SELECT DISTINCT TRIM(GatePass_No) as GatePass_No FROM [Inbound_Receiver_Product_Damage_tbl] WHERE GatePass_No IS NOT NULL
                )
                SELECT DISTINCT agp.GatePass_No, master.Creaetion_Date
                FROM AllGatePasses agp
                INNER JOIN [WHR_Inbound_Genrate_GatePass_tbl] master
                    ON agp.GatePass_No = TRIM(master.GatePass_No) COLLATE DATABASE_DEFAULT
                WHERE TRIM(master.Status) IN ('Product_Scanned_Finished', 'Container_QC')
                ORDER BY agp.GatePass_No DESC
                OFFSET %s ROWS FETCH NEXT %s ROWS ONLY
            """
            cursor.execute(pagination_sql, [offset_val, page_size])
            gp_rows = cursor.fetchall()
            
            if not gp_rows:
                return JsonResponse({"status": "success", "page": page, "count": 0, "data": []})

            # Create a map of GatePassNo -> Date
            # row[0] is GatePass_No, row[1] is Creaetion_Date
            gatepass_list = []
            gatepass_dates = {}
            for row in gp_rows:
                gp_no = row[0]
                gatepass_list.append(gp_no)
                gatepass_dates[gp_no] = str(row[1]) if row[1] else "N/A"

            placeholders = ','.join(['%s'] * len(gatepass_list))
            
            # ---------------------------------------------------------
            # STEP 2: Fetch DETAILED Data
            # ---------------------------------------------------------
            # Note: We don't need to select [Entry_Date] here anymore 
            # because we got it from the master table in Step 1.
            details_sql = f"""
                /* QC Table */
                SELECT 
                    'QC' as SourceTable,
                    [Doc_No], [Container_No], [GatePass_No], [PO_Number], [Bayan_No], [PO_QTY],
                    [Supplier_Name], [Supplier_No], 
                    [Inbound_Initiator_No], [Inbound_Initiater_Name] as Inbound_Initiator_Name,
                    [Inbound_Reciver_No] as Receiver_No, [Inound_Reciver_Name] as Recevicer_Name,
                    [GatePassGeneratedBy_No] as Gate_Pass_Generated_ByNo,
                    [GatePassGeneartedBy_Name] as Gate_Pass_Generated_ByName,
                    [Item_Code], [Item_Desc], [Prod_Part],
                    [Po_Damage_Qty] as Po_Damage_QTY, [Po_Scanned_Qty],
                    NULL as Damage_Image_Count, NULL as Damage_Type, NULL as Damage_ProductNo, NULL as Damage_SerialNo,
                    [Status], NULL as ProductNo, NULL as SerialNo, [Image_Count]
                FROM [Inbound_Receiver_Container_Qc_Tbl]
                WHERE TRIM(GatePass_No) IN ({placeholders})

                UNION ALL

                /* Scanning Table */
                SELECT 
                    'SCAN' as SourceTable,
                    [Doc_No], [Container_No], [GatePass_No], [Po_Number], [Bayan_No], [PO_QTY],
                    [Supplier_Name], [Supplier_No],
                    [Inbound_Initiator_No], [Inbound_Initiator_Name],
                    [Receiver_No], [Recevicer_Name],
                    [Gate_Pass_Generated_ByNo], [Gate_Pass_Generated_ByName],
                    [Item_Code], [ItemDesc] as Item_Desc, [Prod_Part],
                    NULL as Po_Damage_QTY, [Po_Scanned_Qty],
                    NULL as Damage_Image_Count, NULL as Damage_Type, NULL as Damage_ProductNo, NULL as Damage_SerialNo,
                    [Status], [ProductNo], [SerailNo] as SerialNo, NULL as Image_Count
                FROM [Inbound_Receiver_Product_Scaning_tbl]
                WHERE TRIM(GatePass_No) IN ({placeholders})

                UNION ALL

                /* Damage Table */
                SELECT 
                    'DAMAGE' as SourceTable,
                    [Doc_No], [Container_No], [GatePass_No], [Po_Number], [Bayan_No], [PO_QTY],
                    [Supplier_Name], [Supplier_No],
                    [Inbound_Initiator_No], [Inbound_Initiator_Name],
                    [Receiver_No], [Recevicer_Name],
                    [Gate_Pass_Generated_ByNo], [Gate_Pass_Generated_ByName],
                    [Item_Code], [ItemDesc] as Item_Desc, [Prod_Part],
                    [Po_Damage_QTY], [Po_Scanned_Qty],
                    [Damage_Image_Count], [Damage_Type], [Damage_ProductNo], [Damage_serailNo] as Damage_SerialNo,
                    [Status], NULL as ProductNo, NULL as SerialNo, NULL as Image_Count
                FROM [Inbound_Receiver_Product_Damage_tbl]
                WHERE TRIM(GatePass_No) IN ({placeholders})
            """
            
            cursor.execute(details_sql, gatepass_list * 3)
            columns = [col[0] for col in cursor.description]
            all_rows = cursor.fetchall()

        # ---------------------------------------------------------
        # STEP 3: Aggregation
        # ---------------------------------------------------------
        def clean_str(val):
            return val.strip() if isinstance(val, str) else val
            
        gatepass_headers = {}
        gatepass_items = {gp: [] for gp in gatepass_list}

        for row in all_rows:
            row_dict = dict(zip(columns, row))
            gp = row_dict.get('GatePass_No')
            
            if gp not in gatepass_headers:
                gatepass_headers[gp] = {
                    # USE THE DATE WE FETCHED FROM MASTER TABLE
                    "Entry_Date": gatepass_dates.get(gp, "N/A"), 
                    "Doc_No": clean_str(row_dict.get('Doc_No')),
                    "Container_No": clean_str(row_dict.get('Container_No')),
                    "GatePass_No": clean_str(row_dict.get('GatePass_No')),
                    "Po_Number": clean_str(row_dict.get('PO_Number')),
                    "Bayan_No": clean_str(row_dict.get('Bayan_No')),
                    "PO_QTY": row_dict.get('PO_QTY'),
                    "Supplier_Name": clean_str(row_dict.get('Supplier_Name')),
                    "Supplier_No": clean_str(row_dict.get('Supplier_No')),
                    "Inbound_Initiator_No": clean_str(row_dict.get('Inbound_Initiator_No')),
                    "Inbound_Initiator_Name": clean_str(row_dict.get('Inbound_Initiator_Name')),
                    "Receiver_No": clean_str(row_dict.get('Receiver_No')),
                    "Receiver_Name": clean_str(row_dict.get('Recevicer_Name')),
                    "Gate_Pass_Generated_ByNo": clean_str(row_dict.get('Gate_Pass_Generated_ByNo')),
                    "Gate_Pass_Generated_ByName": clean_str(row_dict.get('Gate_Pass_Generated_ByName')),
                }
            
            item = {
                "Item_Code": clean_str(row_dict.get('Item_Code')),
                "Item_Desc": clean_str(row_dict.get('Item_Desc')),
                "Prod_Part": clean_str(row_dict.get('Prod_Part')),
                "PO_QTY": row_dict.get('PO_QTY'),
                "Po_Damage_QTY": row_dict.get('Po_Damage_QTY'),
                "Po_Scanned_Qty": row_dict.get('Po_Scanned_Qty'),
                "Damage_Image_Count": row_dict.get('Damage_Image_Count'),
                "Damage_Type": clean_str(row_dict.get('Damage_Type')),
                "Damage_ProductNo": clean_str(row_dict.get('Damage_ProductNo')),
                "Damage_SerialNo": clean_str(row_dict.get('Damage_SerialNo')),
                "Status": clean_str(row_dict.get('Status')),
                "ProductNo": clean_str(row_dict.get('ProductNo')),
                "SerialNo": clean_str(row_dict.get('SerialNo')),
                "Image_Count": row_dict.get('Image_Count'),
                "SourceTable": clean_str(row_dict.get('SourceTable'))
            }
            if gp in gatepass_items:
                gatepass_items[gp].append(item)

        final_response_data = []
        for gp in gatepass_list:
            header = gatepass_headers.get(gp, {})
            header["ItemDetails"] = gatepass_items.get(gp, [])
            final_response_data.append(header)

        return JsonResponse({"status": "success", "page": page, "count": len(final_response_data), "data": final_response_data})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


# views.py
from django.db import connection
from django.http import JsonResponse
import json
from decimal import Decimal

@csrf_exempt
def check_po_balance(request):

    if request.method not in ['GET', 'POST']:
        return JsonResponse({'error': 'GET or POST method required'}, status=405)

    try:
        # -------------------------------
        # Parse JSON
        # -------------------------------
        if not request.body:
            return JsonResponse({'error': 'Request body is required'}, status=400)

        data = json.loads(request.body)

        supplier_number = data.get('supplier_number')
        po_items = data.get('po_items', [])

        if not supplier_number or not po_items:
            return JsonResponse(
                {'error': 'supplier_number and po_items are required'},
                status=400
            )

        # -------------------------------
        # Prepare data
        # -------------------------------
        po_numbers = set()
        item_codes = set()
        passed_qty_map = {}

        for item in po_items:
            po = item.get('po_number')
            ic = item.get('item_code')
            qty = item.get('passed_qty', 0)

            if not po or not ic:
                continue

            po_numbers.add(po)
            item_codes.add(ic)

            key = f"{po}_{ic}"
            passed_qty_map[key] = (
                passed_qty_map.get(key, Decimal('0')) +
                Decimal(str(qty))
            )

        po_numbers = list(po_numbers)
        item_codes = list(item_codes)

        po_ph = ','.join(['%s'] * len(po_numbers))
        item_ph = ','.join(['%s'] * len(item_codes))

        # =====================================================
        # 1️⃣ PENDING PO → DEFAULT DB
        # =====================================================
        pending_po_map = {}

        pending_sql = f"""
        SELECT
            VENDOR_NUMBER,
            PO_NUMBER,
            ITEM_CODE,
            SUM(PO_QUANTITY) AS TOTAL_PENDING_QTY
        FROM [BUYP].[dbo].[XXALJEBUYP_PENDING_PO] WITH (NOLOCK)
        WHERE VENDOR_NUMBER = %s
          AND PO_NUMBER IN ({po_ph})
          AND ITEM_CODE IN ({item_ph})
        GROUP BY VENDOR_NUMBER, PO_NUMBER, ITEM_CODE
        """

        with connections['default'].cursor() as cursor:
            params = [supplier_number] + po_numbers + item_codes
            cursor.execute(pending_sql, params)

            for row in cursor.fetchall():
                key = f"{row[1]}_{row[2]}"
                pending_po_map[key] = {
                    'VENDOR_NUMBER': row[0],
                    'PO_NUMBER': row[1],
                    'ITEM_CODE': row[2],
                    'TOTAL_PENDING_QTY': float(row[3] or 0)
                }

        # =====================================================
        # 2️⃣ BAYAN DETAILS → INBOUND_DB
        # =====================================================
        bayan_qty_map = {}

        bayan_sql = f"""
        SELECT
            PO_NUMBER,
            ITEM_CODE,
            SUM(PO_QTY) AS TOTAL_BAYAN_QTY
        FROM [BUYP_INBOUND].[dbo].[WHR_INBOUND_BAYAN_DETAILS_TBL] WITH (NOLOCK)
        WHERE PO_NUMBER IN ({po_ph})
          AND ITEM_CODE IN ({item_ph})
        GROUP BY PO_NUMBER, ITEM_CODE
        """

        with connections['Inbound_db'].cursor() as cursor:
            params = po_numbers + item_codes
            cursor.execute(bayan_sql, params)

            for row in cursor.fetchall():
                key = f"{row[0]}_{row[1]}"
                bayan_qty_map[key] = float(row[2] or 0)

        # =====================================================
        # 3️⃣ MERGE RESULTS
        # =====================================================
        results = []

        all_keys = set(pending_po_map.keys()) | set(passed_qty_map.keys())

        for key in all_keys:
            po, item = key.split('_')

            pending_qty = pending_po_map.get(key, {}).get('TOTAL_PENDING_QTY', 0)
            bayan_qty = bayan_qty_map.get(key, 0)
            balance_qty = pending_qty - bayan_qty
            passed_qty = float(passed_qty_map.get(key, 0))

            if key not in pending_po_map:
                status = 'ERROR'
                message = 'PO / Item not found in system'
            elif passed_qty > balance_qty:
                status = 'WARNING'
                message = (
                    f'Passed quantity ({passed_qty}) exceeds '
                    f'available balance ({balance_qty})'
                )
            else:
                status = 'OK'
                message = f'Available balance: {balance_qty}'

            results.append({
                'VENDOR_NUMBER': supplier_number,
                'PO_NUMBER': po,
                'ITEM_CODE': item,
                'TOTAL_PENDING_QTY': pending_qty,
                'TOTAL_BAYAN_QTY': bayan_qty,
                'BALANCE_QTY': balance_qty,
                'PASSED_QTY': passed_qty,
                'STATUS': status,
                'MESSAGE': message
            })

        return JsonResponse({
            'supplier_number': supplier_number,
            'total_items_checked': len(po_items),
            'results': results
        })

    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON format'}, status=400)

    except Exception as e:
        import traceback
        return JsonResponse({
            'error': str(e),
            'trace': traceback.format_exc()
        }, status=500)
    



class ProductScanCheckView(APIView):
    """
    Optimized existence check for Product Scanning and Damage tables.
    """
    authentication_classes = []
    permission_classes = []
 
    def post(self, request, *args, **kwargs):
        doc_no = request.data.get('DocNo')
        item_code = request.data.get('ItemCode')
        product_code = request.data.get('ProductCode')
        serial_no = request.data.get('SerialNo')
 
        if not all([doc_no, item_code, product_code, serial_no]):
             return Response({'error': 'Missing parameters: DocNo, ItemCode, ProductCode, SerialNo'}, status=status.HTTP_400_BAD_REQUEST)
 
        # Using raw SQL for optimization as requested
        # Note: We repeat parameters for the two parts of the UNION ALL
        sql = """
            IF EXISTS (
                SELECT 1
                FROM [BUYP_INBOUND].[dbo].[Inbound_Receiver_Product_Scaning_tbl]
                WHERE Doc_No      = %s
                  AND Item_Code   = %s
                  AND ProductNo   = %s
                  AND SerailNo    = %s
           
                UNION ALL
           
                SELECT 1
                FROM [BUYP_INBOUND].[dbo].[Inbound_Receiver_Product_Damage_tbl]
                WHERE Doc_No             = %s
                  AND Item_Code          = %s
                  AND Damage_ProductNo   = %s
                  AND Damage_serailNo    = %s
            )
                SELECT 'EXISTS' AS Result;
            ELSE
                SELECT 'NOT EXISTS' AS Result;
        """
       
        params = [doc_no, item_code, product_code, serial_no, doc_no, item_code, product_code, serial_no]
 
        try:
            with connection.cursor() as cursor:
                cursor.execute(sql, params)
                row = cursor.fetchone()
                result = row[0] if row else 'NOT EXISTS'
               
            return Response({'Result': result}, status=status.HTTP_200_OK)
           
        except Exception as e:
             logger.error(f"Error in ProductScanCheckView: {e}", exc_info=True)
             return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        



@csrf_exempt
def get_consolidated_gatepass_data_today(request):
    """
    Consolidated view for GatePass data across 3 tables FILTERED BY Inbound_Initiator_No.
    Logic:
    1. Filter CTE by Inbound_Initiator_No.
    2. Fetch distinct paginated GatePass_Nos.
    3. Fetch all rows for those GatePass_Nos from all 3 tables (Normalize columns).
    4. Aggregate in Python to (Header + Items) structure.
    """
    try:
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 50))
        initiator_no = request.GET.get('Inbound_Initiator_No')

        if not initiator_no:
             return JsonResponse({'status': 'error', 'message': 'Inbound_Initiator_No is required'}, status=400)
        
        if page < 1: page = 1
        if page_size < 1: page_size = 50
        offset_val = (page - 1) * page_size

        with connection.cursor() as cursor:
            # ---------------------------------------------------------
            # STEP 1: Get DISTINCT GatePass_No for Pagination (Filtered by Initiator)
            # ---------------------------------------------------------
            # We construct a union of all GatePass_Nos to find the unique set for this page
            # Added Filter: AND Inbound_Initiator_No = %s
            pagination_sql = """
                WITH AllGatePasses AS (
                    SELECT DISTINCT TRIM(GatePass_No) as GatePass_No FROM [dbo].[Inbound_Receiver_Container_Qc_Tbl] WHERE GatePass_No IS NOT NULL AND Inbound_Initiator_No = %s
                    UNION
                    SELECT DISTINCT TRIM(GatePass_No) as GatePass_No FROM [dbo].[Inbound_Receiver_Product_Scaning_tbl] WHERE GatePass_No IS NOT NULL AND Inbound_Initiator_No = %s
                    UNION
                    SELECT DISTINCT TRIM(GatePass_No) as GatePass_No FROM [dbo].[Inbound_Receiver_Product_Damage_tbl] WHERE GatePass_No IS NOT NULL AND Inbound_Initiator_No = %s
                )
                SELECT DISTINCT agp.GatePass_No 
                FROM AllGatePasses agp
                INNER JOIN [dbo].[WHR_Inbound_Genrate_GatePass_tbl] master
                    ON agp.GatePass_No = TRIM(master.GatePass_No) COLLATE DATABASE_DEFAULT
                WHERE TRIM(master.Status) IN ('Product_Scanned_Finished', 'Container_QC')
                ORDER BY agp.GatePass_No DESC
                OFFSET %s ROWS 
                FETCH NEXT %s ROWS ONLY
            """
            cursor.execute(pagination_sql, [initiator_no, initiator_no, initiator_no, offset_val, page_size])
            gp_rows = cursor.fetchall()
            
            if not gp_rows:
                return JsonResponse({
                    "status": "success",
                    "page": page,
                    "count": 0,
                    "data": []
                })

            # Extract list of GatePass_Nos
            gatepass_list = [row[0] for row in gp_rows]
            
            # Format for SQL IN clause (e.g., 'GP01', 'GP02')
            placeholders = ','.join(['%s'] * len(gatepass_list))
            
            # ---------------------------------------------------------
            # STEP 2: Fetch DETAILED Data for these GatePasses
            # ---------------------------------------------------------
            # We select ALL columns needed for Header + Items.
            # We normalize column aliases so they match across tables where possible,
            # or just select NULL for missing columns.
            
            # Common Columns for Header: 
            # Doc_No, Container_No, GatePass_No, Po_Number, Bayan_No, PO_QTY, 
            # Supplier_Name, Supplier_No, Inbound_Initiator_No, Inbound_Initiator_Name,
            # Receiver_No, Receiver_Name, Gate_Pass_Generated_ByNo, Gate_Pass_Generated_ByName

            # Common Columns for Items:
            # Item_Code, Item_Desc, Prod_Part, PO_QTY, Po_Damage_QTY, Po_Scanned_Qty, 
            # Damage_Image_Count, Damage_Type, Damage_ProductNo, Damage_SerialNo, Status,
            # ProductNo, SerialNo, Image_Count

            details_sql = f"""
                /* Table 1: QC Table */
                SELECT 
                    'QC' as SourceTable,
                    [Doc_No], [Container_No], [GatePass_No], [PO_Number], [Bayan_No], [PO_QTY],
                    [Supplier_Name], [Supplier_No], 
                    [Inbound_Initiator_No], [Inbound_Initiater_Name] as Inbound_Initiator_Name, -- Typos in T1
                    [Inbound_Reciver_No] as Receiver_No, [Inound_Reciver_Name] as Recevicer_Name, -- Typos in T1
                    [GatePassGeneratedBy_No] as Gate_Pass_Generated_ByNo, -- Typos in T1
                    [GatePassGeneartedBy_Name] as Gate_Pass_Generated_ByName, -- Typos in T1
                    
                    /* Item Details */
                    [Item_Code], [Item_Desc], [Prod_Part],
                    [Po_Damage_Qty] as Po_Damage_QTY, 
                    [Po_Scanned_Qty],
                    NULL as Damage_Image_Count, NULL as Damage_Type, NULL as Damage_ProductNo, NULL as Damage_SerialNo,
                    [Status],
                    NULL as ProductNo, NULL as SerialNo,
                    [Image_Count] -- Renamed later to Image_Count (it is Image_Count in DB)
                FROM [BUYP_INBOUND].[dbo].[Inbound_Receiver_Container_Qc_Tbl]
                WHERE TRIM(GatePass_No) IN ({placeholders})

                UNION ALL

                /* Table 2: Scanning Table */
                SELECT 
                    'SCAN' as SourceTable,
                    [Doc_No], [Container_No], [GatePass_No], [Po_Number], [Bayan_No], [PO_QTY],
                    [Supplier_Name], [Supplier_No],
                    [Inbound_Initiator_No], [Inbound_Initiator_Name],
                    [Receiver_No], [Recevicer_Name],
                    [Gate_Pass_Generated_ByNo], [Gate_Pass_Generated_ByName],
                    
                    /* Item Details */
                    [Item_Code], [ItemDesc] as Item_Desc, [Prod_Part],
                    NULL as Po_Damage_QTY,
                    [Po_Scanned_Qty],
                    NULL as Damage_Image_Count, NULL as Damage_Type, NULL as Damage_ProductNo, NULL as Damage_SerialNo,
                    [Status],
                    [ProductNo], [SerailNo] as SerialNo, -- Typos in T2
                    NULL as Image_Count
                FROM [BUYP_INBOUND].[dbo].[Inbound_Receiver_Product_Scaning_tbl]
                WHERE TRIM(GatePass_No) IN ({placeholders})

                UNION ALL

                /* Table 3: Damage Table */
                SELECT 
                    'DAMAGE' as SourceTable,
                    [Doc_No], [Container_No], [GatePass_No], [Po_Number], [Bayan_No], [PO_QTY],
                    [Supplier_Name], [Supplier_No],
                    [Inbound_Initiator_No], [Inbound_Initiator_Name],
                    [Receiver_No], [Recevicer_Name],
                    [Gate_Pass_Generated_ByNo], [Gate_Pass_Generated_ByName],
                    
                    /* Item Details */
                    [Item_Code], [ItemDesc] as Item_Desc, [Prod_Part],
                    [Po_Damage_QTY],
                    [Po_Scanned_Qty],
                    [Damage_Image_Count], [Damage_Type], [Damage_ProductNo], [Damage_serailNo] as Damage_SerialNo, -- Typos in T3
                    [Status],
                    NULL as ProductNo, NULL as SerialNo,
                    NULL as Image_Count
                FROM [BUYP_INBOUND].[dbo].[Inbound_Receiver_Product_Damage_tbl]
                WHERE TRIM(GatePass_No) IN ({placeholders})
            """
            
            # Execute with params repeated 3 times for the 3 UNION parts
            params = gatepass_list * 3
            cursor.execute(details_sql, params)
            
            columns = [col[0] for col in cursor.description]
            all_rows = cursor.fetchall()

        # ---------------------------------------------------------
        # STEP 3: Aggregation in Python
        # ---------------------------------------------------------
        
        # Helpers
        def clean_str(val):
            if isinstance(val, str):
                return val.strip()
            return val
            
        gatepass_headers = {}
        gatepass_items = {} # Map[gp] -> list of items

        # Initialize lists to preserve pagination order
        for gp in gatepass_list:
            gatepass_items[gp] = []
            
        for row in all_rows:
            row_dict = dict(zip(columns, row))
            gp = row_dict.get('GatePass_No')
            
            # 1. Capture Header (First occurrence wins)
            if gp not in gatepass_headers:
                gatepass_headers[gp] = {
                    "Doc_No": clean_str(row_dict.get('Doc_No')),
                    "Container_No": clean_str(row_dict.get('Container_No')),
                    "GatePass_No": clean_str(row_dict.get('GatePass_No')),
                    "Po_Number": clean_str(row_dict.get('PO_Number')),
                    "Bayan_No": clean_str(row_dict.get('Bayan_No')),
                    "PO_QTY": row_dict.get('PO_QTY'),
                    "Supplier_Name": clean_str(row_dict.get('Supplier_Name')),
                    "Supplier_No": clean_str(row_dict.get('Supplier_No')),
                    "Inbound_Initiator_No": clean_str(row_dict.get('Inbound_Initiator_No')),
                    "Inbound_Initiator_Name": clean_str(row_dict.get('Inbound_Initiator_Name')),
                    "Receiver_No": clean_str(row_dict.get('Receiver_No')),
                    "Receiver_Name": clean_str(row_dict.get('Recevicer_Name')),
                    "Gate_Pass_Generated_ByNo": clean_str(row_dict.get('Gate_Pass_Generated_ByNo')),
                    "Gate_Pass_Generated_ByName": clean_str(row_dict.get('Gate_Pass_Generated_ByName')),
                }
                
            # 2. Add Item
            item = {
                "Item_Code": clean_str(row_dict.get('Item_Code')),
                "Item_Desc": clean_str(row_dict.get('Item_Desc')),
                "Prod_Part": clean_str(row_dict.get('Prod_Part')),
                "PO_QTY": row_dict.get('PO_QTY'),
                "Po_Damage_QTY": row_dict.get('Po_Damage_QTY'),
                "Po_Scanned_Qty": row_dict.get('Po_Scanned_Qty'),
                "Damage_Image_Count": row_dict.get('Damage_Image_Count'),
                "Damage_Type": clean_str(row_dict.get('Damage_Type')),
                "Damage_ProductNo": clean_str(row_dict.get('Damage_ProductNo')),
                "Damage_SerialNo": clean_str(row_dict.get('Damage_SerialNo')),
                "Status": clean_str(row_dict.get('Status')),
                "ProductNo": clean_str(row_dict.get('ProductNo')),
                "SerialNo": clean_str(row_dict.get('SerialNo')),
                "Image_Count": row_dict.get('Image_Count'),
                "SourceTable": clean_str(row_dict.get('SourceTable'))
            }
            if gp in gatepass_items:
                gatepass_items[gp].append(item)

        # Construct Final Response (ordered)
        final_response_data = []
        for gp in gatepass_list:
            header = gatepass_headers.get(gp, {})
            # Ensure ItemDetails is the last key added
            header["ItemDetails"] = gatepass_items.get(gp, [])
            final_response_data.append(header)

        return JsonResponse({
            "status": "success",
            "page": page,
            "page_size": page_size,
            "count": len(final_response_data),
            "data": final_response_data
        })

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@api_view(['GET'])
def get_item_tracking_history(request):
    """
    Search-based history tracking with nested data.
    Ensures PRODUCT_CODE is correctly matched and returned.
    """
    item_code = request.query_params.get('item_code', '').strip()
    serial_no = request.query_params.get('serial_no', '').strip()
    product_code = request.query_params.get('product_code', '').strip()
 
    if not any([item_code, serial_no, product_code]):
        return Response({"error": "Provide a search filter"}, status=status.HTTP_400_BAD_REQUEST)
 
    base_query = """
    SELECT DISTINCT
        b.[DOC_NO], b.[BAYAN_NO], b.[PO_NUMBER], b.[ITEM_CODE], b.[ITEM_DESC], b.[CONTAINER_NO], b.[PO_QTY],
        gp.[GatePass_No], gp.[Supplier_Name], gp.[Status] as GP_STATUS,
        sn.[MODEL_NO] as PRODUCT_CODE
    FROM [BUYP_INBOUND].[dbo].[WHR_INBOUND_BAYAN_DETAILS_TBL] b
    -- Join Serial table to get the Model No (Product Code)
    LEFT JOIN [BUYP_INBOUND].[dbo].[WHR_INBOUND_BAYAN_SERIALNO_TBL] sn
        ON TRIM(b.[DOC_NO]) = TRIM(sn.[DOC_NO]) AND TRIM(b.[ITEM_CODE]) = TRIM(sn.[MODEL_NO])
    -- Join GatePass table for status
    LEFT JOIN [BUYP_INBOUND].[dbo].[WHR_Inbound_Genrate_GatePass_tbl] gp
        ON TRIM(b.[DOC_NO]) = TRIM(gp.[Doc_No])
        AND TRIM(b.[CONTAINER_NO]) = TRIM(gp.[Container_No])
        AND TRIM(b.[ITEM_CODE]) = TRIM(gp.[Item_Code])
    WHERE 1=1
    """
   
    params = []
    if item_code:
        base_query += " AND (TRIM(b.[ITEM_CODE]) = %s OR TRIM(sn.[MODEL_NO]) = %s)"
        params.extend([item_code, item_code])
    
    if product_code:
        base_query += """ AND (
            TRIM(b.[ITEM_CODE]) = %s OR 
            TRIM(sn.[MODEL_NO]) = %s OR 
            EXISTS (SELECT 1 FROM [BUYP_INBOUND].[dbo].[Inbound_Receiver_Product_Scaning_tbl] sc 
                   WHERE TRIM(sc.[Doc_No]) = TRIM(b.[DOC_NO]) AND TRIM(sc.[Item_Code]) = TRIM(b.[ITEM_CODE]) 
                   AND (TRIM(sc.[ProductNo]) = %s OR TRIM(sc.[Item_Code]) = %s)) OR
            EXISTS (SELECT 1 FROM [BUYP_INBOUND].[dbo].[Inbound_Receiver_Product_Damage_tbl] dm 
                   WHERE TRIM(dm.[Doc_No]) = TRIM(b.[DOC_NO]) AND TRIM(dm.[Item_Code]) = TRIM(b.[ITEM_CODE]) 
                   AND TRIM(dm.[Damage_ProductNo]) = %s)
        )"""
        params.extend([product_code, product_code, product_code, product_code, product_code])
        
    if serial_no:
        base_query += """ AND (
            TRIM(sn.[SERIAL_NO]) = %s OR 
            EXISTS (SELECT 1 FROM [BUYP_INBOUND].[dbo].[WHR_INBOUND_BAYAN_SERIALNO_TBL] sn2 
                   WHERE TRIM(sn2.[DOC_NO]) = TRIM(b.[DOC_NO]) AND TRIM(sn2.[MODEL_NO]) = TRIM(b.[ITEM_CODE]) AND TRIM(sn2.[SERIAL_NO]) = %s) OR
            EXISTS (SELECT 1 FROM [BUYP_INBOUND].[dbo].[Inbound_Receiver_Product_Scaning_tbl] sc 
                   WHERE TRIM(sc.[Doc_No]) = TRIM(b.[DOC_NO]) AND TRIM(sc.[Item_Code]) = TRIM(b.[ITEM_CODE]) AND TRIM(sc.[SerailNo]) = %s) OR
            EXISTS (SELECT 1 FROM [BUYP_INBOUND].[dbo].[Inbound_Receiver_Product_Damage_tbl] dm 
                   WHERE TRIM(dm.[Doc_No]) = TRIM(b.[DOC_NO]) AND TRIM(dm.[Item_Code]) = TRIM(b.[ITEM_CODE]) AND TRIM(dm.[Damage_serailNo]) = %s)
        )"""
        params.extend([serial_no, serial_no, serial_no, serial_no])
 
    base_query += " ORDER BY b.[DOC_NO] DESC"
 
    try:
        with connection.cursor() as cursor:
            cursor.execute(base_query, params)
            columns = [col[0] for col in cursor.description]
            rows = [dict(zip(columns, r)) for r in cursor.fetchall()]
 
            if rows:
                # Optimization: Fetch Scanned and Damaged details in batches (N+1 query solution)
                unique_pairs = list({(r['DOC_NO'].strip() if r['DOC_NO'] else '', r['ITEM_CODE'].strip() if r['ITEM_CODE'] else '') for r in rows})
                scans_map = {}
                damages_map = {}
 
                # Fetch details in chunks to respect MS SQL param limits (max 2100)
                chunk_size = 1000  # 1000 pairs = 2000 parameters
                for i in range(0, len(unique_pairs), chunk_size):
                    chunk = unique_pairs[i:i+chunk_size]
 
                    # Create VALUES clause, apply type cast on the first tuple to avoid type-guessing issues
                    values_clause_parts = []
                    for idx in range(len(chunk)):
                        if idx == 0:
                            values_clause_parts.append("(CAST(%s AS NVARCHAR(4000)), CAST(%s AS NVARCHAR(4000)))")
                        else:
                            values_clause_parts.append("(%s, %s)")
 
                    values_clause = ", ".join(values_clause_parts)
                    flat_params = [val for pair in chunk for val in pair]
 
                    scan_query = f"""
                        SELECT TRIM(tbl.[Doc_No]) as DOC_NO, TRIM(tbl.[Item_Code]) as ITEM_CODE, tbl.[SerailNo], tbl.[ProductNo], tbl.[Status] 
                        FROM [BUYP_INBOUND].[dbo].[Inbound_Receiver_Product_Scaning_tbl] tbl
                        INNER JOIN (VALUES {values_clause}) AS temp(Doc_No, Item_Code)
                        ON TRIM(tbl.[Doc_No]) = temp.Doc_No AND TRIM(tbl.[Item_Code]) = temp.Item_Code
                    """
                    cursor.execute(scan_query, flat_params)
                    scan_cols = [c[0] for c in cursor.description]
                    for s_row in cursor.fetchall():
                        s_dict = dict(zip(scan_cols, s_row))
                        key = (s_dict['DOC_NO'], s_dict['ITEM_CODE'])
                        scans_map.setdefault(key, []).append({
                            'SerailNo': s_dict['SerailNo'],
                            'ProductNo': s_dict['ProductNo'],
                            'Status': s_dict['Status']
                        })
 
                    dmg_query = f"""
                        SELECT TRIM(tbl.[Doc_No]) as DOC_NO, TRIM(tbl.[Item_Code]) as ITEM_CODE, tbl.[Damage_serailNo], tbl.[Damage_Type], tbl.[Damage_ProductNo], tbl.[Status] 
                        FROM [BUYP_INBOUND].[dbo].[Inbound_Receiver_Product_Damage_tbl] tbl
                        INNER JOIN (VALUES {values_clause}) AS temp(Doc_No, Item_Code)
                        ON TRIM(tbl.[Doc_No]) = temp.Doc_No AND TRIM(tbl.[Item_Code]) = temp.Item_Code
                    """
                    cursor.execute(dmg_query, flat_params)
                    dmg_cols = [c[0] for c in cursor.description]
                    for d_row in cursor.fetchall():
                        d_dict = dict(zip(dmg_cols, d_row))
                        key = (d_dict['DOC_NO'], d_dict['ITEM_CODE'])
                        damages_map.setdefault(key, []).append({
                            'Damage_serailNo': d_dict['Damage_serailNo'],
                            'Damage_Type': d_dict['Damage_Type'],
                            'Damage_ProductNo': d_dict['Damage_ProductNo'],
                            'Status': d_dict['Status']
                        })
 
            for row in rows:
                doc_no = row['DOC_NO'].strip() if row['DOC_NO'] else ''
                item_code = row['ITEM_CODE'].strip() if row['ITEM_CODE'] else ''
                key = (doc_no, item_code)
 
                all_scanned = scans_map.get(key, []) if rows else []
                all_damaged = damages_map.get(key, []) if rows else []

                # If searching by serial_no, only return the matching serial's details
                if serial_no:
                    row['SCANNED_DETAILS'] = [
                        s for s in all_scanned
                        if (s.get('SerailNo') or '').strip() == serial_no
                    ]
                    row['DAMAGED_DETAILS'] = [
                        d for d in all_damaged
                        if (d.get('Damage_serailNo') or '').strip() == serial_no
                    ]
                else:
                    row['SCANNED_DETAILS'] = all_scanned
                    row['DAMAGED_DETAILS'] = all_damaged

                # Try to fill PRODUCT_CODE from scanned details if it's currently NULL (join failed)
                if not row.get('PRODUCT_CODE'):
                    if row['SCANNED_DETAILS']:
                        row['PRODUCT_CODE'] = row['SCANNED_DETAILS'][0].get('ProductNo')
                    elif row['DAMAGED_DETAILS']:
                        row['PRODUCT_CODE'] = row['DAMAGED_DETAILS'][0].get('Damage_ProductNo')

                # Trim status for comparison
                gp_status = (row['GP_STATUS'] or "").strip()
 
                # Status Flags for Timeline (Expected by Flutter: QC_STATUS, SCANNING_STATUS)
                row['QC_STATUS'] = "FINISHED" if gp_status in ['Container_QC', 'Product_Scanned_Finished'] else "PENDING"
                row['SCANNING_STATUS'] = "FINISHED" if gp_status == 'Product_Scanned_Finished' else "PENDING"
 
                # Summary Text
                if gp_status == 'Product_Scanned_Finished':
                    row['SUMMARY'] = "Process 100% Completed"
                elif gp_status == 'Container_QC':
                    row['SUMMARY'] = "QC Done, Scanning Pending"
                else:
                    row['SUMMARY'] = gp_status or "In Transit"
 
        return Response({"success": True, "data": rows}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



@csrf_exempt
def create_vehicle_type(request):
    """
    POST API to insert Vechicle_Type and Measurmement into [BUYP_INBOUND].[dbo].[WHR_Vehicle_Type_tbl]
    using raw SQL queries.
    """
    if request.method == 'POST':
        try:
            body = json.loads(request.body)
           
            # Normalize to a list to support both single inserts and bulk inserts
            if isinstance(body, dict):
                body = [body]
               
            if not isinstance(body, list):
                return JsonResponse({
                    "status": "error",
                    "message": "Request body must be a JSON object or a list of JSON objects."
                }, status=400)
 
            if not body:
                return JsonResponse({
                    "status": "error",
                    "message": "The request body cannot be empty."
                }, status=400)
 
            insert_data = []
            for index, item in enumerate(body):
                # Using the original key names expected from JSON (can be mapped directly to DB columns later)
                vehicle_type = item.get('Vehicle_Type') or item.get('Vechicle_Type')
                measurement = item.get('Measurement') or item.get('Measurmement')
               
                # Validation: both fields are required
                if not vehicle_type or not measurement:
                    return JsonResponse({
                        "status": "error",
                        "message": f"Item at index {index} is missing 'Vechicle_Type' or 'Measurmement' (also accepts 'Vehicle_Type' and 'Measurement')."
                    }, status=400)
                   
                insert_data.append([vehicle_type, measurement])
 
            # Execute bulk insert using raw SQL
            with connections['Inbound_db'].cursor() as cursor:
                insert_query = """
                    INSERT INTO [BUYP_INBOUND].[dbo].[WHR_Vehicle_Type_tbl] ([Vechicle_Type], [Measurmement])
                    VALUES (%s, %s)
                """
                cursor.executemany(insert_query, insert_data)
 
            return JsonResponse({
                "status": "success",
                "message": f"Successfully inserted {len(insert_data)} record(s)."
            }, status=201)
 
        except json.JSONDecodeError:
            return JsonResponse({
                "status": "error",
                "message": "Invalid JSON format in the request body."
            }, status=400)
           
        except Exception as e:
            return JsonResponse({
                "status": "error",
                "message": f"A database error occurred: {str(e)}"
            }, status=500)
           
    return JsonResponse({
        "status": "error",
        "message": "Only POST requests are allowed."
    }, status=405)


@csrf_exempt
def get_vehicle_types(request):
    """
    GET API to retrieve all Vechicle_Type and Measurmement records from [BUYP_INBOUND].[dbo].[WHR_Vehicle_Type_tbl]
    using raw SQL queries, optimized for large datasets.
    """
    if request.method == 'GET':
        try:
            with connections['Inbound_db'].cursor() as cursor:
                # Optimized selective query checking specified DB + columns
                select_query = """
                    SELECT TOP (1000) [Id], [Vechicle_Type], [Measurmement]
                    FROM [BUYP_INBOUND].[dbo].[WHR_Vehicle_Type_tbl]
                    ORDER BY [Id] DESC
                """
                cursor.execute(select_query)
               
                # Fetchall automatically pulls all matching records
                rows = cursor.fetchall()
               
                # Zip columns dynamically for faster dict mapping
                columns = [col[0] for col in cursor.description]
                data = [dict(zip(columns, row)) for row in rows]
 
            return JsonResponse({
                "status": "success",
                "count": len(data),
                "data": data
            }, status=200)
 
        except Exception as e:
            return JsonResponse({
                "status": "error",
                "message": f"A database error occurred: {str(e)}"
            }, status=500)
 
    return JsonResponse({
        "status": "error",
        "message": "Only GET requests are allowed."
    }, status=405)